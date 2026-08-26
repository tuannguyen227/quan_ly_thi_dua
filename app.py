import sys
import os
import json
import pandas as pd
from datetime import datetime
import io
import re

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

# Import các thành phần từ thư mục con của dự án
from database.database import init_db, get_session, session_scope
from database.models import User, UserRole, SchoolYear, Branch, RedStar, DutyArea, Assignment, StarEvaluation, WeeklyScore, ViolationCategory, WeeklyViolation, RawScore, MonthlyRecord, ActionLog, ScoreSettings

# Import hàm kiểm tra đăng nhập từ file account_manager
from database.account_manager import verify_external_login, sync_account_to_json, remove_account_from_json, load_external_accounts, save_external_accounts

# Khởi tạo ứng dụng Web Flask
app = Flask(__name__)
# Thiết lập khóa bí mật để sử dụng Flash messages báo lỗi
app.secret_key = "doan_truong_thanh_hoa_secret_key" 

from flask import g

# ==========================================
# --- HÀM HỖ TRỢ: GHI NHẬT KÝ THAO TÁC HỆ THỐNG ---
# ==========================================
def log_system_action(action_type, details=""):
    """
    Hàm lưu vết thao tác vào bộ nhớ đệm (Flask g). 
    Sẽ được ghi vào database SAU KHI xử lý xong luồng chính để tránh lỗi "database is locked".
    """
    if 'action_logs' not in g:
        g.action_logs = []
        
    g.action_logs.append({
        'username': session.get('username', 'Khách'),
        'full_name': session.get('full_name', 'Người dùng Ẩn danh'),
        'action_type': action_type,
        'details': details,
        'timestamp': datetime.utcnow()
    })

@app.after_request
def save_action_logs(response):
    """
    Tự động kích hoạt sau mỗi request: Đổ dữ liệu log từ bộ nhớ đệm vào Database.
    Vì lúc này Database đã được luồng chính giải phóng, nên sẽ KHÔNG BAO GIỜ bị lỗi Locked.
    """
    if hasattr(g, 'action_logs') and g.action_logs:
        try:
            from database.database import session_scope
            from database.models import ActionLog
            
            with session_scope() as db_session:
                for log_data in g.action_logs:
                    new_log = ActionLog(
                        username=log_data['username'],
                        full_name=log_data['full_name'],
                        action_type=log_data['action_type'],
                        details=log_data['details'],
                        timestamp=log_data['timestamp']
                    )
                    db_session.add(new_log)
                db_session.commit()
        except Exception as e:
            print(f"Lỗi lưu file log: {e}")
    return response

# ==========================================
# CÁC HÀM TỰ ĐỘNG KHỞI TẠO HỆ THỐNG
# ==========================================
def auto_init_accounts():
    """Hàm tự động kiểm tra và tạo lại file accounts.json nếu bị mất"""
    data_folder = "data"
    os.makedirs(data_folder, exist_ok=True)
    
    file_path = os.path.join(data_folder, "accounts.json")
    
    if not os.path.exists(file_path):
        default_accounts = [
            {
                "username": "admin",
                "password": "1",  
                "full_name": "Bí thư Đoàn trường",
                "role": "Bí thư"
            }
        ]
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(default_accounts, f, indent=4, ensure_ascii=False)
            print("Hệ thống: Đã tự động khôi phục file accounts.json với tài khoản mặc định.")
        except Exception as e:
            print(f"Hệ thống: Lỗi không thể tạo file accounts.json - {e}")

def create_mock_admin():
    """Tạo một tài khoản Bí thư Đoàn trường mặc định trong SQLite để test"""
    try:
        with session_scope() as db_session:
            admin = db_session.query(User).filter(User.username == "admin").first()
            if not admin:
                new_admin = User(
                    username="admin",
                    password_hash="1",  
                    full_name="Bí thư Đoàn trường",
                    role=UserRole.BI_THU,
                    is_active=True
                )
                db_session.add(new_admin)
                print("Đã khởi tạo tài khoản thành công: User: admin | Pass: 1")
    except Exception as e:
        print(f"Lỗi khởi tạo tài khoản admin: {e}")

@app.before_request
def restrict_access():
    # 1. [BẢN VÁ LỖI PWA]: SỬ DỤNG .path THAY VÌ .endpoint ĐỂ CHỐNG REDIRECT CHO APP MOBILE
    if request.path in ['/sw.js', '/manifest.json'] or request.path.startswith('/static/'):
        return None
        
    # 2. KHAI BÁO DANH SÁCH QUYỀN TRUY CẬP CHO TỪNG ROLE
    allowed_for_gvcn = [
        'login', 'logout', 'ping', 'change_password', 
        'class_dashboard', 'preview_class_dashboard', 'export_class_dashboard', 'submit_appeal',
        'parse_sodaubai', 
        'api_submit_evaluation',
        'api_gvcn_leaderboard',
        'api_gvcn_get_months',
        'update_branch_info'
    ]
    
    allowed_for_saodo = [
        'login', 'logout', 'ping', 'change_password', 
        'mobile_sao_do', '', 
        'quick_log_violation_form', 'handle_raw_scores',
        'sao_do_quick_submit_form', 'submit_mobile_sao_do',
        'api_submit_evaluation'
    ]

    # [MỚI NHẤT]: DANH SÁCH ĐƯỜNG DẪN DÀNH RIÊNG CHO BGH (CHỈ XEM BÁO CÁO & ĐIỀU HÀNH)
    allowed_for_bgh = [
        'login', 'logout', 'ping', 'change_password', 
        'bgh_dashboard',                 
        'class_dashboard',               
        'preview_class_dashboard',       
        'export_class_dashboard',        
        'class_monthly_analysis',        
        'class_semester_analysis',       
        'school_monthly_analysis'        
    ]
    
    # 3. KIỂM TRA QUYỀN TRUY CẬP THEO ROLE
    role = session.get('role')
    
    # Kiểm tra quyền Giáo viên chủ nhiệm
    if role == 'Giáo viên chủ nhiệm':
        if request.endpoint and request.endpoint not in allowed_for_gvcn:
            if request.path.startswith('/api/'):
                return {
                    "success": False,
                    "error": "Bạn không có quyền truy cập API này.",
                    "endpoint": request.endpoint
                }, 403

            flash("⛔ Từ chối truy cập: Quyền GVCN!", "error")
            return redirect(url_for('class_dashboard'))

    # Kiểm tra quyền Ban Giám hiệu (Chỉ cho phép xem báo cáo và điều hành)
    elif role == 'Ban Giám hiệu':
        if request.endpoint and request.endpoint not in allowed_for_bgh:
            if request.path.startswith('/api/'):
                return {
                    "success": False,
                    "error": "Ban Giám hiệu chỉ có quyền xem báo cáo và điều hành.",
                    "endpoint": request.endpoint
                }, 403

            flash("⛔ Từ chối truy cập: Tài khoản BGH chỉ có quyền xem báo cáo và điều hành!", "error")
            return redirect(url_for('bgh_dashboard'))

@app.route('/ping')
def ping():
    return "<h1>Kết nối thành công! Máy chủ đang hoạt động.</h1>"

# =====================================================================
# API: BÓC TÁCH DỮ LIỆU TỪ FILE SỔ ĐẦU BÀI (TỐI ƯU CHỐNG SÓT ĐIỂM)
@app.route('/api/parse_sodaubai', methods=['POST'])
@app.route('/weekly/api/parse_sodaubai', methods=['POST'])
def parse_sodaubai():
    if 'excel_file' not in request.files:
        return {"error": "Không tìm thấy file!"}, 400
        
    file = request.files['excel_file']
    if file.filename == '':
        return {"error": "Chưa chọn file nào!"}, 400
        
    # Lấy tên lớp dự kiến từ giao diện để đối chiếu
    expected_branch = request.form.get('expected_branch_name', '').strip().upper()
        
    try:
        import pandas as pd
        import re
        
        df = pd.read_excel(file, header=None)
        
        # =================================================================
        # THUẬT TOÁN KHIÊN BẢO VỆ: CHỐNG UPLOAD NHẦM FILE LỚP KHÁC
        # =================================================================
        if expected_branch:
            found_class_name = None
            # Quét tối đa 50 dòng đầu và toàn bộ cột để tìm chữ "Lớp: ..."
            for i in range(min(50, len(df))):
                for j in range(len(df.columns)):
                    cell_val = str(df.iloc[i, j]).strip()
                    if cell_val.lower().startswith('lớp:'):
                        match = re.search(r'Lớp:\s*([A-Za-z0-9]+)', cell_val, re.IGNORECASE)
                        if match:
                            found_class_name = match.group(1).upper()
                            break
                if found_class_name:
                    break
            
            # Khóa nòng: Trùng thì đi tiếp, Không trùng thì từ chối ngay!
            if found_class_name and found_class_name != expected_branch:
                return {
                    "error": f"⛔ CẢNH BÁO: FILE SỔ ĐẦU BÀI KHÔNG KHỚP!\nBạn đang ở form nhập điểm của lớp {expected_branch}, nhưng file Excel bạn vừa tải lên lại là Sổ đầu bài của lớp {found_class_name}. Vui lòng chọn lại đúng file!"
                }, 400
        # =================================================================
        
        try:
            start_row = df[df[0].astype(str).str.contains('Thứ \nngày tháng', na=False, case=False)].index[0]
        except:
            return {"error": "Hệ thống không nhận diện được biểu mẫu Sổ Đầu Bài này!"}, 400
            
        c10 = c9 = c8 = 0
        subject_scores = {}  # Phân loại điểm tốt (8, 9, 10) theo Tên Môn Học
        diem_kem_list = []   # Danh sách điểm kém (1, 2) theo định dạng tên học sinh
        diem_khong_list = [] # Danh sách học sinh bị điểm 0 (Không học bài)
        
        # Bắt đầu quét từ start_row + 3 (bỏ qua dòng tiêu đề và hàng số thứ tự 1-10)
        for i in range(start_row + 3, len(df)):
            if i >= len(df): break
            
            cot0_text = str(df.iloc[i, 0])
            if "Ý kiến nhận xét" in cot0_text or "Tổng số tiết" in cot0_text: break
                
            # Lấy tên môn học (Cột 3 theo biểu mẫu Sổ đầu bài)
            try: mon = str(df.iloc[i, 3]).strip()
            except: mon = "Khác"
            if not mon or mon.lower() == 'nan': mon = "Khác"
                
            # Gộp dữ liệu từ các cột chứa điểm (Cột 13, 14, 15, 18)
            row_scores = []
            for col in [13, 14, 15, 18]:
                if col < len(df.columns):
                    val = str(df.iloc[i, col]).strip()
                    if val.lower() != 'nan': row_scores.append(val)
                    
            diem_raw = " ".join(row_scores)
            if not diem_raw or 'Ý kiến' in diem_raw or 'BAN GIÁM' in diem_raw:
                continue
                
            # THUẬT TOÁN MỚI: Tách theo dấu phẩy/chấm phẩy, trích xuất điểm bất chấp có nhận xét kèm theo phía sau
            entries = re.split(r'[,;]+', diem_raw)
            parsed_any = False
            
            for entry in entries:
                entry = entry.strip()
                if not entry: continue
                
                # Tìm cặp [Tên học sinh] và [Con số điểm 0-10] ở bất kỳ vị trí nào trong đoạn phân tách
                match = re.search(r'([A-ZÀ-Ỹa-zà-ỹ\s]+?)\s*[:\-]?\s*\b(10|[0-9])\b', entry)
                if match:
                    parsed_any = True
                    raw_name = match.group(1).strip()
                    # Lấy từ cuối cùng hoặc 2 từ cuối làm tên học sinh nếu chuỗi tên quá dài do dính chữ
                    name_words = raw_name.split()
                    name_part = name_words[-1].title() if name_words else "Học sinh"
                    score_val = int(match.group(2))
                    
                    tiet = str(df.iloc[i, 2]).strip()
                    tiet_str = f"Tiết {tiet}" if tiet != 'nan' else "Tiết học"
                    
                    if score_val == 10:
                        c10 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c10'] += 1
                    elif score_val == 9:
                        c9 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c9'] += 1
                    elif score_val == 8:
                        c8 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c8'] += 1
                    elif score_val == 0:
                        if name_part: diem_khong_list.append(f"{name_part} ({tiet_str} môn {mon})")
                        else: diem_khong_list.append(f"{tiet_str} môn {mon}")
                    elif score_val in [1, 2]:
                        if name_part: diem_kem_list.append(f"{name_part} - {score_val} điểm ({tiet_str} môn {mon})")
                        else: diem_kem_list.append(f"{tiet_str} môn {mon} ({score_val} điểm)")
            
            # THUẬT TOÁN DỰ PHÒNG: Nếu không tách được theo tên, quét toàn bộ số nguyên hợp lệ trong ô
            if not parsed_any:
                numbers = re.findall(r'\b(10|9|8|0|[1-2])\b', diem_raw)
                for num_str in numbers:
                    num = int(num_str)
                    tiet = str(df.iloc[i, 2]).strip()
                    tiet_str = f"Tiết {tiet}" if tiet != 'nan' else "Tiết học"
                    
                    if num == 10:
                        c10 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c10'] += 1
                    elif num == 9:
                        c9 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c9'] += 1
                    elif num == 8:
                        c8 += 1
                        if mon not in subject_scores: subject_scores[mon] = {'c10': 0, 'c9': 0, 'c8': 0}
                        subject_scores[mon]['c8'] += 1
                    elif num == 0:
                        diem_khong_list.append(f"{tiet_str} môn {mon}")
                    elif num in [1, 2]:
                        diem_kem_list.append(f"{tiet_str} môn {mon} ({num} điểm)")

        # THUẬT TOÁN XẾP LOẠI TUẦN THEO QUY CHẾ CỦA TRƯỜNG
        xep_loai = "Bình thường"
        summary_row = df[df[0].astype(str).str.contains('Tổng số giờ xếp loại', na=False, case=False)]
        if not summary_row.empty:
            sum_text = str(summary_row.iloc[0, 0])
            match_tot = re.search(r'\[\s*(\d+)\s*Tốt', sum_text, re.IGNORECASE)
            match_kha = re.search(r';\s*(\d+)\s*Khá', sum_text, re.IGNORECASE)
            match_tb  = re.search(r';\s*(\d+)\s*(?:Đạt|Trung bình|TB)', sum_text, re.IGNORECASE)
            match_yeu = re.search(r';\s*(\d+)\s*Yếu', sum_text, re.IGNORECASE)
            
            so_tot = int(match_tot.group(1)) if match_tot else 0
            so_kha = int(match_kha.group(1)) if match_kha else 0
            so_tb  = int(match_tb.group(1)) if match_tb else 0
            so_yeu = int(match_yeu.group(1)) if match_yeu else 0
            
            tong_tiet = so_tot + so_kha + so_tb + so_yeu
            
            if tong_tiet > 0:
                if so_tot == tong_tiet:
                    xep_loai = "Tuần Tốt"
                elif so_yeu == 0 and so_tb == 0 and so_kha > 0:
                    xep_loai = "Tuần Khá"

        # ĐÓNG GÓI LỖI VÀO Ô GHI CHÚ (SỔ ĐEN)
        note_fragments = []
        if diem_khong_list:
            for item in set(diem_khong_list):
                note_fragments.append(f"Không học bài x1 [{item}]")
        if diem_kem_list:
            for item in set(diem_kem_list):
                note_fragments.append(f"Bị điểm kém x1 [{item}]")

        # ĐÓNG GÓI MẢNG ĐIỂM MÔN HỌC CHI TIẾT
        raw_list = []
        for subj, pts in subject_scores.items():
            raw_list.append({"subj": subj, "c10": pts['c10'], "c9": pts['c9'], "c8": pts['c8']})

        return {
            "success": True,
            "c10": c10, "c9": c9, "c8": c8,
            "xep_loai": xep_loai,
            "note": " ; ".join(note_fragments),
            "raw_list": raw_list
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"error": f"Lỗi xử lý file Sổ Đầu Bài: {str(e)}"}, 500

@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_data = verify_external_login(username, password)

        if user_data: 
            session['username'] = user_data.get('username', username)
            session['role'] = user_data.get('role', 'Quản trị viên')
            session['full_name'] = user_data.get('full_name', 'Người dùng')
            
            log_system_action("ĐĂNG NHẬP", f"Tài khoản {session['username']} ({session['full_name']}) đã truy cập hệ thống.")
            
            # [NÂNG CẤP LÕI]: Bẻ lái tự động theo phân quyền
            if session['role'] == "Giáo viên chủ nhiệm":
                return redirect(url_for('class_dashboard'))
            elif session['role'] == "Sao đỏ":
                return redirect(url_for('mobile_sao_do')) # Bắn thẳng vào Web App Mobile cho học sinh
            elif session['role'] == "Ban Giám hiệu":
                return redirect(url_for('bgh_dashboard')) # [MỚI]: Chuyển thẳng BGH vào Trung tâm điều hành
            else:
                return redirect(url_for('dashboard')) # Admin / Bí thư vào trang tổng quan
        else:
            flash("Tài khoản hoặc mật khẩu không chính xác!", "error")
            return redirect(url_for('login'))
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    log_system_action("ĐĂNG XUẤT", f"Tài khoản {session.get('username', '')} đã rời khỏi hệ thống.")
    session.clear() 
    return redirect(url_for('login'))

# [NÂNG CẤP LÕI]: API XỬ LÝ PHÚC KHẢO DÀNH CHO BCH / ADMIN
@app.route('/resolve_appeal', methods=['POST'])
def resolve_appeal():
    # Chống GVCN can thiệp
    if session.get('role') == 'Giáo viên chủ nhiệm':
        return redirect(url_for('login'))
        
    score_id = request.form.get('score_id', type=int)
    response_text = request.form.get('response_text', '').strip()
    
    if not score_id or not response_text:
        flash("Vui lòng nhập nội dung phản hồi!", "error")
        return redirect(request.referrer or url_for('dashboard'))
        
    try:
        with session_scope() as db_session:
            score = db_session.query(WeeklyScore).filter_by(id=score_id).first()
            if score:
                score.appeal_response = response_text
                log_system_action("XỬ LÝ PHÚC KHẢO", f"Đã phản hồi khiếu nại của lớp {score.branch.name} Tuần {score.week}.")
                flash(f"✅ Đã gửi phản hồi thành công cho lớp {score.branch.name}!", "success")
    except Exception as e:
        flash(f"Lỗi: {e}", "error")
        
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    try:
        with session_scope() as db_session:
            from sqlalchemy import func
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            total_branches = 0
            current_week = "Chưa có"
            avg_score = 0.0
            chart_labels = []
            chart_data = []
            pending_appeals_data = []
            
            # [TÍNH NĂNG MỚI]: BẢNG BÁO ĐỘNG HỌC SINH CÁ BIỆT TOÀN TRƯỜNG
            global_warnings = []

            if active_year:
                total_branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).count()
                latest_score = db_session.query(WeeklyScore).join(Branch).filter(Branch.school_year_id == active_year.id).order_by(WeeklyScore.id.desc()).first()
                
                if latest_score:
                    current_week = latest_score.week
                    avg_val = db_session.query(func.avg(WeeklyScore.total_score)).join(Branch).filter(
                        Branch.school_year_id == active_year.id,
                        WeeklyScore.week == current_week
                    ).scalar()
                    
                    try: avg_score = round(float(avg_val), 1) if avg_val is not None else 0.0
                    except: avg_score = 0.0

                    top_10 = db_session.query(Branch.name, WeeklyScore.total_score).join(WeeklyScore).filter(
                        Branch.school_year_id == active_year.id,
                        WeeklyScore.week == current_week
                    ).order_by(WeeklyScore.total_score.desc()).limit(10).all()

                    chart_labels = [item[0] for item in top_10]
                    chart_data = [round(float(item[1]), 1) if item[1] is not None else 0.0 for item in top_10]

                    # --- THUẬT TOÁN QUÉT SỔ ĐEN TOÀN TRƯỜNG (CHỈ XÉT TUẦN HIỆN TẠI) ---
                    all_scores_current_week = db_session.query(WeeklyScore).join(Branch).filter(
                        WeeklyScore.week == current_week,
                        Branch.school_year_id == active_year.id
                    ).all()
                    
                    for score in all_scores_current_week:
                        student_viol_counts = {}
                        for viol in score.violations:
                            if viol.student_name and str(viol.student_name).strip() != "":
                                raw_names = str(viol.student_name).replace(';', ',').split(',')
                                names = [n.strip().upper() for n in raw_names if n.strip()]
                                for name in names:
                                    qty = int(viol.quantity) if viol.quantity else 1
                                    student_viol_counts[name] = student_viol_counts.get(name, 0) + qty
                        
                        # Nếu ai >= 3 lỗi, ném ngay ra bảng phong thần
                        for name, count in student_viol_counts.items():
                            if count >= 3:
                                global_warnings.append({
                                    'branch_name': score.branch.name,
                                    'student_name': name.title(),
                                    'count': count
                                })
                    
                    # Xếp người vi phạm nhiều nhất lên đầu
                    global_warnings.sort(key=lambda x: x['count'], reverse=True)

                if session.get('role') != 'Giáo viên chủ nhiệm':
                    appeals = db_session.query(WeeklyScore).join(Branch).filter(
                        Branch.school_year_id == active_year.id,
                        WeeklyScore.is_appealed == True,
                        (WeeklyScore.appeal_response == None) | (WeeklyScore.appeal_response == "")
                    ).all()
                    for p in appeals:
                        pending_appeals_data.append({
                            'score_id': p.id, 'branch_name': p.branch.name,
                            'week': p.week, 'reason': p.appeal_reason
                        })

            return render_template(
                'dashboard.html',
                user_fullname=session.get('full_name', "Người dùng"),
                user_role=session.get('role', "Quản trị viên"),
                total_branches=total_branches,
                current_week=current_week,
                avg_score=avg_score,
                chart_labels=chart_labels,
                chart_data=chart_data,
                pending_appeals=pending_appeals_data,
                global_warnings=global_warnings # Truyền biến báo động ra giao diện
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải bảng điều khiển: {e}", "error")
        return redirect(url_for('login'))
    
# ==========================================
# MODULE: ĐỔI MẬT KHẨU TRỰC TUYẾN (ÉP ĐỔI)
# ==========================================
@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if 'username' not in session:
        flash("Vui lòng đăng nhập trước khi đổi mật khẩu!", "warning")
        return redirect(url_for('login'))

    if request.method == 'POST':
        login_id = request.form.get('login_id', '').strip()
        new_pwd = request.form.get('new_password', '').strip()
        confirm_pwd = request.form.get('confirm_password', '').strip()

        if not login_id or not new_pwd or not confirm_pwd:
            flash("Vui lòng nhập đầy đủ thông tin Tên đăng nhập và Mật khẩu mới!", "error")
            return redirect(url_for('change_password'))

        if len(new_pwd) < 3:
            flash("Mật khẩu mới phải có ít nhất 3 ký tự!", "error")
            return redirect(url_for('change_password'))

        if new_pwd != confirm_pwd:
            flash("Mật khẩu xác nhận không khớp!", "error")
            return redirect(url_for('change_password'))

        try:
            accounts = load_external_accounts()
            target_acc = None
            for acc in accounts:
                if acc.get("username") == login_id:
                    target_acc = acc
                    break

            if not target_acc:
                flash(f"Không tìm thấy tài khoản nào có Tên đăng nhập là '{login_id}'!", "error")
                return redirect(url_for('change_password'))

            target_acc["password"] = new_pwd
            save_external_accounts(accounts)

            with session_scope() as session_db:
                db_user = session_db.query(User).filter(User.username == login_id).first()
                if db_user:
                    db_user.password_hash = new_pwd

            log_system_action("BẢO MẬT", f"Đã ép đổi mật khẩu cho tài khoản: {login_id}")

            flash(f"✅ Đã thay đổi mật khẩu thành công cho tài khoản: {login_id}! Vui lòng đăng nhập lại.", "success")
            session.clear() 
            return redirect(url_for('login'))

        except Exception as e:
            flash(f"Lỗi hệ thống khi đổi mật khẩu: {str(e)}", "error")
            return redirect(url_for('change_password'))

    return render_template('change_password.html', default_username=session.get('username', ''))

# ==========================================
# MODULE: XEM NHẬT KÝ HỆ THỐNG (CHỈ DÀNH CHO ADMIN)
# ==========================================
@app.route('/action_logs')
def action_logs():
    user_role = session.get('role', '')
    
    if user_role not in ['Quản trị viên', 'Admin', 'Bí thư', 'Bí thư Đoàn trường']:
        flash("Bạn không có quyền xem Nhật ký hệ thống!", "error")
        return redirect(url_for('dashboard'))
        
    try:
        import datetime as dt
        with session_scope() as db_session:
            logs = db_session.query(ActionLog).order_by(ActionLog.timestamp.desc()).limit(500).all()
            
            logs_data = []
            for log in logs:
                local_time = log.timestamp + dt.timedelta(hours=7) if log.timestamp else dt.datetime.now()
                logs_data.append({
                    'time': local_time.strftime("%d/%m/%Y - %H:%M:%S"),
                    'username': log.username,
                    'full_name': log.full_name,
                    'action_type': log.action_type,
                    'details': log.details
                })
                
        return render_template('action_logs.html', logs=logs_data)
    except Exception as e:
        flash(f"Lỗi tải nhật ký: {e}", "error")
        return redirect(url_for('dashboard'))

# ==========================================
# CẤU HÌNH HỆ THỐNG: QUẢN LÝ NĂM HỌC
# ==========================================
@app.route('/school-years', methods=['GET'])
def school_years():
    try:
        with session_scope() as db_session:
            years = db_session.query(SchoolYear).order_by(SchoolYear.id.desc()).all()
            return render_template('school_years.html', years=years)
    except Exception as e:
        flash(f"Lỗi tải danh sách năm học: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/add_school_year', methods=['POST'])
def add_school_year():
    name = request.form.get('name', '').strip()
    if not name:
        flash("Vui lòng điền tên năm học!", "error")
        return redirect(url_for('school_years'))

    try:
        with session_scope() as db_session:
            is_first = db_session.query(SchoolYear).count() == 0
            new_year = SchoolYear(name=name, is_active=is_first)
            db_session.add(new_year)
            log_system_action("CẤU HÌNH", f"Thêm năm học mới: {name}")
        flash(f"Đã thêm năm học mới: {name}", "success")
    except Exception as e:
        flash(f"Lỗi khi thêm năm học: {e}", "error")
    return redirect(url_for('school_years'))

@app.route('/set_active_year/<int:id>', methods=['POST'])
def set_active_year(id):
    try:
        with session_scope() as db_session:
            db_session.query(SchoolYear).update({SchoolYear.is_active: False})
            selected_year = db_session.query(SchoolYear).filter_by(id=id).first()
            if selected_year:
                selected_year.is_active = True
                log_system_action("CẤU HÌNH", f"Kích hoạt năm học: {selected_year.name}")
                flash(f"Đã chuyển sang năm học: {selected_year.name}", "success")
    except Exception as e:
        flash(f"Lỗi kích hoạt năm học: {e}", "error")
    return redirect(url_for('school_years'))

@app.route('/delete_school_year/<int:id>', methods=['POST'])
def delete_school_year(id):
    try:
        with session_scope() as db_session:
            year = db_session.query(SchoolYear).filter_by(id=id).first()
            if year:
                if year.is_active:
                    flash("Không thể xóa năm học đang được kích hoạt!", "error")
                else:
                    name = year.name
                    db_session.delete(year)
                    log_system_action("CẤU HÌNH", f"Xóa năm học: {name}")
                    flash("Đã xóa năm học thành công!", "success")
    except Exception as e:
        flash(f"Lỗi khi xóa năm học: {e}", "error")
    return redirect(url_for('school_years'))


# ==========================================
# CẤU HÌNH HỆ THỐNG: QUẢN LÝ NGƯỜI DÙNG
# ==========================================
@app.route('/users', methods=['GET'])
def users():
    try:
        with session_scope() as db_session:
            user_list = db_session.query(User).order_by(User.id).all()
            return render_template('users.html', users=user_list)
    except Exception as e:
        flash(f"Lỗi tải danh sách người dùng: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/add_user', methods=['POST'])
def add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    full_name = request.form.get('full_name', '').strip()
    role_text = request.form.get('role', '').strip()

    if not username or not full_name or not password:
        flash("Vui lòng điền đầy đủ Tên đăng nhập, Họ tên và Mật khẩu!", "error")
        return redirect(url_for('users'))

    try:
        with session_scope() as db_session:
            # [NÂNG CẤP]: Khớp nối chính xác quyền GVCN vào CSDL
            role_enum = UserRole.BCH
            if "Quản trị" in role_text or "Admin" in role_text: role_enum = UserRole.ADMIN
            elif "Bí thư" in role_text: role_enum = UserRole.BI_THU
            elif "Giáo viên chủ nhiệm" in role_text: role_enum = UserRole.GVCN

            exist = db_session.query(User).filter_by(username=username).first()
            if exist:
                flash(f"Tên đăng nhập '{username}' đã tồn tại trên hệ thống!", "error")
            else:
                new_user = User(username=username, password_hash=password, full_name=full_name, role=role_enum, is_active=True)
                db_session.add(new_user)
                
                try: sync_account_to_json(username, full_name, password, role_text, True)
                except Exception as e: print(f"Lỗi đồng bộ JSON: {e}")
                
                log_system_action("TẠO TÀI KHOẢN", f"Đã cấp tài khoản mới: {username} (Họ tên: {full_name}, Quyền: {role_text})")
                flash(f"Đã cấp tài khoản thành công cho: {full_name}", "success")
    except Exception as e:
        flash(f"Lỗi khi thêm tài khoản: {e}", "error")
    return redirect(url_for('users'))

@app.route('/edit_user/<int:id>', methods=['POST'])
def edit_user(id):
    new_username = request.form.get('edit_username', '').strip()
    new_fullname = request.form.get('edit_fullname', '').strip()
    new_password = request.form.get('edit_password', '').strip()
    new_role_text = request.form.get('edit_role', '').strip()

    if not new_username or not new_fullname:
        flash("Tên đăng nhập và Họ tên không được để trống!", "error")
        return redirect(url_for('users'))

    try:
        with session_scope() as db_session:
            user = db_session.query(User).filter_by(id=id).first()
            if not user:
                flash("Không tìm thấy tài khoản này!", "error")
                return redirect(url_for('users'))

            old_username = user.username
            changes_made = False
            change_details = []

            if new_username != old_username:
                exist = db_session.query(User).filter_by(username=new_username).first()
                if exist:
                    flash(f"Tên đăng nhập '{new_username}' đã có người sử dụng!", "error")
                    return redirect(url_for('users'))
                user.username = new_username
                changes_made = True
                change_details.append(f"Tên ĐN: '{old_username}' -> '{new_username}'")

            if new_fullname != user.full_name:
                change_details.append(f"Họ tên: '{user.full_name}' -> '{new_fullname}'")
                user.full_name = new_fullname
                changes_made = True

            # [NÂNG CẤP]: Khớp nối chính xác quyền GVCN vào CSDL
            new_role_enum = UserRole.BCH
            if "Quản trị" in new_role_text or "Admin" in new_role_text: new_role_enum = UserRole.ADMIN
            elif "Bí thư" in new_role_text: new_role_enum = UserRole.BI_THU
            elif "Giáo viên chủ nhiệm" in new_role_text: new_role_enum = UserRole.GVCN
                
            if user.role != new_role_enum:
                change_details.append(f"Quyền: '{user.role.value}' -> '{new_role_text}'")
                user.role = new_role_enum
                changes_made = True

            if new_password:
                if len(new_password) < 3:
                    flash("Mật khẩu mới phải có ít nhất 3 ký tự!", "error")
                    return redirect(url_for('users'))
                user.password_hash = new_password
                changes_made = True
                change_details.append("Đã đổi mật khẩu")

            if changes_made:
                try:
                    if new_username != old_username: remove_account_from_json(old_username)
                    sync_account_to_json(user.username, user.full_name, user.password_hash, new_role_text, user.is_active)
                except: pass

                log_system_action("CHỈNH SỬA TÀI KHOẢN", f"Cập nhật '{old_username}': " + "; ".join(change_details))
                flash(f"Đã cập nhật thông tin tài khoản '{new_username}'!", "success")
            else:
                flash("Chưa có thông tin nào được thay đổi.", "info")

    except Exception as e:
        flash(f"Lỗi cập nhật tài khoản: {e}", "error")
    return redirect(url_for('users'))
# ==========================================
# API: THAO TÁC HÀNG LOẠT VỚI TÀI KHOẢN (XÓA / KHÓA)
# ==========================================
@app.route('/bulk_user_action', methods=['POST'])
def bulk_user_action():
    # Chỉ Admin/Bí thư mới được làm điều này
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Bạn không có quyền thực hiện chức năng này!", "error")
        return redirect(url_for('users'))

    action = request.form.get('action') # Nhận lệnh: 'delete' (xóa) hoặc 'toggle_lock' (khóa/mở)
    user_ids = request.form.getlist('user_ids') # Lấy danh sách các ID được tích chọn

    if not user_ids:
        flash("Vui lòng tích chọn ít nhất 1 tài khoản để thao tác!", "warning")
        return redirect(url_for('users'))

    try:
        with session_scope() as db_session:
            count = 0
            for uid in user_ids:
                user = db_session.query(User).filter_by(id=int(uid)).first()
                if not user:
                    continue

                # CƠ CHẾ BẢO VỆ LÕI: Tuyệt đối không cho chạm vào tài khoản admin tối cao
                if user.username.lower() == 'admin':
                    continue

                if action == 'delete':
                    name = user.username
                    db_session.delete(user)
                    try: remove_account_from_json(name)
                    except: pass
                    count += 1
                    
                elif action == 'toggle_lock':
                    user.is_active = not user.is_active
                    role_text = user.role.value if hasattr(user.role, 'value') else str(user.role)
                    try: sync_account_to_json(user.username, user.full_name, user.password_hash, role_text, user.is_active)
                    except: pass
                    count += 1

            # Phản hồi theo từng hành động
            if action == 'delete':
                log_system_action("XÓA TÀI KHOẢN", f"Đã xóa hàng loạt {count} tài khoản khỏi hệ thống.")
                flash(f"✅ Đã xóa vĩnh viễn {count} tài khoản thành công!", "success")
            elif action == 'toggle_lock':
                log_system_action("THAY ĐỔI TRẠNG THÁI", f"Đã thay đổi trạng thái hàng loạt {count} tài khoản.")
                flash(f"✅ Đã Khóa / Mở khóa {count} tài khoản thành công!", "success")

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi thao tác hàng loạt: {str(e)}", "error")

    return redirect(url_for('users'))

@app.route('/toggle_user/<int:id>', methods=['POST'])
def toggle_user(id):
    try:
        with session_scope() as db_session:
            user = db_session.query(User).filter_by(id=id).first()
            if user:
                if user.username.lower() == 'admin':
                    flash("Không thể khóa tài khoản quản trị viên tối cao!", "error")
                else:
                    user.is_active = not user.is_active
                    status = "Mở khóa" if user.is_active else "Khóa"
                    
                    try: sync_account_to_json(user.username, user.full_name, user.password_hash, user.role.value, user.is_active)
                    except: pass
                    
                    log_system_action("THAY ĐỔI TRẠNG THÁI", f"Đã {status} tài khoản: {user.username}")
                    flash(f"Đã {status} tài khoản: {user.username}", "success")
    except Exception as e:
        flash(f"Lỗi thay đổi trạng thái user: {e}", "error")
    return redirect(url_for('users'))

@app.route('/delete_user/<int:id>', methods=['POST'])
def delete_user(id):
    try:
        with session_scope() as db_session:
            user = db_session.query(User).filter_by(id=id).first()
            if user:
                if user.username.lower() == 'admin':
                    flash("Hệ thống từ chối xóa tài khoản quản trị tối cao!", "error")
                else:
                    name = user.username
                    db_session.delete(user)
                    
                    try: remove_account_from_json(name)
                    except: pass
                    
                    log_system_action("XÓA TÀI KHOẢN", f"Đã xóa vĩnh viễn tài khoản: {name}")
                    flash(f"Đã xóa vĩnh viễn tài khoản: {name}", "success")
    except Exception as e:
        flash(f"Lỗi xóa tài khoản: {e}", "error")
    return redirect(url_for('users'))

# ==========================================
# MODULE: QUẢN LÝ CHI ĐOÀN
# ==========================================
@app.route('/branches', methods=['GET', 'POST'])
def branches():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            if request.method == 'POST':
                if not active_year:
                    flash("Chưa có năm học nào được kích hoạt!", "error")
                    return redirect(url_for('branches'))
                    
                name = request.form.get('name', '').strip().upper()
                group = request.form.get('group', 'Nhóm 1')
                si_so = request.form.get('si_so', 0)
                gvcn = request.form.get('gvcn', '').strip()
                phone_gvcn = request.form.get('phone_gvcn', '').strip()
                class_monitor = request.form.get('class_monitor', '').strip()
                phone_monitor = request.form.get('phone_monitor', '').strip()
                
                try: si_so = int(si_so) if si_so else 0
                except ValueError: si_so = 0

                if name:
                    exist = db_session.query(Branch).filter(Branch.name == name, Branch.school_year_id == active_year.id).first()
                    if exist:
                        flash(f"Chi đoàn {name} đã tồn tại trong năm học này!", "error")
                    else:
                        new_branch = Branch(
                            name=name, group=group, si_so=si_so, gvcn=gvcn, 
                            phone_gvcn=phone_gvcn, class_monitor=class_monitor, phone_monitor=phone_monitor,
                            school_year_id=active_year.id
                        )
                        db_session.add(new_branch)
                        log_system_action("QUẢN LÝ CHI ĐOÀN", f"Đã thêm Chi đoàn {name}")
                        flash(f"Đã thêm Chi đoàn {name} thành công!", "success")
                else:
                    flash("Vui lòng nhập tên Chi đoàn!", "error")
                    
                return redirect(url_for('branches'))

            branch_list = []
            if active_year:
                branch_list = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                
            return render_template('branches.html', branches=branch_list, active_year=active_year)
    except Exception as e:
        flash(f"Lỗi phân hệ Chi đoàn: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/delete_branch/<int:id>', methods=['POST'])
def delete_branch(id):
    try:
        with session_scope() as db_session:
            branch = db_session.query(Branch).filter(Branch.id == id).first()
            if branch:
                name = branch.name
                db_session.delete(branch)
                log_system_action("QUẢN LÝ CHI ĐOÀN", f"Đã xóa Chi đoàn {name}")
                flash(f"Đã xóa Chi đoàn {name} thành công!", "success")
    except Exception as e:
        flash(f"Lỗi xóa Chi đoàn: {e}", "error")
    return redirect(url_for('branches'))

@app.route('/edit_branch/<int:id>', methods=['POST'])
def edit_branch(id):
    try:
        with session_scope() as db_session:
            branch = db_session.query(Branch).filter(Branch.id == id).first()
            if branch:
                branch.name = request.form.get('edit_name', branch.name).strip().upper()
                branch.group = request.form.get('edit_group', branch.group)
                branch.gvcn = request.form.get('edit_gvcn', branch.gvcn).strip()
                branch.phone_gvcn = request.form.get('edit_phone_gvcn', branch.phone_gvcn).strip()
                branch.class_monitor = request.form.get('edit_class_monitor', branch.class_monitor).strip()
                branch.phone_monitor = request.form.get('edit_phone_monitor', branch.phone_monitor).strip()
                
                si_so = request.form.get('edit_si_so', branch.si_so)
                try: branch.si_so = int(si_so) if si_so else 0
                except ValueError: branch.si_so = 0
                
                log_system_action("QUẢN LÝ CHI ĐOÀN", f"Đã cập nhật thông tin Chi đoàn {branch.name}")
                flash(f"Đã cập nhật thông tin Chi đoàn {branch.name}!", "success")
    except Exception as e:
        flash(f"Lỗi cập nhật Chi đoàn: {e}", "error")
    return redirect(url_for('branches'))

@app.route('/import_branches', methods=['POST'])
def import_branches():
    if 'excel_file' not in request.files:
        flash("Không tìm thấy file tải lên!", "error")
        return redirect(url_for('branches'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash("Chưa chọn file nào!", "error")
        return redirect(url_for('branches'))
        
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            df = pd.read_excel(file)
            with session_scope() as db_session:
                active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
                if not active_year:
                    flash("Chưa có năm học nào được kích hoạt!", "error")
                    return redirect(url_for('branches'))
                    
                count_success = 0
                for index, row in df.iterrows():
                    name = str(row.get('Tên Lớp', '')).strip().upper()
                    if not name or name == 'NAN': continue
                    
                    group = str(row.get('Nhóm', 'Nhóm 1')).strip()
                    gvcn = str(row.get('GVCN', '')).strip()
                    phone_gvcn = str(row.get('SĐT GVCN', '')).strip()
                    class_monitor = str(row.get('Lớp trưởng', '')).strip()
                    phone_monitor = str(row.get('SĐT Lớp trưởng', '')).strip()
                    
                    if gvcn.lower() == 'nan': gvcn = ''
                    if phone_gvcn.lower() == 'nan': phone_gvcn = ''
                    if class_monitor.lower() == 'nan': class_monitor = ''
                    if phone_monitor.lower() == 'nan': phone_monitor = ''
                    
                    try: si_so = int(row.get('Sĩ số', 0))
                    except: si_so = 0
                        
                    exist = db_session.query(Branch).filter(Branch.name == name, Branch.school_year_id == active_year.id).first()
                    if not exist:
                        new_branch = Branch(
                            name=name, group=group, si_so=si_so, gvcn=gvcn, 
                            phone_gvcn=phone_gvcn, class_monitor=class_monitor, phone_monitor=phone_monitor,
                            school_year_id=active_year.id
                        )
                        db_session.add(new_branch)
                        count_success += 1
                        
                log_system_action("QUẢN LÝ CHI ĐOÀN", f"Đã nhập thành công {count_success} Chi đoàn từ file Excel")
                flash(f"Đã nhập thành công {count_success} Chi đoàn từ file Excel!", "success")
        except Exception as e:
            flash(f"Lỗi đọc file Excel: {str(e)}", "error")
    else:
        flash("Vui lòng chọn file Excel hợp lệ (.xlsx, .xls)", "error")
        
    return redirect(url_for('branches'))


# ==========================================
# MODULE: QUẢN LÝ ĐỘI SAO ĐỎ
# ==========================================
@app.route('/red-stars', methods=['GET', 'POST'])
def red_stars():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            if request.method == 'POST':
                if not active_year:
                    flash("Chưa có năm học nào được kích hoạt!", "error")
                    return redirect(url_for('red_stars'))
                    
                full_name = request.form.get('full_name', '').strip()
                gender = request.form.get('gender', 'Nam')
                phone = request.form.get('phone', '').strip()
                branch_id = request.form.get('branch_id')
                notes = request.form.get('notes', '').strip()
                
                if full_name and branch_id:
                    new_star = RedStar(
                        full_name=full_name, gender=gender, phone=phone,
                        branch_id=branch_id, notes=notes, is_active=True
                    )
                    db_session.add(new_star)
                    log_system_action("ĐỘI SAO ĐỎ", f"Đã thêm Sao đỏ {full_name}")
                    flash(f"Đã thêm Sao đỏ {full_name} thành công!", "success")
                else:
                    flash("Vui lòng nhập tên và chọn Chi đoàn!", "error")
                return redirect(url_for('red_stars'))

            stars_list = []
            branches_list = []
            search_name = request.args.get('search_name', '').strip()
            filter_branch = request.args.get('filter_branch', '')

            if active_year:
                branches_list = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                branch_ids = [b.id for b in branches_list]
                
                if branch_ids:
                    query = db_session.query(RedStar).filter(RedStar.branch_id.in_(branch_ids))
                    if search_name:
                        query = query.filter(RedStar.full_name.ilike(f"%{search_name}%"))
                    if filter_branch:
                        query = query.filter(RedStar.branch_id == int(filter_branch))
                    stars_list = query.all()
                    
            return render_template(
                'red_stars.html', stars=stars_list, branches=branches_list, 
                active_year=active_year, search_name=search_name, filter_branch=filter_branch
            )
    except Exception as e:
        flash(f"Lỗi phân hệ Sao đỏ: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/toggle_star_status/<int:id>', methods=['POST'])
def toggle_star_status(id):
    try:
        with session_scope() as db_session:
            star = db_session.query(RedStar).filter(RedStar.id == id).first()
            if star:
                star.is_active = not star.is_active
                status_label = "Đang hoạt động" if star.is_active else "Tạm nghỉ"
                log_system_action("ĐỘI SAO ĐỎ", f"Đã chuyển trạng thái của {star.full_name} sang: {status_label}")
                flash(f"Đã chuyển trạng thái của {star.full_name} sang: {status_label}", "success")
    except Exception as e:
        flash(f"Lỗi đổi trạng thái sao đỏ: {e}", "error")
    return redirect(url_for('red_stars'))

@app.route('/edit_red_star/<int:id>', methods=['POST'])
def edit_red_star(id):
    try:
        with session_scope() as db_session:
            star = db_session.query(RedStar).filter(RedStar.id == id).first()
            if star:
                star.full_name = request.form.get('edit_full_name', star.full_name).strip()
                star.gender = request.form.get('edit_gender', star.gender)
                star.phone = request.form.get('edit_phone', star.phone).strip()
                star.notes = request.form.get('edit_notes', star.notes).strip()
                
                branch_id = request.form.get('edit_branch_id')
                if branch_id: star.branch_id = branch_id
                
                is_active = request.form.get('edit_is_active')
                star.is_active = True if is_active == 'on' else False
                
                log_system_action("ĐỘI SAO ĐỎ", f"Đã cập nhật thông tin Sao đỏ {star.full_name}")
                flash(f"Đã cập nhật thông tin Sao đỏ {star.full_name}!", "success")
    except Exception as e:
        flash(f"Lỗi sửa thông tin sao đỏ: {e}", "error")
    return redirect(url_for('red_stars'))

@app.route('/delete_red_star/<int:id>', methods=['POST'])
def delete_red_star(id):
    try:
        with session_scope() as db_session:
            star = db_session.query(RedStar).filter(RedStar.id == id).first()
            if star:
                name = star.full_name
                db_session.query(Assignment).filter(Assignment.red_star_id == id).delete()
                db_session.query(StarEvaluation).filter(StarEvaluation.red_star_id == id).delete()
                db_session.delete(star)
                log_system_action("ĐỘI SAO ĐỎ", f"Đã xóa vĩnh viễn Sao đỏ {name} và các lịch trực liên quan")
                flash(f"Đã xóa vĩnh viễn Sao đỏ {name} và các lịch trực liên quan!", "success")
    except Exception as e:
        flash(f"Lỗi xóa sao đỏ: {e}", "error")
    return redirect(url_for('red_stars'))

@app.route('/import_red_stars', methods=['POST'])
def import_red_stars():
    if 'excel_file' not in request.files:
        flash("Không tìm thấy file tải lên!", "error")
        return redirect(url_for('red_stars'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash("Chưa chọn file nào!", "error")
        return redirect(url_for('red_stars'))
        
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip().str.upper()
            
            col_name, col_phone, col_branch, col_gender, col_notes = None, None, None, None, None
            for col in df.columns:
                if col in ["HỌ VÀ TÊN", "HỌ TÊN", "TÊN"]: col_name = col
                if col in ["GIỚI TÍNH", "NAM/NỮ", "PHÁI"]: col_gender = col
                if col in ["LỚP", "CHI ĐOÀN"]: col_branch = col
                if col in ["SỐ ĐIỆN THOẠI", "SDT", "PHONE", "ĐIỆN THOẠI"]: col_phone = col
                if col in ["GHI CHÚ", "NOTE", "GHI CHU"]: col_notes = col
                    
            if not col_name or not col_branch:
                flash("Lỗi cấu trúc: File Excel bắt buộc phải có cột 'Họ và tên' và 'Chi đoàn/Lớp'.", "error")
                return redirect(url_for('red_stars'))
                
            with session_scope() as db_session:
                active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
                if not active_year:
                    flash("Chưa có năm học nào được kích hoạt!", "error")
                    return redirect(url_for('red_stars'))
                    
                count_new, count_exist = 0, 0
                for _, row in df.iterrows():
                    if pd.isna(row[col_name]) or pd.isna(row[col_branch]): continue
                        
                    fullname = str(row[col_name]).strip()
                    branch_name = str(row[col_branch]).strip().upper()
                    phone_num = str(row[col_phone]).strip() if col_phone and not pd.isna(row[col_phone]) else ""
                    gender_val = str(row[col_gender]).strip().capitalize() if col_gender and not pd.isna(row[col_gender]) else "Nam"
                    notes_raw = row[col_notes] if col_notes else None
                    notes_val = str(notes_raw).strip() if pd.notna(notes_raw) and str(notes_raw).strip().lower() != 'nan' else ""
                    
                    branch = db_session.query(Branch).filter(Branch.name == branch_name, Branch.school_year_id == active_year.id).first()
                    if branch:
                        exist_rs = db_session.query(RedStar).filter(RedStar.full_name == fullname, RedStar.branch_id == branch.id).first()
                        if not exist_rs:
                            new_rs = RedStar(
                                full_name=fullname, gender=gender_val, phone=phone_num, 
                                branch_id=branch.id, notes=notes_val, is_active=True
                            )
                            db_session.add(new_rs)
                            count_new += 1
                        else: count_exist += 1
                            
                log_system_action("ĐỘI SAO ĐỎ", f"Đã nạp thành công {count_new} Sao đỏ mới từ Excel")
                flash(f"Đã nạp thành công {count_new} Sao đỏ mới (Bỏ qua {count_exist} học sinh bị trùng)", "success")
        except Exception as e:
            flash(f"Lỗi đọc file Excel: {str(e)}", "error")
    else:
        flash("Vui lòng chọn file Excel hợp lệ (.xlsx, .xls)", "error")
        
    return redirect(url_for('red_stars'))
# --- API LẤY LỊCH SỬ TRỰC CỦA TỪNG SAO ĐỎ ---
@app.route('/api/red_star_history/<int:star_id>')
def api_red_star_history(star_id):
    try:
        from database.database import session_scope
        from database.models import Assignment, DutyArea, RedStar
        import json
        import os
        
        with session_scope() as db_session:
            # 1. Kiểm tra sao đỏ
            star = db_session.query(RedStar).filter_by(id=star_id).first()
            if not star:
                return {"error": "Không tìm thấy thông tin Sao đỏ này!"}
                
            # 2. Đọc Sơ đồ Cụm trực để dịch từ Tên Khu Vực ra Tên Lớp
            zones_map = {}
            if os.path.exists("config/class_zones.json"):
                with open("config/class_zones.json", "r", encoding="utf-8") as f:
                    zones_map = json.load(f)
                    
            # 3. Quét toàn bộ lịch sử phân công
            history_data = []
            assignments = db_session.query(Assignment).filter(
                Assignment.red_star_id == star_id
            ).order_by(Assignment.week_number.desc(), Assignment.shift).all()
            
            for a in assignments:
                # Lấy tên khu vực
                area_name = a.duty_area.name if a.duty_area else "Không rõ"
                
                # Dịch ra các lớp phải trực
                classes = zones_map.get(area_name, [])
                class_str = ", ".join(classes) if classes else "Khu vực chung / Cổng"
                
                history_data.append({
                    "week": f"Tuần {a.week_number}",
                    "shift": a.shift,
                    "area": area_name,
                    "classes": class_str
                })
                
            return {"success": True, "data": history_data}
    except Exception as e:
        return {"error": str(e)}
    
# --- 1. QUẢN LÝ KHU VỰC TRỰC ---
@app.route('/duty-areas', methods=['GET', 'POST'])
def duty_areas():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            if request.method == 'POST':
                name = request.form.get('name', '').strip()
                try: req_stars = int(request.form.get('required_stars', 2))
                except: req_stars = 2
                
                selected_classes = request.form.getlist('branch_names')
                
                if name:
                    exist = db_session.query(DutyArea).filter_by(name=name).first()
                    if not exist:
                        new_area = DutyArea(name=name, required_stars=req_stars)
                        db_session.add(new_area)
                    else:
                        exist.required_stars = req_stars
                    
                    if selected_classes:
                        zones_map = {}
                        if os.path.exists("config/class_zones.json"):
                            with open("config/class_zones.json", "r", encoding="utf-8") as f:
                                zones_map = json.load(f)
                        
                        zones_map[name] = selected_classes
                        os.makedirs("config", exist_ok=True)
                        with open("config/class_zones.json", "w", encoding="utf-8") as f:
                            json.dump(zones_map, f, ensure_ascii=False, indent=4)
                            
                    log_system_action("CỤM TRỰC", f"Đã cập nhật thành công Khu vực/Cụm trực: {name}")
                    flash(f"Đã cập nhật thành công Khu vực/Cụm trực: {name}", "success")
                else:
                    flash("Vui lòng nhập tên khu vực!", "error")
                return redirect(url_for('duty_areas'))
                
            areas = db_session.query(DutyArea).all()
            branches = []
            if active_year:
                branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
            
            zones_map = {}
            assigned_classes = set() 
            if os.path.exists("config/class_zones.json"):
                with open("config/class_zones.json", "r", encoding="utf-8") as f:
                    zones_map = json.load(f)
                    for classes_in_zone in zones_map.values():
                        assigned_classes.update(classes_in_zone)
                    
            return render_template('duty_areas.html', areas=areas, zones_map=zones_map, branches=branches, assigned_classes=assigned_classes)
    except Exception as e:
        flash(f"Lỗi phân hệ khu vực trực: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/delete_duty_area/<int:id>', methods=['POST'])
def delete_duty_area(id):
    try:
        with session_scope() as db_session:
            area = db_session.query(DutyArea).filter(DutyArea.id == id).first()
            if area:
                zone_name = area.name
                
                # [VÁ LỖI TRIỆT ĐỂ]: Duyệt qua các phân công và xóa thủ công dựa trên ID object liên kết
                all_assignments = db_session.query(Assignment).all()
                for assign in all_assignments:
                    if assign.duty_area and assign.duty_area.id == id:
                        db_session.delete(assign)
                        
                db_session.delete(area)
                
                import os, json
                if os.path.exists("config/class_zones.json"):
                    with open("config/class_zones.json", "r", encoding="utf-8") as f:
                        zones_map = json.load(f)
                    if zone_name in zones_map:
                        del zones_map[zone_name]
                        with open("config/class_zones.json", "w", encoding="utf-8") as f:
                            json.dump(zones_map, f, ensure_ascii=False, indent=4)
                            
                log_system_action("CỤM TRỰC", f"Đã xóa khu vực trực: {zone_name}")
                flash("Đã xóa khu vực trực và các lịch trực liên quan thành công!", "success")
    except Exception as e:
        flash(f"Lỗi xóa khu vực: {e}", "error")
    return redirect(url_for('duty_areas'))
@app.route('/edit_duty_area/<int:id>', methods=['POST'])
def edit_duty_area(id):
    try:
        with session_scope() as db_session:
            area = db_session.query(DutyArea).filter(DutyArea.id == id).first()
            if area:
                old_name = area.name
                new_name = request.form.get('edit_name', old_name).strip()
                try: new_req_stars = int(request.form.get('edit_required_stars', area.required_stars))
                except: new_req_stars = 2
                
                selected_classes = request.form.getlist('edit_branch_names')
                
                # Kiểm tra nếu đổi tên mà tên mới bị trùng
                if new_name != old_name:
                    exist = db_session.query(DutyArea).filter_by(name=new_name).first()
                    if exist:
                        flash(f"Tên khu vực '{new_name}' đã tồn tại!", "error")
                        return redirect(url_for('duty_areas'))
                        
                # Cập nhật Database
                area.name = new_name
                area.required_stars = new_req_stars
                
                # Cập nhật Sơ đồ lớp (File JSON)
                import os, json
                config_path = "config/class_zones.json"
                zones_map = {}
                if os.path.exists(config_path):
                    with open(config_path, "r", encoding="utf-8") as f:
                        try: zones_map = json.load(f)
                        except: pass
                        
                # Nếu đổi tên Cụm, xóa tên cũ trong JSON
                if old_name != new_name and old_name in zones_map:
                    del zones_map[old_name]
                    
                # Cập nhật danh sách lớp mới
                zones_map[new_name] = selected_classes
                
                os.makedirs("config", exist_ok=True)
                with open(config_path, "w", encoding="utf-8") as f:
                    json.dump(zones_map, f, ensure_ascii=False, indent=4)
                    
                log_system_action("CỤM TRỰC", f"Đã cập nhật khu vực: {old_name} -> {new_name}")
                flash(f"Đã cập nhật thông tin Cụm trực {new_name} thành công!", "success")
    except Exception as e:
        flash(f"Lỗi cập nhật khu vực trực: {e}", "error")
    return redirect(url_for('duty_areas'))

@app.route('/import_class_zones', methods=['POST'])
def import_class_zones():
    if 'excel_file' not in request.files:
        flash("Không tìm thấy file tải lên!", "error")
        return redirect(url_for('duty_areas'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash("Chưa chọn file nào!", "error")
        return redirect(url_for('duty_areas'))
        
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            import numpy as np
            
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip().str.upper()
            
            if 'LỚP' not in df.columns:
                flash("Lỗi cấu trúc: File Excel bắt buộc phải có cột tiêu đề là 'LỚP'.", "error")
                return redirect(url_for('duty_areas'))
                
            dynamic_zones = {}
            cum_col = None
            for col in df.columns:
                if "CỤM" in col or "CUM" in col:
                    cum_col = col
                    break
            
            if cum_col:
                df[cum_col] = df[cum_col].replace('', np.nan).ffill()
                for index, row in df.iterrows():
                    zone_name = str(row[cum_col]).strip()
                    class_name = str(row['LỚP']).strip()
                    
                    if pd.notna(row['LỚP']) and class_name.lower() != 'nan' and class_name != '':
                        if zone_name.lower() != 'nan' and zone_name != '':
                            if zone_name not in dynamic_zones: 
                                dynamic_zones[zone_name] = []
                            dynamic_zones[zone_name].append(class_name)
            else:
                class_list = df['LỚP'].dropna().astype(str).tolist()
                grades = {}
                for c in class_list:
                    c_clean = c.strip()
                    if not c_clean or c_clean.lower() == 'nan': continue
                    grade = c_clean[:2]
                    if grade not in grades: grades[grade] = []
                    grades[grade].append(c_clean)
                for grade, classes in grades.items():
                    for i in range(0, len(classes), 2):
                        zone_name = f"CỤM_{grade}_{(i//2)+1:02d}"
                        dynamic_zones[zone_name] = classes[i:i+2]
            
            os.makedirs("config", exist_ok=True)
            with open("config/class_zones.json", "w", encoding="utf-8") as f:
                json.dump(dynamic_zones, f, ensure_ascii=False, indent=4)
                
            with session_scope() as db_session:
                count_new = 0
                for zone_name in dynamic_zones.keys():
                    exist = db_session.query(DutyArea).filter_by(name=zone_name).first()
                    if not exist:
                        new_area = DutyArea(name=zone_name, required_stars=2)
                        db_session.add(new_area)
                        count_new += 1
                
            log_system_action("CỤM TRỰC", f"Đã nhập Sơ đồ lớp từ Excel, tạo {count_new} khu vực")
            flash(f"✅ Đã xử lý thành công file Sơ đồ lớp! Tạo {len(dynamic_zones)} Cụm trực và thêm mới {count_new} khu vực vào hệ thống.", "success")
        except Exception as e:
            flash(f"Lỗi đọc file Excel: {str(e)}", "error")
    else:
        flash("Vui lòng chọn file Excel hợp lệ (.xlsx, .xls)", "error")
    return redirect(url_for('duty_areas'))

# --- 2. PHÂN CÔNG LỊCH TRỰC ---
@app.route('/assignments', methods=['GET', 'POST'])
def assignments():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            # --- [NÂNG CẤP LÕI]: TỰ ĐỘNG NHẬN DIỆN TUẦN TRỰC MỚI NHẤT ---
            week_param = request.args.get('week')
            if week_param:
                current_week = int(week_param)
            else:
                # Quét CSDL tìm tuần mới nhất vừa được xếp lịch
                latest_assign = db_session.query(Assignment).order_by(Assignment.week_number.desc()).first()
                current_week = latest_assign.week_number if latest_assign else 1
            # -------------------------------------------------------------
            
            areas = db_session.query(DutyArea).all()
            active_stars = db_session.query(RedStar).filter(RedStar.is_active == True).all() 
            
            assign_list = []
            start_date_str = "" # Biến chứa ngày áp dụng để hiển thị lên Form
            
            if active_year:
                assign_list = db_session.query(Assignment).join(RedStar).join(Branch).filter(
                    Assignment.week_number == current_week,
                    Branch.school_year_id == active_year.id
                ).order_by(Assignment.shift, Assignment.id).all()
                
                # Trích xuất "Ngày áp dụng" từ bản ghi đầu tiên của Tuần này
                if assign_list and assign_list[0].date:
                    start_date_str = assign_list[0].date.strftime('%Y-%m-%d')
            
            return render_template('assignments.html', 
                                   active_year=active_year, 
                                   areas=areas, 
                                   stars=active_stars, 
                                   assign_list=assign_list, 
                                   current_week=current_week,
                                   start_date=start_date_str) # Truyền ngày ra màn hình HTML
    except Exception as e:
        flash(f"Lỗi tải lịch phân công: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/delete_assignment/<int:id>', methods=['POST'])
def delete_assignment(id):
    try:
        with session_scope() as db_session:
            assign = db_session.query(Assignment).filter(Assignment.id == id).first()
            week = assign.week_number if assign else 1
            if assign:
                db_session.delete(assign)
                log_system_action("LỊCH TRỰC", f"Đã thu hồi 1 phân công trong Tuần {week}")
                flash("Đã thu hồi lịch phân công!", "success")
        return redirect(url_for('assignments', week=week))
    except Exception as e:
        flash(f"Lỗi thu hồi lịch: {e}", "error")
        return redirect(url_for('assignments'))

@app.route('/auto_assign', methods=['POST'])
def auto_assign():
    week_number = int(request.form.get('auto_week_number', 1))
    date_str = request.form.get('auto_date')
    try: start_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.now().date()
    except: start_date = datetime.now().date()
    
    shifts = request.form.getlist('shifts')
    if not shifts:
        flash("Vui lòng chọn ít nhất 1 ca trực!", "error")
        return redirect(url_for('assignments', week=week_number))
        
    try:
        with session_scope() as db_session:
            import random
            import json
            import os
            import re
            
            # 1. Xóa toàn bộ lịch cũ của tuần này để xếp lại từ đầu
            db_session.query(Assignment).filter(Assignment.week_number == week_number).delete()
            
            # 2. Lấy dữ liệu Cụm trực và Sao đỏ đang hoạt động
            areas = db_session.query(DutyArea).all()
            stars = db_session.query(RedStar).filter_by(is_active=True).all()
            
            if not areas or not stars:
                flash("Lỗi: Thiếu dữ liệu Khu vực trực hoặc Đội Sao đỏ để phân công!", "error")
                return redirect(url_for('assignments', week=week_number))
                
            # 3. Đọc cấu hình Sơ đồ lớp để né (Không trực lớp mình)
            base_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(base_dir, "config", "class_zones.json")
            zones_map = {}
            if os.path.exists(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    try: zones_map = json.load(f)
                    except: pass
            
            # HÀM BỔ TRỢ: Trích xuất Khối (10, 11, 12) từ tên lớp
            def get_grade(class_name):
                match = re.search(r'(10|11|12)', str(class_name))
                return match.group(1) if match else ""

            # 4. [NÂNG CẤP LÕI - QUY TẮC 2]: TRÍCH XUẤT LỊCH SỬ ĐỂ ÉP LUẬT XOAY VÒNG
            history_counts = {star.id: {} for star in stars}
            past_assignments = db_session.query(Assignment).filter(Assignment.week_number < week_number).all()
            
            for pa in past_assignments:
                # [ĐÃ VÁ LỖI]: Gọi thẳng Object pa.duty_area.id thay vì gọi cột database
                if pa.red_star_id in history_counts and pa.duty_area:
                    area_id_val = pa.duty_area.id
                    history_counts[pa.red_star_id][area_id_val] = history_counts[pa.red_star_id].get(area_id_val, 0) + 1

            # Biến đếm khối lượng công việc trong Tuần hiện tại
            current_week_shift_counts = {star.id: 0 for star in stars}
            success_count = 0
            
            # Xáo trộn mảng cụm trực để đổi mới ngẫu nhiên thứ tự bốc thăm
            random.shuffle(areas)
            
            # 5. Bắt đầu xếp lịch
            for shift in shifts:
                available_stars = list(stars)
                random.shuffle(available_stars) # Trộn ngẫu nhiên ban đầu
                
                for area in areas:
                    req_count = area.required_stars or 2
                    assigned_count = 0
                    
                    # Xác định CÁC KHỐI LỚP (10, 11, 12) có mặt trong Cụm trực này
                    area_classes = [c.strip().upper() for c in zones_map.get(area.name, [])]
                    area_grades = {get_grade(c) for c in area_classes if get_grade(c)}
                    
                    # [NÂNG CẤP LÕI - QUY TẮC 1]: Lọc ra danh sách Sao đỏ hợp lệ
                    valid_stars_for_area = []
                    for star in available_stars:
                        star_class = star.branch.name.strip().upper() if star.branch else ""
                        star_grade = get_grade(star_class)
                        
                        is_conflict = False
                        # Nếu Cụm trực chứa lớp có cùng Khối với Sao đỏ -> XUNG ĐỘT
                        if star_grade and star_grade in area_grades:
                            is_conflict = True
                        
                        # Xử lý dự phòng cho cụm trực chưa có trong Sơ đồ lớp
                        if not is_conflict and "KHỐI" in area.name.upper():
                            if star_grade and star_grade in area.name:
                                is_conflict = True
                                
                        if not is_conflict:
                            valid_stars_for_area.append(star)
                            
                    # Sắp xếp theo ưu tiên: 1. Ít trực cụm này nhất -> 2. Ít việc trong tuần nhất
                    valid_stars_for_area.sort(key=lambda s: (
                        history_counts[s.id].get(area.id, 0),
                        current_week_shift_counts[s.id]
                    ))
                    
                    # Tiến hành bốc người vào Cụm trực
                    stars_to_remove_from_shift = []
                    for star in valid_stars_for_area:
                        if assigned_count >= req_count:
                            break
                            
                        # Chốt phân công
                        new_assign = Assignment(
                            week_number=week_number,
                            shift=shift,
                            date=start_date,
                            red_star=star,      
                            duty_area=area      
                        )
                        db_session.add(new_assign)
                        
                        current_week_shift_counts[star.id] += 1
                        history_counts[star.id][area.id] = history_counts[star.id].get(area.id, 0) + 1
                        assigned_count += 1
                        stars_to_remove_from_shift.append(star)
                        success_count += 1
                            
                    # Trực xong 1 ca rồi thì loại ra khỏi vòng lặp của buổi đó
                    for s in stars_to_remove_from_shift:
                        if s in available_stars:
                            available_stars.remove(s)
            
            db_session.commit()
            
            if success_count > 0:
                log_system_action("LỊCH TRỰC", f"Đã chạy thuật toán xếp lịch tự động cho Tuần {week_number}")
                flash(f"✅ Đã phân công tự động Tuần {week_number} thành công! ({success_count} lượt trực)", "success")
            else:
                flash("⚠️ Thuật toán chạy xong nhưng không thể xếp lịch (Có thể do không đủ người đáp ứng điều kiện chéo tuyến).", "warning")
                
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi hệ thống khi phân công: {str(e)}", "error")
        
    return redirect(url_for('assignments', week=week_number))

# --- XUẤT EXCEL LỊCH TRỰC ---
@app.route('/export_schedule/<int:week>')
def export_schedule(week):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('assignments', week=week))
                
            assignments = db_session.query(Assignment).join(RedStar).join(Branch).filter(
                Assignment.week_number == week,
                Branch.school_year_id == active_year.id
            ).order_by(Assignment.shift, Assignment.id).all()
            
            if not assignments:
                flash(f"Chưa có dữ liệu lịch trực tuần {week} để xuất!", "error")
                return redirect(url_for('assignments', week=week))

            zones_map = {}
            if os.path.exists("config/class_zones.json"):
                with open("config/class_zones.json", "r", encoding="utf-8") as f:
                    zones_map = json.load(f)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Tuan_{week}"
            
            ws['A1'] = "ĐOÀN TRƯỜNG THPT THANH HÒA"
            ws['D1'] = "ĐOÀN TNCS HỒ CHÍ MINH"
            ws['A1'].font = Font(name="Times New Roman", size=11, bold=True)
            ws['D1'].font = Font(name="Times New Roman", size=11, bold=True)
            ws['D1'].alignment = Alignment(horizontal="right")
            
            ws['A3'] = f"LỊCH TRỰC SAO ĐỎ - TUẦN {week}"
            ws['A3'].font = Font(name="Times New Roman", size=14, bold=True)
            
            headers = ["STT", "Họ Và Tên", "Vị trí trực", "Ghi chú"]
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            for col_num, h_title in enumerate(headers, 1):
                c = ws.cell(row=5, column=col_num, value=h_title)
                c.font = Font(name="Times New Roman", size=11, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
                
            for idx, assign in enumerate(assignments, 1):
                area_name = assign.duty_area.name if assign.duty_area else ""
                disp_area = ", ".join(zones_map.get(area_name, [])) if area_name in zones_map else area_name
                star_name = assign.red_star.full_name if assign.red_star else ""
                branch_name = assign.red_star.branch.name if assign.red_star and assign.red_star.branch else ""
                
                row_idx = idx + 5
                c1 = ws.cell(row=row_idx, column=1, value=idx)
                c2 = ws.cell(row=row_idx, column=2, value=f"{star_name} ({branch_name})")
                c3 = ws.cell(row=row_idx, column=3, value=f"{disp_area} ({assign.shift})")
                c4 = ws.cell(row=row_idx, column=4, value="")
                
                for c in [c1, c2, c3, c4]:
                    c.font = Font(name="Times New Roman", size=11)
                    c.border = border
                c1.alignment = Alignment(horizontal="center")
                
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 20
            
            log_system_action("XUẤT EXCEL", f"Xuất lịch trực Tuần {week}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name=f"Lich_Truc_Tuan_{week}.xlsx", as_attachment=True)
    except Exception as e:
        flash(f"Lỗi xuất excel lịch trực: {e}", "error")
        return redirect(url_for('assignments', week=week))

# --- API LẤY DANH SÁCH GỢI Ý ĐỔI NGƯỜI THÔNG MINH ---
@app.route('/api/get_swap_candidates/<int:assign_id>')
def api_get_swap_candidates(assign_id):
    try:
        with session_scope() as db_session:
            assign = db_session.query(Assignment).filter_by(id=assign_id).first()
            if not assign: return {"error": "Không tìm thấy lịch trực"}
            
            week_num = assign.week_number
            shift = assign.shift
            area_name = assign.duty_area.name if assign.duty_area else ""
            current_star_id = assign.red_star_id
            
            active_stars = db_session.query(RedStar).filter_by(is_active=True).all()
            shift_assignments = db_session.query(Assignment).filter_by(week_number=week_num, shift=shift).all()
            busy_map = {a.red_star_id: (a.id, a.duty_area.name if a.duty_area else "") for a in shift_assignments}
            
            zones_map = {}
            if os.path.exists("config/class_zones.json"):
                with open("config/class_zones.json", "r", encoding="utf-8") as f:
                    zones_map = json.load(f)
                    
            restricted_classes = [c.upper() for c in zones_map.get(area_name, [])]
            
            is_gate_chinh = "Cổng chính" in area_name
            is_gate_phu = "Cổng phụ" in area_name
            is_giam_sat = "Giám sát" in area_name
            is_cum = not (is_gate_chinh or is_gate_phu or is_giam_sat)
            
            free_list = []
            busy_list = []
            
            for star in active_stars:
                if star.id == current_star_id: continue
                
                star_class_name = star.branch.name.upper() if star.branch else ""
                is_owner = False
                if star_class_name in restricted_classes: is_owner = True
                elif "Khối" in area_name:
                    grade_num = "".join(filter(str.isdigit, area_name))
                    if grade_num and star_class_name.startswith(grade_num): is_owner = True
                        
                if is_owner: continue
                    
                is_busy = star.id in busy_map
                swap_valid = True
                target_assign_id = None
                target_area_name = ""
                
                if is_busy:
                    target_assign_id, target_area_name = busy_map[star.id]
                    t_gate_chinh = "Cổng chính" in target_area_name
                    t_gate_phu = "Cổng phụ" in target_area_name
                    t_giam_sat = "Giám sát" in target_area_name
                    t_cum = not (t_gate_chinh or t_gate_phu or t_giam_sat)
                    
                    rule_matched = False
                    if is_gate_chinh and t_gate_phu: rule_matched = True
                    elif is_gate_phu and t_gate_chinh: rule_matched = True
                    elif is_cum and t_giam_sat: rule_matched = True
                    elif is_giam_sat and t_cum: rule_matched = True
                    
                    if not rule_matched: swap_valid = False
                    
                    if swap_valid:
                        curr_class = assign.red_star.branch.name.upper() if assign.red_star and assign.red_star.branch else ""
                        target_restricted = [c.upper() for c in zones_map.get(target_area_name, [])]
                        if curr_class in target_restricted: swap_valid = False
                        elif "Khối" in target_area_name:
                            t_grade = "".join(filter(str.isdigit, target_area_name))
                            if t_grade and curr_class.startswith(t_grade): swap_valid = False
                
                if swap_valid:
                    item = {
                        "star_id": star.id,
                        "star_name": f"{star.full_name} ({star.branch.name if star.branch else ''})",
                        "target_assign_id": target_assign_id,
                        "target_area_name": target_area_name
                    }
                    if is_busy: busy_list.append(item)
                    else: free_list.append(item)
                    
            return {"free": free_list[:10], "busy": busy_list}
    except Exception as e:
        return {"error": str(e)}

@app.route('/execute_swap', methods=['POST'])
def execute_swap():
    assign_id = int(request.form.get('assign_id'))
    new_star_id = int(request.form.get('new_star_id'))
    target_assign_id = request.form.get('target_assign_id')
    
    try:
        with session_scope() as db_session:
            assign = db_session.query(Assignment).filter_by(id=assign_id).first()
            if not assign:
                flash("Lỗi: Không tìm thấy phân công gốc!", "error")
                return redirect(url_for('assignments'))
                
            week = assign.week_number
            if target_assign_id and target_assign_id != "None":
                target_assign = db_session.query(Assignment).filter_by(id=int(target_assign_id)).first()
                if target_assign:
                    temp = assign.red_star_id
                    assign.red_star_id = target_assign.red_star_id
                    target_assign.red_star_id = temp
                    log_system_action("LỊCH TRỰC", f"Hoán đổi vị trí trực chéo thành công trong Tuần {week}")
                    flash("✅ Đã hoán đổi chéo vị trí trực thành công!", "success")
            else:
                assign.red_star_id = new_star_id
                log_system_action("LỊCH TRỰC", f"Đổi người rảnh vào ca trực Tuần {week}")
                flash("✅ Đã thay thế người rảnh vào ca trực thành công!", "success")
                
            return redirect(url_for('assignments', week=week))
    except Exception as e:
        flash(f"Lỗi thực hiện đổi người trực: {e}", "error")
        return redirect(url_for('assignments'))
        
# --- 3. ĐÁNH GIÁ VÀ CHẤM ĐIỂM SAO ĐỎ (KPI 360 ĐỘ) ---

@app.route('/api/submit_evaluation', methods=['POST'])
def api_submit_evaluation():
    """API Nhận dữ liệu đánh giá trắc nghiệm - Bản bọc thép chống sập HTML"""
    import traceback # Import trực tiếp bên trong để đảm bảo luôn có hàm in lỗi
    
    try:
        # 1. BẮT DỮ LIỆU ĐẦU VÀO CỰC KỲ CHẶT CHẼ
        # Sử dụng get_json(force=True) để ép Flask đọc JSON dù thiếu Header
        data = request.get_json(silent=True, force=True) 
        if not data:
            data = request.form.to_dict() # Dự phòng nếu gửi bằng form
            
        if not data:
            # Trả về Dictionaries thuần túy, Flask hiện đại tự convert thành JSON
            return {"success": False, "error": "Máy chủ không nhận được dữ liệu (Payload trống)!"}, 400

        evaluatee_id_raw = data.get('evaluatee_id')
        week_name = data.get('week_name')
        
        if not evaluatee_id_raw or not week_name:
            return {"success": False, "error": "Thiếu mã học sinh hoặc tên tuần!"}, 400

        try:
            evaluatee_id = int(evaluatee_id_raw)
        except ValueError:
            return {"success": False, "error": "Mã học sinh không đúng định dạng số!"}, 400
        
        score_gio_giac = int(data.get('score_gio_giac', 5))
        score_tac_phong = int(data.get('score_tac_phong', 5))
        score_thai_do = int(data.get('score_thai_do', 5))
        score_cong_tam = int(data.get('score_cong_tam', 5))
        comment = str(data.get('comment', '')).strip()

        evaluator_username = session.get('username', 'Unknown')
        evaluator_role = session.get('role', 'Giáo viên/Đoàn')

        # 2. XỬ LÝ CƠ SỞ DỮ LIỆU
        with session_scope() as db_session:
            # Kiểm tra Sao đỏ
            target_star = db_session.query(RedStar).filter_by(id=evaluatee_id).first()
            if not target_star:
                return {"success": False, "error": "Học sinh Sao đỏ này không còn tồn tại trong hệ thống!"}, 404

            # Cập nhật hoặc tạo mới
            existing_eval = db_session.query(StarEvaluation).filter_by(
                evaluator_username=str(evaluator_username),
                evaluatee_id=evaluatee_id,
                week_name=str(week_name)
            ).first()
            
            if existing_eval:
                existing_eval.score_gio_giac = score_gio_giac
                existing_eval.score_tac_phong = score_tac_phong
                existing_eval.score_thai_do = score_thai_do
                existing_eval.score_cong_tam = score_cong_tam
                existing_eval.comment = comment
                msg = "Đã lưu bản cập nhật phiếu đánh giá thành công!"
            else:
                new_eval = StarEvaluation(
                    evaluator_username=str(evaluator_username),
                    evaluator_role=str(evaluator_role),
                    evaluatee_id=evaluatee_id,
                    week_name=str(week_name),
                    score_gio_giac=score_gio_giac,
                    score_tac_phong=score_tac_phong,
                    score_thai_do=score_thai_do,
                    score_cong_tam=score_cong_tam,
                    comment=comment
                )
                db_session.add(new_eval)
                msg = "Đã nộp phiếu đánh giá mới thành công!"
                
            log_system_action("ĐÁNH GIÁ KPI", f"User {evaluator_username} chấm điểm Sao đỏ ID {evaluatee_id}")
            
            # Trả về thành công
            return {"success": True, "message": msg}, 200

    except Exception as e:
        traceback.print_exc() # In màn hình đen (Console)
        # Bắt mọi lỗi sập nguồn và trả về JSON an toàn
        return {"success": False, "error": f"Lỗi nội bộ Server: {str(e)}"}, 500
    
@app.route('/star-evaluations', methods=['GET', 'POST'])
def star_evaluations():
    """Trang quản trị xem và quản lý kết quả KPI (Dành cho Admin/Bí thư)"""
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            # Tự động nhận diện Tuần Đánh giá mới nhất
            week_param = request.args.get('week')
            if week_param:
                current_week = int(week_param)
            else:
                latest_eval = db_session.query(StarEvaluation).order_by(StarEvaluation.id.desc()).first()
                if latest_eval:
                    # Lấy số từ chuỗi (Ví dụ "Tuần 5" -> 5)
                    import re
                    match = re.search(r'\d+', latest_eval.week_name)
                    current_week = int(match.group()) if match else 1
                else:
                    latest_assign = db_session.query(Assignment).order_by(Assignment.week_number.desc()).first()
                    current_week = latest_assign.week_number if latest_assign else 1

            week_name_str = f"Tuần {current_week}"
            stars = db_session.query(RedStar).filter(RedStar.is_active == True).all()
            
            # Lấy toàn bộ đánh giá của tuần
            evaluations = db_session.query(StarEvaluation).filter(StarEvaluation.week_name == week_name_str).all()
            
            # Gom nhóm đánh giá theo từng Sao đỏ để tính KPI trung bình
            eval_dict = {}
            for e in evaluations:
                if e.evaluatee_id not in eval_dict:
                    eval_dict[e.evaluatee_id] = []
                eval_dict[e.evaluatee_id].append(e)

            # Tính toán Tổng điểm KPI hiển thị ra màn hình
            kpi_summary = {}
            for star in stars:
                if star.id in eval_dict:
                    evals = eval_dict[star.id]
                    total_points = sum((ev.score_gio_giac + ev.score_tac_phong + ev.score_thai_do + ev.score_cong_tam) for ev in evals)
                    avg_score = total_points / len(evals) # Điểm trung bình / Lượt đánh giá
                    kpi_percent = (avg_score / 20) * 100  # Quy ra % (Tối đa 20đ/lượt)
                    
                    kpi_summary[star.id] = {
                        'count': len(evals),
                        'avg_score': round(avg_score, 1),
                        'kpi_percent': round(kpi_percent, 1),
                        'evals_detail': evals
                    }

            return render_template('star_evaluations.html', 
                                   active_year=active_year, 
                                   stars=stars, 
                                   kpi_summary=kpi_summary, 
                                   current_week=current_week)
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi phân hệ đánh giá KPI: {e}", "error")
        return redirect(url_for('dashboard'))

# ==========================================
# MODULE: XẾP HẠNG THI ĐUA ĐỘI SAO ĐỎ (THÁNG / HỌC KỲ / NĂM)
# ==========================================
@app.route('/star-ranking/<report_type>', methods=['GET', 'POST'])
def star_ranking(report_type):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            available_weeks = []
            if active_year:
                # Tìm các tuần đã có dữ liệu đánh giá
                weeks_db = db_session.query(StarEvaluation.week_name).distinct().all()
                import re
                week_nums = []
                for w in weeks_db:
                    match = re.search(r'\d+', w[0])
                    if match: week_nums.append(int(match.group()))
                available_weeks = sorted(list(set(week_nums)))

            ranking_data = []
            selected_title = ""
            selected_weeks = []
            display_selected_weeks = []

            if request.method == 'POST' or report_type == 'yearly':
                if report_type == 'monthly':
                    selected_title = request.form.get('title', 'Tháng 9')
                    display_selected_weeks = [int(w) for w in request.form.getlist('weeks')]
                    selected_weeks = [f"Tuần {w}" for w in display_selected_weeks]
                elif report_type == 'semester':
                    selected_title = request.form.get('title', 'Học Kỳ 1')
                    display_selected_weeks = [int(w) for w in request.form.getlist('weeks')]
                    selected_weeks = [f"Tuần {w}" for w in display_selected_weeks]
                elif report_type == 'yearly':
                    selected_title = f"Năm học {active_year.name}" if active_year else "Năm học"
                    display_selected_weeks = available_weeks
                    selected_weeks = [f"Tuần {w}" for w in display_selected_weeks]

                if not selected_weeks and report_type != 'yearly':
                    flash("Vui lòng chọn ít nhất 1 tuần để tổng hợp điểm!", "error")
                elif active_year:
                    stars = db_session.query(RedStar).filter(RedStar.is_active == True).all()
                    for star in stars:
                        if star.branch and star.branch.school_year_id == active_year.id:
                            # [NÂNG CẤP LÕI]: Dùng evaluatee_id theo cấu trúc mới
                            evals = db_session.query(StarEvaluation).filter(
                                StarEvaluation.evaluatee_id == star.id,
                                StarEvaluation.week_name.in_(selected_weeks)
                            ).all()
                            
                            if evals:
                                # [NÂNG CẤP LÕI]: Tính tổng từ 4 cột trắc nghiệm
                                total_score = sum((e.score_gio_giac + e.score_tac_phong + e.score_thai_do + e.score_cong_tam) for e in evals)
                                eval_count = len(evals)
                                avg_score = total_score / eval_count
                                kpi_percent = (avg_score / 20) * 100
                                
                                ranking_data.append({
                                    'star_name': star.full_name,
                                    'branch_name': star.branch.name,
                                    'total_score': total_score,
                                    'eval_count': eval_count,
                                    'avg_score': round(avg_score, 1),
                                    'kpi_percent': round(kpi_percent, 1)
                                })
                    
                    # Sắp xếp dựa trên KPI Trung bình để công bằng cho người trực ít / trực nhiều
                    ranking_data = sorted(ranking_data, key=lambda x: x['avg_score'], reverse=True)
                    current_rank = 1
                    for i, d in enumerate(ranking_data):
                        if i > 0 and d['avg_score'] < ranking_data[i-1]['avg_score']:
                            current_rank = i + 1
                        d['rank'] = current_rank

            if report_type == 'monthly': page_title = "Đánh giá Tháng"
            elif report_type == 'semester': page_title = "Đánh giá Học kỳ"
            else: page_title = "Tổng kết Năm học"
            
            return render_template(
                'star_ranking.html', 
                report_type=report_type,
                page_title=page_title,
                available_weeks=available_weeks, 
                ranking_data=ranking_data,
                selected_title=selected_title, 
                selected_weeks=display_selected_weeks,
                active_year=active_year
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi hệ thống khi tổng hợp đánh giá Sao đỏ: {e}", "error")
        return redirect(url_for('dashboard'))

# ==========================================
# API: KIỂM TRA VÀ ĐỔI TRẠNG THÁI KHÓA SỔ TUẦN
# ==========================================
@app.route('/api/toggle_week_lock', methods=['POST'])
def api_toggle_week_lock():
    try:
        data = request.get_json()
        week_name = data.get('week_name')
        
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                return {"success": False, "error": "Chưa có năm học kích hoạt!"}
                
            scores = db_session.query(WeeklyScore).join(Branch).filter(
                WeeklyScore.week == week_name,
                Branch.school_year_id == active_year.id
            ).all()
            
            current_lock_status = any(getattr(s, 'is_locked', False) for s in scores)
            new_status = not current_lock_status
            
            for s in scores:
                if new_status == True and not s.is_locked: # Chỉ kích hoạt gộp lỗi khi Khóa sổ
                    if s.note:
                        import re
                        all_categories = db_session.query(ViolationCategory).filter_by(school_year_id=s.branch.school_year_id).all()
                        sorted_cats = sorted(all_categories, key=lambda x: len(x.name), reverse=True)
                        parsed_errors = {}
                        
                        parts = re.split(r'[,;+\n](?![^\[]*\])(?![^\(]*\))', s.note)
                        for part in parts:
                            part_clean = part.strip()
                            if not part_clean: continue
                            
                            stu_name_raw = ""
                            match_stu = re.search(r'\[(.*?)\]', part_clean)
                            if match_stu: stu_name_raw = match_stu.group(1).strip()
                            
                            # Chuẩn hóa tên học sinh: Bỏ khoảng trắng thừa, viết hoa chữ cái đầu (Không phân biệt hoa/thường)
                            stu_name_normalized = " ".join(stu_name_raw.split()).title() if stu_name_raw else ""
                            stu_name_key = stu_name_normalized.lower() # Dùng key chữ thường để so sánh chính xác tuyệt đối
                            
                            matched = False
                            for cat in sorted_cats:
                                if cat.name.lower() in part_clean.lower():
                                    match_qty = re.search(r'(?:x|:|-)\s*(\d+)', part_clean.lower())
                                    qty = int(match_qty.group(1)) if match_qty else 1
                                    
                                    # [TIÊU CHÍ CỐT LÕI]: Chỉ gộp chung khi CÙNG TÊN LỖI (cat.name) và CÙNG TÊN HỌC SINH (stu_name_key)
                                    key = (cat.name, stu_name_key, stu_name_normalized)
                                    if key not in parsed_errors:
                                        parsed_errors[key] = 0
                                    parsed_errors[key] += qty
                                    matched = True
                                    break
                            
                            if not matched:
                                parsed_errors[("MANUAL", part_clean.lower(), part_clean)] = 1
                                
                        final_parts = []
                        for (cat_name, stu_key, stu_display), qty in parsed_errors.items():
                            if cat_name == "MANUAL":
                                final_parts.append(stu_display)
                            else:
                                if stu_display: 
                                    final_parts.append(f"{cat_name} x{qty} [{stu_display}]")
                                else: 
                                    final_parts.append(f"{cat_name} x{qty}")
                                
                        s.note = " ; ".join(final_parts)
                
                s.is_locked = new_status
                
            status_text = "Khóa sổ (Đã chốt)" if new_status else "Mở khóa sổ"
            log_system_action("CHỐT SỔ", f"Đã chuyển trạng thái {week_name} sang: {status_text}")
            
            return {"success": True, "is_locked": new_status, "message": f"Đã {status_text} thành công {week_name}!"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.route('/api/get_week_lock_status/<week_name>')
def api_get_week_lock_status(week_name):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return {"is_locked": False}
            
            scores = db_session.query(WeeklyScore).join(Branch).filter(
                WeeklyScore.week == week_name,
                Branch.school_year_id == active_year.id
            ).all()
            
            is_locked = any(getattr(s, 'is_locked', False) for s in scores)
            return {"is_locked": is_locked}
    except Exception as e:
        return {"is_locked": False}
# ==========================================
# MODULE: NHẬP ĐIỂM TUẦN & TỰ ĐỘNG BÓC TÁCH LỖI VÀO SỔ ĐEN
# ==========================================
@app.route('/weekly', methods=['GET', 'POST'])
def weekly():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            if request.method == 'POST':
                if not active_year: flash("Chưa có năm học nào được kích hoạt!", "error"); return redirect(url_for('weekly'))
                    
                week_name = request.form.get('week_name', 'Tuần 1')
                scores_check = db_session.query(WeeklyScore).join(Branch).filter(WeeklyScore.week == week_name, Branch.school_year_id == active_year.id).all()
                if any(getattr(s, 'is_locked', False) for s in scores_check):
                    flash(f"⚠️ {week_name} đã được chốt sổ (khóa điểm)! Toàn bộ dữ liệu đã được đóng băng, không thể chỉnh sửa.", "error")
                    return redirect(url_for('weekly', week=week_name))

                start_date_val = request.form.get('start_date', '').strip()
                end_date_val = request.form.get('end_date', '').strip()
                branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                
                DIEM_8, DIEM_9, DIEM_10 = 1.0, 3.0, 5.0; TUAN_KHA, TUAN_TOT = 20.0, 30.0
                settings = db_session.query(ScoreSettings).filter_by(school_year_id=active_year.id).first()
                if settings:
                    DIEM_8, DIEM_9, DIEM_10 = float(settings.diem_8), float(settings.diem_9), float(settings.diem_10)
                    TUAN_KHA, TUAN_TOT = float(settings.diem_tuan_kha), float(settings.diem_tuan_tot)
                    
                all_categories = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).all() if active_year else []
                actor_username = session.get('username', 'Hệ thống'); actor_fullname = session.get('full_name', 'Người dùng')
                
                for branch in branches:
                    b_id = str(branch.id); b_group = str(branch.group) if branch.group else "1"
                    rating = request.form.get(f'rating_{b_id}', 'Bình thường')
                    try: c_8 = int(request.form.get(f'c8_{b_id}', 0) or 0)
                    except: c_8 = 0
                    try: c_9 = int(request.form.get(f'c9_{b_id}', 0) or 0)
                    except: c_9 = 0
                    try: c_10 = int(request.form.get(f'c10_{b_id}', 0) or 0)
                    except: c_10 = 0
                    try: truc = float(request.form.get(f'truc_{b_id}', 100.0) or 100.0)
                    except: truc = 100.0
                    try: cong = float(request.form.get(f'cong_{b_id}', 0.0) or 0.0)
                    except: cong = 0.0
                    
                    note = request.form.get(f'note_{b_id}', '').strip()
                    
                    diem_quy_uoc = 0.0
                    if "1" in b_group: diem_quy_uoc = (c_9 * DIEM_9) + (c_10 * DIEM_10)
                    elif "2" in b_group: diem_quy_uoc = (c_8 * DIEM_8) + (c_9 * DIEM_9) + (c_10 * DIEM_10)
                    else: diem_quy_uoc = (c_9 * DIEM_9) + (c_10 * DIEM_10)

                    diem_xep_loai = 0.0
                    if rating == "Tuần Tốt": diem_xep_loai = TUAN_TOT
                    elif rating == "Tuần Khá": diem_xep_loai = TUAN_KHA

                    diem_tru_auto = 0.0; new_violations = [] 
                    
                    if note:
                        parts = re.split(r'[,;+\n](?![^\[]*\])(?![^\(]*\))', note)
                        sorted_cats = sorted(all_categories, key=lambda x: len(x.name), reverse=True)
                        for part in parts:
                            part_clean = part.strip()
                            if not part_clean: continue
                            part_lower = part_clean.lower()
                            for cat in sorted_cats:
                                if cat.name.lower() in part_lower:
                                    match_qty = re.search(r'(?:x|:|-)\s*(\d+)', part_lower)
                                    qty = int(match_qty.group(1)) if match_qty else 1 
                                    if getattr(cat, 'point_type', 'Điểm trừ') != 'Điểm cộng': diem_tru_auto += float(cat.penalty_points * qty)
                                    match_name = re.search(r'\[(.*?)\]|\((.*?)\)', part_clean)
                                    student_name = None
                                    if match_name:
                                        student_name = match_name.group(1) if match_name.group(1) is not None else match_name.group(2)
                                        student_name = student_name.strip()
                                    new_violations.append({'violation_id': cat.id, 'quantity': qty, 'student_name': student_name})
                                    break

                    # [VÁ LỖI TRỪ ĐIỂM KÉP]: Điểm trừ tổng ĐƯỢC QUÉT LẠI HOÀN TOÀN từ Ghi chú
                    tong_diem_tru = diem_tru_auto
                    tru = diem_tru_auto # Đồng bộ lại biến tru để lưu vào CSDL
                    
                    total_val = truc + diem_xep_loai + diem_quy_uoc + cong - tong_diem_tru
                    
                    score = db_session.query(WeeklyScore).filter_by(branch_id=branch.id, week=week_name).first()
                    old_val = float(score.total_score) if score and score.total_score is not None else None
                    
                    if score:
                        if old_val is not None and old_val != total_val:
                            log_details = f"Sửa điểm lớp {branch.name} ({week_name}): Từ {old_val}đ thành {total_val}đ"
                            db_session.add(ActionLog(username=actor_username, full_name=actor_fullname, action_type="THAY ĐỔI ĐIỂM", details=log_details))
                            
                        score.week_rating = rating; score.count_8 = c_8; score.count_9 = c_9; score.count_10 = c_10; score.score_truc = truc; score.score_cong = cong; score.score_tru = tru; score.note = note; score.total_score = total_val; score.start_date = start_date_val; score.end_date = end_date_val
                    else:
                        score = WeeklyScore(branch_id=branch.id, week=week_name, week_rating=rating, count_8=c_8, count_9=c_9, count_10=c_10, score_truc=truc, score_cong=cong, score_tru=tru, note=note, total_score=total_val, start_date=start_date_val, end_date=end_date_val)
                        db_session.add(score); db_session.flush() 
                        log_details = f"Khởi tạo điểm mới lớp {branch.name} ({week_name}): {total_val}đ"
                        db_session.add(ActionLog(username=actor_username, full_name=actor_fullname, action_type="KHỞI TẠO ĐIỂM", details=log_details))
                        
                    db_session.query(WeeklyViolation).filter_by(weekly_score_id=score.id).delete()
                    for v in new_violations:
                        db_session.add(WeeklyViolation(weekly_score_id=score.id, violation_id=v['violation_id'], quantity=v['quantity'], student_name=v['student_name']))
                
                log_system_action("LƯU ĐIỂM TUẦN", f"Đã lưu và cập nhật bảng điểm {week_name}")
                flash(f"Đã lưu và cập nhật chính xác bảng điểm {week_name}!", "success")
                return redirect(url_for('weekly', week=week_name))

            current_week = request.args.get('week', 'Tuần 1'); branches_data = []; cat_list = []
            categories = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).all() if active_year else []
            for c in categories: cat_list.append({"name": c.name, "points": float(c.penalty_points), "type": getattr(c, 'point_type', 'Điểm trừ')})
            categories_json = json.dumps(cat_list)

            start_date_str = ""; end_date_str = ""
            existing_score = db_session.query(WeeklyScore).join(Branch).filter(WeeklyScore.week == current_week, Branch.school_year_id == active_year.id if active_year else True).first()
            if existing_score and existing_score.start_date: start_date_str = existing_score.start_date; end_date_str = existing_score.end_date or ""
            else:
                try:
                    import datetime as dt
                    week_num = int(current_week.replace("Tuần ", "").strip())
                    assign_record = db_session.query(Assignment).filter(Assignment.week_number == week_num).first()
                    if assign_record and hasattr(assign_record, 'date') and assign_record.date:
                        py_date = assign_record.date; start_date_str = py_date.strftime("%Y-%m-%d")
                        day_of_week = py_date.weekday()
                        if day_of_week <= 5: days_to_add = 5 - day_of_week
                        else: days_to_add = 6
                        end_date = py_date + dt.timedelta(days=days_to_add)
                        end_date_str = end_date.strftime("%Y-%m-%d")
                except Exception as e: pass

            if active_year:
                branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                for b in branches:
                    sc = db_session.query(WeeklyScore).filter_by(branch_id=b.id, week=current_week).first()
                    branches_data.append({'branch': b, 'score': sc})
                    
            settings = db_session.query(ScoreSettings).filter_by(school_year_id=active_year.id).first() if active_year else None
            return render_template('weekly.html', branches_data=branches_data, active_year=active_year, current_week=current_week, categories_json=categories_json, start_date=start_date_str, end_date=end_date_str, score_settings=settings)
    except Exception as e:
        flash(f"Lỗi nhập điểm tuần: {e}", "error")
        return redirect(url_for('dashboard'))
# ==========================================
# API: LƯU CẤU HÌNH BAREM ĐIỂM
# ==========================================
@app.route('/save_settings', methods=['POST'])
def save_settings():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Chỉ Ban chấp hành Đoàn trường mới có quyền thay đổi Barem!", "error")
        return redirect(request.referrer or url_for('dashboard'))
        
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(request.referrer or url_for('dashboard'))
                
            # Lấy dữ liệu từ form trên giao diện
            diem_8 = float(request.form.get('diem_8', 1.0))
            diem_9 = float(request.form.get('diem_9', 3.0))
            diem_10 = float(request.form.get('diem_10', 5.0))
            tuan_kha = float(request.form.get('tuan_kha', 20.0))
            tuan_tot = float(request.form.get('tuan_tot', 30.0))
            max_tot = int(request.form.get('max_tot', 14))
            max_mon = int(request.form.get('max_mon', 4))
            
            # Cập nhật vào DB
            settings = db_session.query(ScoreSettings).filter_by(school_year_id=active_year.id).first()
            if settings:
                settings.diem_8 = diem_8; settings.diem_9 = diem_9; settings.diem_10 = diem_10
                settings.diem_tuan_kha = tuan_kha; settings.diem_tuan_tot = tuan_tot
                settings.max_diem_tot = max_tot; settings.max_diem_mon = max_mon
            else:
                settings = ScoreSettings(
                    school_year_id=active_year.id,
                    diem_8=diem_8, diem_9=diem_9, diem_10=diem_10,
                    diem_tuan_kha=tuan_kha, diem_tuan_tot=tuan_tot,
                    max_diem_tot=max_tot, max_diem_mon=max_mon
                )
                db_session.add(settings)
                
            log_system_action("CẤU HÌNH", f"Cập nhật Barem điểm: Đ.10={diem_10}, Đ.9={diem_9}, Đ.8={diem_8}")
            flash("✅ Đã cập nhật cấu hình Barem điểm thành công! Hệ thống sẽ áp dụng Barem mới.", "success")
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi lưu Barem: {e}", "error")
        
    return redirect(request.referrer or url_for('weekly'))    
# ==========================================
# MODULE TẠO FILE EXCEL SỔ ĐEN
# ==========================================
@app.route('/preview_blacklist')
def preview_blacklist():
    try:
        with session_scope() as db_session:
            week_name = request.args.get('week', 'Tuần 1')
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            
            if not active_year:
                flash("Chưa có năm học nào được kích hoạt!", "error")
                return redirect(url_for('weekly', week=week_name))
                
            raw_violations = db_session.query(WeeklyViolation, WeeklyScore, Branch, ViolationCategory)\
                .join(WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id)\
                .join(Branch, WeeklyScore.branch_id == Branch.id)\
                .join(ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id)\
                .filter(
                    WeeklyScore.week == week_name, 
                    Branch.school_year_id == active_year.id
                ).all()
                
            violations = []
            for v, s, b, c in raw_violations:
                if v.student_name and str(v.student_name).strip() != "":
                    violations.append({
                        'branch_name': b.name,
                        'student_name': v.student_name,
                        'violation_name': c.name,
                        'quantity': v.quantity
                    })
                    
            # Sắp xếp danh sách vi phạm theo tên Chi đoàn (từ A-Z)
            violations.sort(key=lambda x: x['branch_name'])
                
            return render_template('preview_blacklist.html', violations=violations, week_name=week_name)
    except Exception as e:
        flash(f"Lỗi xem trước sổ đen: {str(e)}", "error")
        return redirect(url_for('weekly'))
    
@app.route('/export_blacklist', methods=['GET', 'POST'])
def export_blacklist():
    try:
        with session_scope() as db_session:
            week_name = request.args.get('week', 'Tuần 1')
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            
            if not active_year:
                flash("Chưa có năm học nào được kích hoạt!", "error")
                return redirect(url_for('weekly', week=week_name))
                
            raw_violations = db_session.query(WeeklyViolation, WeeklyScore, Branch, ViolationCategory)\
                .join(WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id)\
                .join(Branch, WeeklyScore.branch_id == Branch.id)\
                .join(ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id)\
                .filter(
                    WeeklyScore.week == week_name, 
                    Branch.school_year_id == active_year.id
                ).all()
                
            violations = []
            for v, s, b, c in raw_violations:
                if v.student_name and str(v.student_name).strip() != "":
                    violations.append((v, s, b, c))
                    
            violations.sort(key=lambda x: x[2].name)
                
            if not violations:
                flash(f"Tuyệt vời! Trong {week_name} không có cá nhân nào bị ghi tên vi phạm vào Sổ đen.", "success")
                return redirect(url_for('weekly', week=week_name))
                
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"So_Den_{week_name}"
            
            ws['A1'] = "ĐOÀN TRƯỜNG THPT THANH HÒA"
            ws['A1'].font = Font(name="Times New Roman", size=11, bold=True)
            ws['A3'] = f"TRÍCH LỤC SỔ ĐEN (CÁ NHÂN VI PHẠM) - {week_name.upper()}"
            ws['A3'].font = Font(name="Times New Roman", size=14, bold=True)
            ws['A3'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A3:E3')
            
            headers = ["STT", "Chi đoàn", "Họ và Tên Học Sinh", "Lỗi Vi Phạm", "Số Lần"]
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=5, column=col, value=h)
                c.font = Font(name="Times New Roman", size=11, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
                
            for idx, (v, s, b, c) in enumerate(violations, 1):
                row_idx = idx + 5
                c1 = ws.cell(row=row_idx, column=1, value=idx)
                c2 = ws.cell(row=row_idx, column=2, value=b.name)
                c3 = ws.cell(row=row_idx, column=3, value=v.student_name)
                c4 = ws.cell(row=row_idx, column=4, value=c.name)
                c5 = ws.cell(row=row_idx, column=5, value=v.quantity)
                
                for cell in [c1, c2, c3, c4, c5]:
                    cell.font = Font(name="Times New Roman", size=11)
                    cell.border = border
                c1.alignment = Alignment(horizontal="center")
                c2.alignment = Alignment(horizontal="center")
                c5.alignment = Alignment(horizontal="center")
                
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            
            log_system_action("XUẤT EXCEL", f"Xuất Sổ đen Vi phạm {week_name}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            
            return send_file(out, download_name=f"So_Den_{week_name.replace(' ', '_')}.xlsx", as_attachment=True)
    except Exception as e:
        import traceback
        print(traceback.format_exc()) 
        flash(f"Lỗi xuất sổ đen: {str(e)}", "error")
        return redirect(url_for('weekly'))

# ==========================================
# MODULE: QUẢN LÝ NGÂN HÀNG LỖI
# ==========================================
@app.route('/violation-categories', methods=['GET'])
def violation_categories():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Cần kích hoạt Năm học trước khi quản lý Thư viện lỗi!", "error")
                return redirect(url_for('dashboard'))
                
            categories = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).order_by(ViolationCategory.name).all()
            return render_template('violation_categories.html', categories=categories, active_year=active_year)
    except Exception as e:
        flash(f"Lỗi truy cập thư viện lỗi: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/add_violation', methods=['POST'])
def add_violation():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            name = request.form.get('name', '').strip()
            try: points = float(request.form.get('penalty_points', 1.0))
            except ValueError: points = 1.0
            point_type = request.form.get('point_type', 'Điểm trừ')

            if name and active_year:
                exist = db_session.query(ViolationCategory).filter_by(name=name, school_year_id=active_year.id).first()
                if exist:
                    flash(f"Lỗi '{name}' đã tồn tại trong thư viện của năm học này!", "error")
                else:
                    new_cat = ViolationCategory(
                        name=name, penalty_points=points, 
                        point_type=point_type, school_year_id=active_year.id 
                    )
                    db_session.add(new_cat)
                    log_system_action("NGÂN HÀNG LỖI", f"Thêm quy định mới: {name} ({points} điểm)")
                    flash(f"Đã thêm lỗi '{name}' vào thư viện của {active_year.name}!", "success")
    except Exception as e:
        flash(f"Lỗi: {str(e)}", "error")
    return redirect(url_for('violation_categories'))

@app.route('/edit_violation/<int:id>', methods=['POST'])
def edit_violation(id):
    try:
        with session_scope() as db_session:
            cat = db_session.query(ViolationCategory).filter_by(id=id).first()
            if cat:
                cat.name = request.form.get('edit_name', cat.name).strip()
                try: cat.penalty_points = float(request.form.get('edit_penalty_points', cat.penalty_points))
                except: pass
                cat.point_type = request.form.get('edit_point_type', cat.point_type)
                log_system_action("NGÂN HÀNG LỖI", f"Cập nhật quy định lỗi: {cat.name}")
                flash(f"Đã cập nhật quy định cho lỗi '{cat.name}'!", "success")
    except Exception as e:
        flash(f"Lỗi cập nhật: {e}", "error")
    return redirect(url_for('violation_categories'))

@app.route('/delete_violation/<int:id>', methods=['POST'])
def delete_violation(id):
    try:
        with session_scope() as db_session:
            cat = db_session.query(ViolationCategory).filter_by(id=id).first()
            if cat:
                name = cat.name
                db_session.delete(cat)
                log_system_action("NGÂN HÀNG LỖI", f"Xóa quy định lỗi: {name}")
                flash(f"Đã xóa vĩnh viễn '{name}' khỏi hệ thống!", "success")
    except Exception as e:
        flash(f"Lỗi xóa lỗi: {e}", "error")
    return redirect(url_for('violation_categories'))

@app.route('/import_violations', methods=['POST'])
def import_violations():
    if 'excel_file' not in request.files:
        flash("Không tìm thấy file tải lên!", "error")
        return redirect(url_for('violation_categories'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash("Chưa chọn file nào!", "error")
        return redirect(url_for('violation_categories'))
        
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip().str.upper() 
            
            with session_scope() as db_session:
                active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
                if not active_year:
                    flash("Chưa có năm học nào được kích hoạt!", "error")
                    return redirect(url_for('violation_categories'))
                    
                name_col = next((c for c in df.columns if "TÊN" in c or "LỖI" in c or "NỘI DUNG" in c), None)
                point_col = next((c for c in df.columns if "ĐIỂM" in c), None)
                type_col = next((c for c in df.columns if "LOẠI" in c or "CỘNG/TRỪ" in c), None)

                if not name_col or not point_col:
                    flash("Lỗi cấu trúc: File Excel cần có ít nhất cột 'Tên lỗi' và 'Số điểm'!", "error")
                    return redirect(url_for('violation_categories'))

                count_new = 0
                for index, row in df.iterrows():
                    name = str(row[name_col]).strip()
                    if not name or name.lower() == 'nan': continue

                    try: points = float(row[point_col])
                    except: points = 1.0

                    point_type = "Điểm trừ"
                    if type_col and pd.notna(row[type_col]):
                        val = str(row[type_col]).strip().lower()
                        if "cộng" in val or "+" in val:
                            point_type = "Điểm cộng"

                    exist = db_session.query(ViolationCategory).filter_by(name=name, school_year_id=active_year.id).first()
                    if not exist:
                        new_cat = ViolationCategory(
                            name=name, penalty_points=points, 
                            point_type=point_type, school_year_id=active_year.id
                        )
                        db_session.add(new_cat)
                        count_new += 1
                        
                log_system_action("NGÂN HÀNG LỖI", f"Đã nhập thành công {count_new} Quy định mới từ Excel")
                flash(f"Đã nhập thành công {count_new} Quy định mới từ file Excel!", "success")
        except Exception as e:
            flash(f"Lỗi đọc file Excel: {str(e)}", "error")
    else:
        flash("Vui lòng chọn file Excel hợp lệ (.xlsx, .xls)", "error")
        
    return redirect(url_for('violation_categories'))

# ==========================================
# MODULE: THI ĐUA THÁNG
# ==========================================
@app.route('/monthly', methods=['GET', 'POST'])
def monthly():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            available_weeks = []
            available_months = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
            
            if active_year:
                weeks_db = db_session.query(WeeklyScore.week).join(Branch).filter(Branch.school_year_id == active_year.id).distinct().all()
                available_weeks = sorted([w[0] for w in weeks_db])

            monthly_data = {}
            selected_month = request.form.get('month') or request.args.get('month') or "Tháng 9"
            selected_weeks = request.form.getlist('weeks')

            used_by_other_months = set()
            current_month_weeks = set()
            special_periods = ["Thi đua chào mừng 20/11", "Thi đua chào mừng 26/3"]

            if active_year:
                all_records = db_session.query(MonthlyRecord).filter(
                    MonthlyRecord.school_year_id == active_year.id
                ).all()

                for rec in all_records:
                    if rec.weeks_used:
                        w_list = [w.strip() for w in rec.weeks_used.split(",") if w.strip()]
                        if rec.month_name == selected_month:
                            for w in w_list: current_month_weeks.add(w)
                        elif rec.month_name not in special_periods:
                            for w in w_list: used_by_other_months.add(w)

            if request.method == 'POST':
                if not selected_weeks:
                    flash("Vui lòng chọn ít nhất 1 tuần để tổng hợp điểm!", "error")
                elif active_year:
                    branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                    
                    temp_groups = {}
                    for b in branches:
                        grp = b.group or "Nhóm 1"
                        if grp not in temp_groups:
                            temp_groups[grp] = []
                            
                        scores = db_session.query(WeeklyScore).filter(
                            WeeklyScore.branch_id == b.id,
                            WeeklyScore.week.in_(selected_weeks)
                        ).all()
                        
                        total_score = sum([(s.total_score or 0.0) for s in scores])
                        week_scores_dict = {s.week: (s.total_score or 0.0) for s in scores}
                        
                        temp_groups[grp].append({
                            'branch_id': b.id,
                            'branch_name': b.name,
                            'group': grp,
                            'gvcn': b.gvcn,
                            'total_score': total_score,
                            'weeks_count': len(scores),
                            'week_scores': week_scores_dict 
                        })
                    
                    for grp, lst in temp_groups.items():
                        lst.sort(key=lambda x: x['total_score'], reverse=True)
                        current_rank = 1
                        for i, d in enumerate(lst):
                            if i > 0 and d['total_score'] < lst[i-1]['total_score']:
                                current_rank = i + 1
                            d['rank'] = current_rank
                        monthly_data[grp] = lst
                        
                    db_session.query(MonthlyRecord).filter(
                        MonthlyRecord.school_year_id == active_year.id, 
                        MonthlyRecord.month_name == selected_month
                    ).delete()
                    
                    w_str = ", ".join(selected_weeks)
                    for grp, lst in monthly_data.items():
                        for d in lst: 
                            db_session.add(MonthlyRecord(
                                school_year_id=active_year.id, 
                                month_name=selected_month, 
                                branch_id=d['branch_id'], 
                                total_score=d['total_score'], 
                                rank=d['rank'], 
                                weeks_used=w_str
                            ))
                    db_session.commit()
                    
                    log_system_action("LƯU ĐIỂM THÁNG", f"Đã tính toán và chốt sổ điểm {selected_month} (gộp từ: {w_str}).")
                    flash(f"✅ Đã lưu thành công dữ liệu xếp hạng {selected_month}!", "success")
            else:
                selected_weeks = list(current_month_weeks)

            user_role = session.get('role', '')
            is_admin = (user_role and ("Quản trị" in user_role or "Bí thư" in user_role or "Admin" in user_role))

            return render_template(
                'monthly.html', 
                available_weeks=available_weeks, 
                available_months=available_months,
                monthly_data=monthly_data,
                selected_month=selected_month, 
                selected_weeks=selected_weeks,
                used_by_other_months=used_by_other_months,
                is_admin=is_admin,
                active_year=active_year
            )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi phân hệ thi đua tháng: {e}", "error")
        return redirect(url_for('dashboard'))
    
# =========================================================================
# THUẬT TOÁN LÕI: TÍNH ĐIỂM TỐT CHUẨN XÁC TỪ BẢNG RAW_SCORES (BẢNG VÀNG)
# =========================================================================
def calculate_trimmed_details(raw_scores, branch_group, max_mon, max_tot):
    """Tính lại chi tiết số lượng điểm 10, 9, 8 dựa trên dữ liệu thô và barem động"""
    f10, f9, f8 = 0, 0, 0
    for r in raw_scores:
        try: c10, c9, c8 = int(r.c10 or 0), int(r.c9 or 0), int(r.c8 or 0)
        except ValueError: continue

        if "1" in str(branch_group): c8 = 0

        k10 = min(c10, max_mon)
        k9 = min(c9, max_mon - k10)
        k8 = min(c8, max_mon - k10 - k9)
        f10 += k10; f9 += k9; f8 += k8

    total = f10 + f9 + f8
    if total > max_tot:
        excess = total - max_tot
        cut_8 = min(f8, excess)
        f8 -= cut_8; excess -= cut_8
        if excess > 0:
            cut_9 = min(f9, excess)
            f9 -= cut_9; excess -= cut_9
        if excess > 0:
            f10 -= excess
    return f10, f9, f8

# ==========================================
# MODULE: BÁO CÁO & THỐNG KÊ
# ==========================================
@app.route('/reports', methods=['GET', 'POST'])
def reports():
    try:
        with session_scope() as db_session:
            years = db_session.query(SchoolYear).order_by(SchoolYear.id.desc()).all()
            active_year = next((y for y in years if y.is_active), None)

            selected_year_id = request.args.get('year_id', type=int)
            if not selected_year_id and active_year: 
                selected_year_id = active_year.id

            selected_month_filter = request.args.get('time_filter', 'Tất cả')
            active_tab = request.args.get('active_tab', 'tab-dashboard')

            chart_data = {
                "bar_labels": [], "bar_values": [], "bar_colors": [],
                "line_labels": [], "line_values": [],
                "pie_bad_labels": [], "pie_bad_values": [],
                "pie_good_labels": [], "pie_good_values": []
            }
            analysis_general = []
            analysis_details = []
            available_months = []

            if selected_year_id:
                months_db = db_session.query(MonthlyRecord.month_name).filter(
                    MonthlyRecord.school_year_id == selected_year_id,
                    MonthlyRecord.month_name.like('Tháng%')
                ).distinct().all()
                school_order = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
                available_months = sorted([m[0] for m in months_db if m[0]], key=lambda x: school_order.index(x) if x in school_order else 99)

                scores_raw = db_session.query(
                    Branch.name.label('Lớp'), Branch.group.label('Nhóm'),
                    WeeklyScore.week.label('Tuần'), WeeklyScore.total_score.label('Điểm'),
                    WeeklyScore.count_8.label('C8'), WeeklyScore.count_9.label('C9'), WeeklyScore.count_10.label('C10')
                ).join(Branch).filter(Branch.school_year_id == selected_year_id)
                
                vios_raw = db_session.query(
                    Branch.name.label('Lớp'), ViolationCategory.name.label('Lỗi'), 
                    WeeklyViolation.quantity.label('SL'), WeeklyScore.week.label('Tuần')
                ).join(WeeklyViolation, ViolationCategory.id == WeeklyViolation.violation_id)\
                 .join(WeeklyScore, WeeklyScore.id == WeeklyViolation.weekly_score_id)\
                 .join(Branch, Branch.id == WeeklyScore.branch_id)\
                 .filter(Branch.school_year_id == selected_year_id)

                if selected_month_filter != 'Tất cả' and selected_month_filter in available_months:
                    month_rec = db_session.query(MonthlyRecord.weeks_used).filter(
                        MonthlyRecord.school_year_id == selected_year_id,
                        MonthlyRecord.month_name == selected_month_filter
                    ).first()
                    if month_rec and month_rec.weeks_used:
                        valid_weeks = [w.strip() for w in month_rec.weeks_used.split(',') if w.strip()]
                        scores_raw = scores_raw.filter(WeeklyScore.week.in_(valid_weeks))
                        vios_raw = vios_raw.filter(WeeklyScore.week.in_(valid_weeks))
                    else:
                        scores_raw = scores_raw.filter(WeeklyScore.week == 'NONE')

                df = pd.DataFrame(scores_raw.all())
                df_vios = pd.DataFrame(vios_raw.all())

                if not df.empty:
                    df['Điểm'] = pd.to_numeric(df['Điểm'], errors='coerce').fillna(0)
                    for col in ['C8', 'C9', 'C10']:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                        
                    def calc_tot_qty(row):
                        nhom = str(row['Nhóm'])
                        if "1" in nhom: return row['C9'] + row['C10']
                        elif "2" in nhom: return row['C8'] + row['C9'] + row['C10']
                        return row['C9'] + row['C10']
                    df['Điểm Tốt'] = df.apply(calc_tot_qty, axis=1)

                    sum_scores = df.groupby('Lớp')['Điểm'].sum().sort_values(ascending=False)
                    if len(sum_scores) >= 6:
                        top_bottom = pd.concat([sum_scores.head(3), sum_scores.tail(3)])
                        colors = ['#28a745']*3 + ['#dc3545']*3
                    else:
                        top_bottom = sum_scores
                        colors = ['#17a2b8'] * len(sum_scores)
                        
                    chart_data['bar_labels'] = top_bottom.index.tolist()
                    chart_data['bar_values'] = top_bottom.values.round(1).tolist()
                    chart_data['bar_colors'] = colors

                    trend_df = df.groupby('Tuần')['Điểm'].mean().reset_index()
                    trend_df['Tuần_Num'] = trend_df['Tuần'].str.extract(r'(\d+)').astype(float)
                    trend_df = trend_df.sort_values(by='Tuần_Num')
                    chart_data['line_labels'] = [str(w).replace("Tuần ", "T.") for w in trend_df['Tuần']]
                    chart_data['line_values'] = trend_df['Điểm'].round(1).tolist()

                    good_counts = df.groupby('Lớp')['Điểm Tốt'].sum().sort_values(ascending=False)
                    good_counts = good_counts[good_counts > 0]
                    if not good_counts.empty:
                        if len(good_counts) > 5:
                            top_good = good_counts.head(4)
                            top_good['Các lớp khác'] = good_counts.iloc[4:].sum()
                            good_counts = top_good
                        chart_data['pie_good_labels'] = good_counts.index.tolist()
                        chart_data['pie_good_values'] = good_counts.values.tolist()

                    df['Khối'] = df['Lớp'].str.extract(r'(\d+)')
                    khoi_sum_avg = df.groupby(['Khối', 'Lớp'])['Điểm'].sum().groupby('Khối').mean()
                    if not khoi_sum_avg.empty:
                        analysis_general.append(("Khối dẫn đầu phong trào", f"Khối {khoi_sum_avg.idxmax()} ({khoi_sum_avg.max():.1f}đ/lớp)", "Tuyên dương tinh thần khối"))
                    
                    if not sum_scores.empty:
                        analysis_general.append(("Chi đoàn Xuất sắc nhất", f"{sum_scores.idxmax()} ({int(sum_scores.max())}đ)", "Đề xuất biểu dương"))
                        
                    sum_by_group = df.groupby(['Nhóm', 'Lớp'])['Điểm'].sum().reset_index()
                    for gk, gn in [('1', 'Nhóm 1'), ('2', 'Nhóm 2')]:
                        g_df = sum_by_group[sum_by_group['Nhóm'].astype(str).str.contains(gk, na=False)]
                        if not g_df.empty:
                            best = g_df.loc[g_df['Điểm'].idxmax()]
                            analysis_general.append((f"Chi đoàn Tốt nhất ({gn})", f"{best['Lớp']} ({int(best['Điểm'])}đ)", "Khen thưởng theo nhóm"))
                            worst = g_df.loc[g_df['Điểm'].idxmin()]
                            analysis_general.append((f"Chi đoàn Yếu kém nhất ({gn})", f"{worst['Lớp']} ({int(worst['Điểm'])}đ)", "Nhắc nhở, đôn đốc chấn chỉnh"))

                    top_3_tot = good_counts.head(3)
                    if not top_3_tot.empty:
                        analysis_general.append(("Nhiều Điểm Tốt nhất", " | ".join([f"{cls} ({int(score)} lượt)" for cls, score in top_3_tot.items() if cls != 'Các lớp khác']), "Khích lệ phong trào"))
                        
                    overall_avg = df['Điểm'].mean()
                    trend = "Tốt" if overall_avg >= 85 else "Khá" if overall_avg >= 70 else "Thấp"
                    analysis_general.append(("Phong độ Toàn trường (ĐTB)", f"{overall_avg:.1f} điểm", f"Đánh giá: {trend}"))

                if not df_vios.empty:
                    df_vios['SL'] = pd.to_numeric(df_vios['SL'], errors='coerce').fillna(0)
                    bad_counts = df_vios.groupby('Lỗi')['SL'].sum().sort_values(ascending=False)
                    if not bad_counts.empty:
                        if len(bad_counts) > 5:
                            top_bad = bad_counts.head(4)
                            top_bad['Các lỗi khác'] = bad_counts.iloc[4:].sum()
                            bad_counts = top_bad
                        chart_data['pie_bad_labels'] = bad_counts.index.tolist()
                        chart_data['pie_bad_values'] = bad_counts.values.tolist()

                    class_error_summary = df_vios.groupby(['Lớp', 'Lỗi'])['SL'].sum().reset_index()
                    error_dict = {}; count_dict = {}
                    for _, row in class_error_summary.iterrows():
                        cls, err, qty = row['Lớp'], row['Lỗi'], int(row['SL'])
                        if cls not in error_dict: error_dict[cls] = []; count_dict[cls] = 0
                        error_dict[cls].append(f"{err} (x{qty})")
                        count_dict[cls] += qty

                    for cls, total_qty in sorted(count_dict.items(), key=lambda x: x[1], reverse=True):
                        analysis_details.append({"class": cls, "qty": total_qty, "details": " | ".join(error_dict[cls])})

            # --- 2. LOGIC BẢNG VÀNG HỌC TẬP ---
            honor_time = request.args.get('honor_time', 'Cả năm')
            honor_grade = request.args.get('honor_grade', 'Tất cả các khối')
            honor_data = []
            time_options_honor = ["Cả năm"]

            if selected_year_id:
                time_options_honor.extend(available_months)
                if len(time_options_honor) == 1: 
                    time_options_honor.extend(["Tháng 9", "Học kỳ 1", "Học kỳ 2"])

                max_tot, max_mon = 14, 4
                try:
                    settings = db_session.query(ScoreSettings).filter_by(school_year_id=selected_year_id).first()
                    if settings:
                        max_tot = int(getattr(settings, 'max_diem_tot', 14))
                        max_mon = int(getattr(settings, 'max_diem_mon', 4))
                except: pass

                valid_weeks_honor = []
                if honor_time == "Cả năm":
                    valid_weeks_honor = [w[0] for w in db_session.query(WeeklyScore.week).join(Branch).filter(Branch.school_year_id == selected_year_id).distinct().all()]
                elif "Học kì 1" in honor_time or "Học kỳ 1" in honor_time:
                    valid_weeks_honor = [f"Tuần {i}" for i in range(1, 19)]
                elif "Học kì 2" in honor_time or "Học kỳ 2" in honor_time:
                    valid_weeks_honor = [f"Tuần {i}" for i in range(19, 38)]
                else:
                    m_rec = db_session.query(MonthlyRecord).filter(MonthlyRecord.school_year_id == selected_year_id, MonthlyRecord.month_name == honor_time).first()
                    if m_rec and m_rec.weeks_used:
                        valid_weeks_honor = [w.strip() for w in m_rec.weeks_used.split(',') if w.strip()]
                
                if not valid_weeks_honor: 
                    valid_weeks_honor = ["_NO_DATA_"]

                branches_data = {}
                for b in db_session.query(Branch).filter(Branch.school_year_id == selected_year_id).all():
                    match = re.search(r'\d+', b.name)
                    branches_data[b.name] = {
                        "name": b.name, "khoi": f"Khối {match.group()}" if match else "Khác",
                        "c8": 0, "c9": 0, "c10": 0, "tuan_tot": 0, "total": 0, "group": str(b.group or "1")
                    }

                scores = db_session.query(WeeklyScore).join(Branch).filter(Branch.school_year_id == selected_year_id, WeeklyScore.week.in_(valid_weeks_honor)).all()
                for s in scores:
                    b_name = s.branch.name
                    b_group = branches_data[b_name]["group"]
                    
                    raw_scores = db_session.query(RawScore).filter_by(week=s.week, branch_name=b_name).all()
                    if raw_scores:
                        f10, f9, f8 = calculate_trimmed_details(raw_scores, b_group, max_mon, max_tot)
                    else:
                        f10, f9, f8 = int(s.count_10 or 0), int(s.count_9 or 0), int(s.count_8 or 0)
                        if "1" in b_group: f8 = 0
                    
                    is_tot = 1 if s.week_rating and 'Tốt' in s.week_rating else 0
                    branches_data[b_name]["c10"] += f10
                    branches_data[b_name]["c9"] += f9
                    branches_data[b_name]["c8"] += f8
                    branches_data[b_name]["tuan_tot"] += is_tot
                    branches_data[b_name]["total"] += (f10 + f9 + f8)

                for d in branches_data.values():
                    if honor_grade == "Tất cả các khối" or d["khoi"] == honor_grade:
                        honor_data.append(d)
                honor_data.sort(key=lambda x: (x["total"], x["tuan_tot"], x["c10"], x["c9"], x["c8"]), reverse=True)

            return render_template(
                'reports.html', 
                years=years, 
                selected_year_id=selected_year_id,
                available_months=available_months,
                selected_month_filter=selected_month_filter,
                chart_data_json=json.dumps(chart_data),
                analysis_general=analysis_general,
                analysis_details=analysis_details,
                honor_data=honor_data,
                honor_time=honor_time,
                honor_grade=honor_grade,
                time_options_honor=time_options_honor,
                active_tab=active_tab
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải báo cáo: {e}", "error")
        return redirect(url_for('dashboard'))

@app.route('/export_academic_honor', methods=['POST'])
def export_academic_honor():
    try:
        time_filter = request.form.get('time_filter', 'Cả năm')
        table_data_json = request.form.get('table_data')
        data = json.loads(table_data_json) if table_data_json else []

        if not data:
            flash("Không có dữ liệu để xuất!", "error")
            return redirect(url_for('reports', active_tab='tab-academic'))

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bảng Vàng Học Tập"
        title_font = Font(name='Arial', size=14, bold=True, color="C00000")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0054a6", end_color="0054a6", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f"BẢNG VÀNG THÀNH TÍCH HỌC TẬP - {time_filter.upper()}"
        ws['A1'].font = title_font; ws['A1'].alignment = center_align
        
        headers = ["STT", "Tên Chi đoàn", "Khối", "Tổng Điểm 8", "Tổng Điểm 9", "Tổng Điểm 10", "Tổng Tuần Loại Tốt", "TỔNG ĐIỂM TỐT"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
            
        for r, d in enumerate(data):
            row_vals = [r+1, d['name'], d['khoi'], d['c8'], d['c9'], d['c10'], d['tuan_tot'], d['total']]
            for c, val in enumerate(row_vals, 1):
                cell = ws.cell(row=4+r, column=c, value=val)
                cell.alignment = center_align; cell.border = thin_border
                if c in [7, 8]: cell.font = Font(color="C00000", bold=True)
                
        for col, w in zip(['A','B','C','D','E','F','G','H'], [6, 15, 12, 15, 15, 15, 20, 18]):
            ws.column_dimensions[col].width = w
            
        log_system_action("XUẤT EXCEL", f"Xuất Bảng Vàng Học Tập - {time_filter}")
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name=f"Bang_Vang_{time_filter.replace(' ', '_')}.xlsx", as_attachment=True)
    except Exception as e:
        flash(f"Lỗi xuất Excel: {e}", "error")
        return redirect(url_for('reports', active_tab='tab-academic'))

@app.route('/export_chronic_violations', methods=['POST'])
def export_chronic_violations():
    try:
        with session_scope() as db_session:
            year_id = request.form.get('export_year_id', type=int)
            month_name = request.form.get('export_month')
            active_year = db_session.query(SchoolYear).filter_by(id=year_id).first()
            if not active_year or not month_name:
                flash("Thiếu thông tin xuất báo cáo!", "error"); return redirect(url_for('reports'))

            month_rec = db_session.query(MonthlyRecord).filter(MonthlyRecord.school_year_id == year_id, MonthlyRecord.month_name == month_name).first()
            if not month_rec or not month_rec.weeks_used:
                flash(f"Chưa có dữ liệu chốt tháng cho {month_name}!", "error"); return redirect(url_for('reports'))
                
            valid_weeks = [w.strip() for w in month_rec.weeks_used.split(',') if w.strip()]
            branches = db_session.query(Branch).filter(Branch.school_year_id == year_id).all()
            b_ids = [b.id for b in branches]
            
            scores = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(b_ids), WeeklyScore.week.in_(valid_weeks)).all()
            
            score_map = {}
            for s in scores:
                diem_tru = float(s.score_tru or 0.0)
                if diem_tru > 0:
                    if s.branch_id not in score_map: score_map[s.branch_id] = {'tru': 0, 'weeks': 0}
                    score_map[s.branch_id]['tru'] += diem_tru
                    score_map[s.branch_id]['weeks'] += 1

            export_data = []
            for b in branches:
                if b.id in score_map:
                    d = score_map[b.id]
                    if d['weeks'] >= 2 or d['tru'] >= 5:
                        export_data.append([0, b.name, d['weeks'], d['tru'], "Vi phạm nhiều lần/Nghiêm trọng"])

            if not export_data:
                flash(f"Tuyệt vời! Không có lớp nào bị cảnh báo vi phạm trong {month_name}!", "success"); return redirect(url_for('reports'))

            export_data.sort(key=lambda x: x[3], reverse=True)

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cảnh báo Vi phạm"
            ws.merge_cells('A1:E1'); ws['A1'] = f"BÁO CÁO CẢNH BÁO VI PHẠM NỀ NẾP - {month_name.upper()}"; ws['A1'].font = Font(size=14, bold=True); ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A2:E2'); ws['A2'] = f"(Gộp dữ liệu từ: {month_rec.weeks_used})"; ws['A2'].font = Font(italic=True); ws['A2'].alignment = Alignment(horizontal="center")
            
            bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            headers = ["STT", "Lớp", "Số tuần vi phạm", "Tổng điểm bị trừ", "Đánh giá sơ bộ"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col, value=h); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center"); c.border = bd
                
            for idx, row in enumerate(export_data, 1):
                row[0] = idx
                for col, val in enumerate(row, 1):
                    c = ws.cell(row=idx+4, column=col, value=val); c.border = bd
                    if col in [1, 2, 3, 4]: c.alignment = Alignment(horizontal="center")

            ws.column_dimensions['B'].width = 15; ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 18; ws.column_dimensions['E'].width = 35
            
            log_system_action("XUẤT EXCEL", f"Xuất báo cáo Cảnh báo Vi phạm - {month_name}")
            out = io.BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, download_name=f"Canh_Bao_Vi_Pham_{month_name.replace(' ', '_')}.xlsx", as_attachment=True)
    except Exception as e:
        flash(f"Lỗi: {e}", "error"); return redirect(url_for('reports'))

@app.route('/export_monthly_summary', methods=['POST'])
def export_monthly_summary():
    try:
        with session_scope() as db_session:
            year_id = request.form.get('export_year_id', type=int)
            month_name = request.form.get('export_month')
            active_year = db_session.query(SchoolYear).filter_by(id=year_id).first()
            if not active_year or not month_name:
                flash("Thiếu thông tin xuất báo cáo!", "error"); return redirect(url_for('reports'))

            month_rec = db_session.query(MonthlyRecord).filter(MonthlyRecord.school_year_id == year_id, MonthlyRecord.month_name == month_name).first()
            if not month_rec or not month_rec.weeks_used:
                flash(f"Chưa có dữ liệu chốt tháng cho {month_name}!", "error"); return redirect(url_for('reports'))
                
            valid_weeks = [w.strip() for w in month_rec.weeks_used.split(',') if w.strip()]
            branches = db_session.query(Branch).filter(Branch.school_year_id == year_id).all()
            b_ids = [b.id for b in branches]
            
            scores = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(b_ids), WeeklyScore.week.in_(valid_weeks)).all()
            
            score_map = {}
            for s in scores:
                if s.branch_id not in score_map: score_map[s.branch_id] = {'tot': 0, 'tru': 0, 'tong': 0, 'count': 0}
                sl_tot = int(s.count_8 or 0) + int(s.count_9 or 0) + int(s.count_10 or 0)
                score_map[s.branch_id]['tot'] += sl_tot
                score_map[s.branch_id]['tru'] += float(s.score_tru or 0.0)
                score_map[s.branch_id]['tong'] += float(s.total_score or 0.0)
                score_map[s.branch_id]['count'] += 1

            export_data = []
            for b in branches:
                if b.id in score_map:
                    d = score_map[b.id]
                    avg = d['tong'] / d['count'] if d['count'] > 0 else 0
                    export_data.append([0, b.name, b.group, d['tot'], d['tru'], round(avg, 1)])

            if not export_data:
                flash(f"Chưa có điểm số nào!", "error"); return redirect(url_for('reports'))

            export_data.sort(key=lambda x: x[5], reverse=True)

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tổng hợp Thi đua"
            ws.merge_cells('A1:F1'); ws['A1'] = f"TỔNG HỢP KẾT QUẢ THI ĐUA - {month_name.upper()}"; ws['A1'].font = Font(size=14, bold=True); ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A2:F2'); ws['A2'] = f"(Gộp dữ liệu từ: {month_rec.weeks_used})"; ws['A2'].font = Font(italic=True); ws['A2'].alignment = Alignment(horizontal="center")
            
            bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            headers = ["STT", "Lớp", "Nhóm", "Tổng Điểm Tốt", "Tổng Điểm Bị Trừ", "Điểm TB Tháng"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col, value=h); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center"); c.border = bd
                
            for idx, row in enumerate(export_data, 1):
                row[0] = idx
                for col, val in enumerate(row, 1):
                    c = ws.cell(row=idx+4, column=col, value=val); c.border = bd
                    c.alignment = Alignment(horizontal="center")

            ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 18; ws.column_dimensions['E'].width = 18; ws.column_dimensions['F'].width = 18
            
            log_system_action("XUẤT EXCEL", f"Xuất Tổng hợp Kế quả Thi đua - {month_name}")
            out = io.BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, download_name=f"Tong_Hop_Thi_Dua_{month_name.replace(' ', '_')}.xlsx", as_attachment=True)
    except Exception as e:
        flash(f"Lỗi: {e}", "error"); return redirect(url_for('reports'))

# ==========================================
# MODULE: TRA CỨU CHI ĐOÀN (CLASS DASHBOARD)
# ==========================================
@app.route('/class-dashboard', methods=['GET', 'POST'])
def class_dashboard():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            # [NÂNG CẤP LÕI]: Khoanh vùng dữ liệu theo Quyền đăng nhập
            user_role = session.get('role', '')
            session_username = session.get('username', '').strip().upper()
            is_gvcn = (user_role == 'Giáo viên chủ nhiệm')
            
            branches = []
            if active_year:
                if is_gvcn:
                    branches = db_session.query(Branch).filter(Branch.name == session_username, Branch.school_year_id == active_year.id).all()
                else:
                    branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).order_by(Branch.name).all()
            
            selected_branch_id = request.args.get('branch_id', type=int)
            if not selected_branch_id and branches:
                selected_branch_id = branches[0].id
                
            selected_branch = None
            weekly_scores = []
            assignments = []
            monitoring_assignments = []
            warning_students = []
            
            if selected_branch_id:
                if is_gvcn and branches and selected_branch_id != branches[0].id:
                    selected_branch_id = branches[0].id
                    
                selected_branch = db_session.query(Branch).filter_by(id=selected_branch_id).first()
                if selected_branch:
                    group_val = selected_branch.group or "Nhóm 1"
                    weekly_scores_db = db_session.query(WeeklyScore).filter_by(branch_id=selected_branch.id).order_by(WeeklyScore.id).all()
                    
                    # ==========================================================
                    # [THUẬT TOÁN ĐẾM LỖI TỐI ƯU]: Quét học sinh vi phạm >= 3 lỗi 
                    # ==========================================================
                    if weekly_scores_db:
                        # VÁ LỖI: Duyệt ngược từ tuần mới nhất trở về trước
                        for score in reversed(weekly_scores_db):
                            student_viol_counts = {}
                            for viol in score.violations:
                                if viol.student_name and str(viol.student_name).strip() != "":
                                    raw_names = str(viol.student_name).replace(';', ',').split(',')
                                    names = [n.strip().upper() for n in raw_names if n.strip()]
                                    for name in names:
                                        qty = int(viol.quantity) if viol.quantity else 1
                                        student_viol_counts[name] = student_viol_counts.get(name, 0) + qty
                            
                            current_warnings = [{'name': name.title(), 'count': count, 'week': score.week} 
                                                for name, count in student_viol_counts.items() if count >= 3]
                            if current_warnings:
                                warning_students = sorted(current_warnings, key=lambda x: x['count'], reverse=True)
                                break
                    # ==========================================================
                    
                    for sc in weekly_scores_db:
                        all_in_week = db_session.query(WeeklyScore).join(Branch).filter(
                            WeeklyScore.week == sc.week,
                            Branch.school_year_id == active_year.id
                        ).all()
                        
                        same_group_scores = [s for s in all_in_week if (s.branch.group or "Nhóm 1") == group_val]
                        same_group_scores.sort(key=lambda x: float(x.total_score or 0), reverse=True)
                        
                        rk = 1
                        for i, s in enumerate(same_group_scores):
                            if i > 0 and float(s.total_score or 0) < float(same_group_scores[i-1].total_score or 0): rk = i + 1
                            if s.branch_id == selected_branch.id: break
                        
                        so_luong_diem_tot = int(sc.count_9 or 0) + int(sc.count_10 or 0)
                        if "2" in str(group_val): so_luong_diem_tot += int(sc.count_8 or 0)
                            
                        weekly_scores.append({
                            'score_id': sc.id,             
                            'is_locked': getattr(sc, 'is_locked', False),     
                            'is_appealed': getattr(sc, 'is_appealed', False), 
                            'appeal_reason': getattr(sc, 'appeal_reason', ''),
                            'appeal_response': getattr(sc, 'appeal_response', ''),
                            'week': sc.week,
                            'rating': sc.week_rating or "-",
                            'score_truc': sc.score_truc or 100,
                            'diem_tot': so_luong_diem_tot,
                            'score_cong': sc.score_cong or 0,
                            'score_tru': sc.score_tru or 0,
                            'total_score': sc.total_score or 0,
                            'note': sc.note or "",
                            'rank': rk,
                            'is_appeal_expired': getattr(sc, 'is_appeal_expired', False)
                        })
                    
                    red_star_ids = [rs.id for rs in selected_branch.red_stars] if hasattr(selected_branch, 'red_stars') else []
                    if not red_star_ids: red_star_ids = [rs.id for rs in db_session.query(RedStar).filter_by(branch_id=selected_branch.id).all()]
                    if red_star_ids: assignments = db_session.query(Assignment).filter(Assignment.red_star_id.in_(red_star_ids)).order_by(Assignment.week_number.desc(), Assignment.shift).all()

                    import json, os
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    zones_map = {}
                    if os.path.exists(os.path.join(base_dir, "config", "class_zones.json")):
                        with open(os.path.join(base_dir, "config", "class_zones.json"), "r", encoding="utf-8") as f:
                            try: zones_map = json.load(f)
                            except: pass
                    
                    target_zone_names = [zone for zone, classes in zones_map.items() if isinstance(classes, list) and str(selected_branch.name or '').strip().upper() in [str(c).strip().upper() for c in classes]]
                    if target_zone_names: monitoring_assignments = db_session.query(Assignment).join(DutyArea).join(RedStar).join(Branch).filter(DutyArea.name.in_(target_zone_names), Branch.school_year_id == active_year.id).order_by(Assignment.week_number.desc(), Assignment.shift).all()

            return render_template('class_dashboard.html', 
                                   active_year=active_year, 
                                   branches=branches, 
                                   selected_branch=selected_branch, 
                                   weekly_scores=weekly_scores, 
                                   assignments=assignments, 
                                   monitoring_assignments=monitoring_assignments,
                                   warning_students=warning_students,
                                   is_gvcn=is_gvcn 
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tra cứu chi đoàn: {e}", "error")
        return redirect(url_for('dashboard'))
    
# ==========================================
# MODULE: TRANG XEM TRƯỚC HỒ SƠ LỚP MỚI
# ==========================================
@app.route('/preview_class_dashboard/<int:branch_id>')
def preview_class_dashboard(branch_id):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return redirect(url_for('class_dashboard'))

            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            if not branch: return redirect(url_for('class_dashboard'))

            group_val = branch.group or "Nhóm 1"
            weekly_scores = []
            weekly_scores_db = db_session.query(WeeklyScore).filter_by(branch_id=branch.id).order_by(WeeklyScore.id).all()
            for sc in weekly_scores_db:
                all_in_week = db_session.query(WeeklyScore).join(Branch).filter(
                    WeeklyScore.week == sc.week, Branch.school_year_id == active_year.id
                ).all()
                same_group_scores = [s for s in all_in_week if (s.branch.group or "Nhóm 1") == group_val]
                same_group_scores.sort(key=lambda x: float(x.total_score or 0), reverse=True)
                rk = 1
                for i, s in enumerate(same_group_scores):
                    if i > 0 and float(s.total_score or 0) < float(same_group_scores[i-1].total_score or 0): rk = i + 1
                    if s.branch_id == branch.id: break
                
                # [ĐÃ NÂNG CẤP LẠI]: Tính Tổng số con điểm Tốt (Không nhân hệ số)
                so_luong_diem_tot = int(sc.count_9 or 0) + int(sc.count_10 or 0)
                if "2" in str(group_val): so_luong_diem_tot += int(sc.count_8 or 0)
                
                weekly_scores.append({
                    'week': sc.week, 'rating': sc.week_rating or "-", 'score_truc': sc.score_truc or 100,
                    'diem_tot': so_luong_diem_tot, 'score_cong': sc.score_cong or 0, 'score_tru': sc.score_tru or 0,
                    'total_score': sc.total_score or 0, 'note': sc.note or "", 'rank': rk
                })
            
            red_star_ids = [rs.id for rs in db_session.query(RedStar).filter_by(branch_id=branch.id).all()]
            assignments = []
            if red_star_ids:
                assignments = db_session.query(Assignment).filter(Assignment.red_star_id.in_(red_star_ids)).order_by(Assignment.week_number.desc(), Assignment.shift).all()

            import json, os
            base_dir = os.path.dirname(os.path.abspath(__file__))
            zones_map = {}
            monitoring_assignments = []
            if os.path.exists(os.path.join(base_dir, "config", "class_zones.json")):
                with open(os.path.join(base_dir, "config", "class_zones.json"), "r", encoding="utf-8") as f:
                    try: zones_map = json.load(f)
                    except: pass
                    
            target_zone_names = []
            current_branch_clean = str(branch.name or '').strip().upper()
            for zone, classes in zones_map.items():
                if isinstance(classes, list):
                    if current_branch_clean in [str(c).strip().upper() for c in classes]:
                        target_zone_names.append(zone)
                        
            if target_zone_names:
                monitoring_assignments = db_session.query(Assignment).join(DutyArea).join(RedStar).join(Branch).filter(
                    DutyArea.name.in_(target_zone_names), Branch.school_year_id == active_year.id
                ).order_by(Assignment.week_number.desc(), Assignment.shift).all()
                
            return render_template('preview_class_dashboard.html', branch=branch, weekly_scores=weekly_scores, assignments=assignments, monitoring_assignments=monitoring_assignments)
    except Exception as e:
        flash(f"Lỗi xem trước: {e}", "error")
        return redirect(url_for('class_dashboard'))

# ==========================================
# MODULE: XUẤT EXCEL HỒ SƠ CHI ĐOÀN
# ==========================================
@app.route('/export_class_dashboard/<int:branch_id>')
def export_class_dashboard(branch_id):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return redirect(url_for('class_dashboard'))

            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            if not branch: return redirect(url_for('class_dashboard'))

            group_val = branch.group or "Nhóm 1"
            weekly_scores = []
            weekly_scores_db = db_session.query(WeeklyScore).filter_by(branch_id=branch.id).order_by(WeeklyScore.id).all()
            for sc in weekly_scores_db:
                all_in_week = db_session.query(WeeklyScore).join(Branch).filter(WeeklyScore.week == sc.week, Branch.school_year_id == active_year.id).all()
                same_group_scores = [s for s in all_in_week if (s.branch.group or "Nhóm 1") == group_val]
                same_group_scores.sort(key=lambda x: float(x.total_score or 0), reverse=True)
                rk = 1
                for i, s in enumerate(same_group_scores):
                    if i > 0 and float(s.total_score or 0) < float(same_group_scores[i-1].total_score or 0): rk = i + 1
                    if s.branch_id == branch.id: break
                
                # [ĐÃ NÂNG CẤP LẠI]: Tính Tổng số con điểm Tốt (Không nhân hệ số)
                so_luong_diem_tot = int(sc.count_9 or 0) + int(sc.count_10 or 0)
                if "2" in str(group_val): so_luong_diem_tot += int(sc.count_8 or 0)
                
                weekly_scores.append({
                    'week': sc.week, 'rating': sc.week_rating or "-", 'score_truc': sc.score_truc or 100,
                    'diem_tot': so_luong_diem_tot, 'note': sc.note or "", 'total_score': sc.total_score or 0, 'rank': rk
                })
            
            red_star_ids = [rs.id for rs in db_session.query(RedStar).filter_by(branch_id=branch.id).all()]
            assignments = db_session.query(Assignment).filter(Assignment.red_star_id.in_(red_star_ids)).order_by(Assignment.week_number.desc(), Assignment.shift).all() if red_star_ids else []

            import json, os
            base_dir = os.path.dirname(os.path.abspath(__file__))
            zones_map = {}
            monitoring_assignments = []
            if os.path.exists(os.path.join(base_dir, "config", "class_zones.json")):
                with open(os.path.join(base_dir, "config", "class_zones.json"), "r", encoding="utf-8") as f:
                    try: zones_map = json.load(f)
                    except: pass
            target_zone_names = [zone for zone, classes in zones_map.items() if str(branch.name or '').strip().upper() in [str(c).strip().upper() for c in classes]] if zones_map else []
            if target_zone_names:
                monitoring_assignments = db_session.query(Assignment).join(DutyArea).join(RedStar).join(Branch).filter(
                    DutyArea.name.in_(target_zone_names), Branch.school_year_id == active_year.id
                ).order_by(Assignment.week_number.desc(), Assignment.shift).all()

            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            import io
            from flask import send_file

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Ho_So_{branch.name}"

            font_title = Font(name="Times New Roman", size=14, bold=True)
            font_header = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
            font_normal = Font(name="Times New Roman", size=12)
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

            ws.merge_cells('A1:G1')
            ws['A1'] = f"HỒ SƠ TRA CỨU CHI ĐOÀN {branch.name.upper()}"
            ws['A1'].font = font_title; ws['A1'].alignment = align_center

            ws.merge_cells('A2:G2')
            ws['A2'] = f"GVCN: {branch.gvcn or '....................'} | Sĩ số: {branch.si_so} | Nhóm: {branch.group or '1'}"
            ws['A2'].font = Font(name="Times New Roman", size=12, italic=True); ws['A2'].alignment = align_center

            current_row = 4

            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "I. KẾT QUẢ THI ĐUA TỪNG TUẦN"
            ws[f'A{current_row}'].font = Font(name="Times New Roman", size=13, bold=True, color="C00000")
            current_row += 1

            # Đã đổi Header: Điểm Tốt -> SL Điểm Tốt
            headers_1 = ["Tuần", "Tổng điểm", "Hạng", "Xếp loại", "SL Điểm Tốt", "Ghi chú Vi phạm", ""]
            for col_num, h_text in enumerate(headers_1, 1):
                c = ws.cell(row=current_row, column=col_num, value=h_text)
                c.font = font_header; c.alignment = align_center; c.border = border_thin; c.fill = fill_header
            ws.cell(row=current_row, column=7).border = border_thin
            ws.cell(row=current_row, column=7).fill = fill_header
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
            current_row += 1

            if weekly_scores:
                for sc in weekly_scores:
                    row_data = [sc['week'], sc['total_score'], f"Hạng {sc['rank']}", sc['rating'], sc['diem_tot'], sc['note']]
                    for col_num, val in enumerate(row_data, 1):
                        c = ws.cell(row=current_row, column=col_num, value=val)
                        c.font = font_normal; c.border = border_thin
                        c.alignment = Alignment(horizontal="left", vertical="center") if col_num == 6 else align_center
                    ws.cell(row=current_row, column=7).border = border_thin
                    ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
                    current_row += 1
            else:
                ws.merge_cells(f'A{current_row}:G{current_row}')
                ws[f'A{current_row}'] = "Chưa có dữ liệu thi đua"
                ws[f'A{current_row}'].font = font_normal; ws[f'A{current_row}'].alignment = align_center; ws[f'A{current_row}'].border = border_thin
                current_row += 1

            current_row += 2

            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "II. DANH SÁCH SAO ĐỎ ĐÃ TRỰC CHẤM ĐIỂM LỚP"
            ws[f'A{current_row}'].font = Font(name="Times New Roman", size=13, bold=True, color="C00000")
            current_row += 1

            headers_3 = ["Tuần", "Ca trực", "Học sinh trực", "Thuộc Chi đoàn", "Khu vực trực", "Đánh giá KQ", "Ghi chú"]
            for col_num, h_text in enumerate(headers_3, 1):
                c = ws.cell(row=current_row, column=col_num, value=h_text)
                c.font = font_header; c.alignment = align_center; c.border = border_thin; c.fill = fill_header
            current_row += 1

            if monitoring_assignments:
                for assign in monitoring_assignments:
                    row_data = [f"Tuần {assign.week_number}", assign.shift, assign.red_star.full_name if assign.red_star else "-", 
                                assign.red_star.branch.name if assign.red_star and assign.red_star.branch else "-", 
                                assign.duty_area.name if assign.duty_area else "-", "", ""]
                    for col_num, val in enumerate(row_data, 1):
                        c = ws.cell(row=current_row, column=col_num, value=val)
                        c.font = font_normal; c.border = border_thin; c.alignment = align_center
                    current_row += 1
            else:
                ws.merge_cells(f'A{current_row}:G{current_row}')
                ws[f'A{current_row}'] = "Không có thông tin"
                ws[f'A{current_row}'].font = font_normal; ws[f'A{current_row}'].alignment = align_center; ws[f'A{current_row}'].border = border_thin
                current_row += 1

            current_row += 2

            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "III. LỊCH PHÂN CÔNG TRỰC CỦA HỌC SINH LỚP NÀY"
            ws[f'A{current_row}'].font = Font(name="Times New Roman", size=13, bold=True, color="C00000")
            current_row += 1

            headers_2 = ["Tuần", "Ca trực", "Học sinh trực", "Khu vực trực", "Nhiệm vụ", "Đánh giá KQ", "Ghi chú"]
            for col_num, h_text in enumerate(headers_2, 1):
                c = ws.cell(row=current_row, column=col_num, value=h_text)
                c.font = font_header; c.alignment = align_center; c.border = border_thin; c.fill = fill_header
            current_row += 1

            if assignments:
                for assign in assignments:
                    row_data = [f"Tuần {assign.week_number}", assign.shift, assign.red_star.full_name if assign.red_star else "-", 
                                assign.duty_area.name if assign.duty_area else "-", "", "", ""]
                    for col_num, val in enumerate(row_data, 1):
                        c = ws.cell(row=current_row, column=col_num, value=val)
                        c.font = font_normal; c.border = border_thin; c.alignment = align_center
                    current_row += 1
            else:
                ws.merge_cells(f'A{current_row}:G{current_row}')
                ws[f'A{current_row}'] = "Không có lịch phân công"
                ws[f'A{current_row}'].font = font_normal; ws[f'A{current_row}'].alignment = align_center; ws[f'A{current_row}'].border = border_thin
                current_row += 1

            ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 15; ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 30

            log_system_action("XUẤT EXCEL", f"Xuất Hồ sơ tra cứu Chi đoàn {branch.name}")
            out = io.BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, download_name=f"Ho_So_{branch.name.replace(' ', '_')}.xlsx", as_attachment=True)

    except Exception as e:
        import traceback
        traceback.print_exc() 
        flash(f"Lỗi xuất Excel: {e}", "error")
        return redirect(url_for('class_dashboard'))
    
# ==========================================
# API: ĐỌC / GHI ĐIỂM GỐC TỪNG MÔN HỌC (RAWSCORE)
# ==========================================
@app.route('/api/raw_scores/<week_name>/<int:branch_id>', methods=['GET', 'POST'])
def handle_raw_scores(week_name, branch_id):
    try:
        with session_scope() as db_session:
            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            if not branch: 
                return {"error": "Lớp không tồn tại"}, 404

            if request.method == 'GET':
                records = db_session.query(RawScore).filter_by(week=week_name, branch_name=branch.name).all()
                data = [{"subj": r.subject, "c10": r.c10, "c9": r.c9, "c8": r.c8} for r in records]
                return {"data": data}

            if request.method == 'POST':
                raw_list = request.json.get('raw_list', [])
                db_session.query(RawScore).filter_by(week=week_name, branch_name=branch.name).delete()
                
                for item in raw_list:
                    subj = str(item.get("subj", "")).strip()
                    if not subj or subj == "Điểm đã nhập": continue
                    rs = RawScore(
                        week=week_name,
                        branch_name=branch.name,
                        subject=subj,
                        c10=int(item.get("c10", 0)),
                        c9=int(item.get("c9", 0)),
                        c8=int(item.get("c8", 0))
                    )
                    db_session.add(rs)
                return {"status": "success"}
    except Exception as e:
        return {"error": str(e)}, 500

# ==========================================
# MODULE: XEM TRƯỚC EXCEL BÁO CÁO TUẦN
# ==========================================
@app.route('/preview_report')
def preview_report():
    try:
        with session_scope() as db_session:
            week_name = request.args.get('week', 'Tuần 1')
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('weekly'))
            
            current_scores = db_session.query(WeeklyScore).join(Branch).filter(
                WeeklyScore.week == week_name, Branch.school_year_id == active_year.id
            ).all()
            
            try:
                week_num = int(week_name.replace("Tuần ", "").strip())
                prev_week_name = f"Tuần {week_num - 1}"
            except:
                prev_week_name = None
                
            prev_rank_map = {}
            if prev_week_name:
                prev_scores = db_session.query(WeeklyScore).join(Branch).filter(
                    WeeklyScore.week == prev_week_name, Branch.school_year_id == active_year.id
                ).all()
                
                prev_data = {}
                for sc in prev_scores:
                    grp = sc.branch.group or "Nhóm 1"
                    if grp not in prev_data: prev_data[grp] = []
                    prev_data[grp].append(sc)
                for grp, lst in prev_data.items():
                    lst.sort(key=lambda x: float(x.total_score or 0), reverse=True)
                    rk = 1
                    for i, s in enumerate(lst):
                        if i > 0 and float(s.total_score or 0) < float(lst[i-1].total_score or 0):
                            rk = i + 1
                        prev_rank_map[s.branch_id] = rk

            report_data = {}
            start_date_str = ""
            end_date_str = ""
            
            for sc in current_scores:
                if sc.start_date: start_date_str = sc.start_date
                if sc.end_date: end_date_str = sc.end_date
                
                b = sc.branch
                grp = b.group or "Nhóm 1"
                if grp not in report_data: report_data[grp] = []
                
                so_luong_diem_tot = int(sc.count_9 or 0) + int(sc.count_10 or 0)
                if "2" in str(grp): 
                    so_luong_diem_tot += int(sc.count_8 or 0)
                
                report_data[grp].append({
                    'branch_id': b.id,
                    'branch_name': b.name,
                    'total_score': float(sc.total_score or 0),
                    'diem_tot': so_luong_diem_tot,
                    'note': sc.note or "",
                    'prev_rank': prev_rank_map.get(b.id, "N/A"),
                    'current_rank': 0
                })
                
            for grp, lst in report_data.items():
                lst.sort(key=lambda x: x['total_score'], reverse=True)
                rk = 1
                for i, item in enumerate(lst):
                    if i > 0 and item['total_score'] < lst[i-1]['total_score']: rk = i + 1
                    item['current_rank'] = rk

            if start_date_str and "-" in start_date_str:
                try: 
                    p = start_date_str.split('-')
                    start_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass
            if end_date_str and "-" in end_date_str:
                try:
                    p = end_date_str.split('-')
                    end_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass

            # [BẢN VÁ]: Nếu CSDL chưa lưu ngày, tự động gắn dấu chấm để giữ khung văn bản
            if not start_date_str: start_date_str = "......"
            if not end_date_str: end_date_str = "......"

            return render_template('preview_report.html', 
                                   report_data=report_data, 
                                   week_name=week_name, 
                                   start_date=start_date_str,
                                   end_date=end_date_str)
    except Exception as e:
        flash(f"Lỗi xem trước báo cáo: {e}", "error")
        return redirect(url_for('weekly'))

@app.route('/export_weekly_excel')
def export_weekly_excel():
    try:
        with session_scope() as db_session:
            week_name = request.args.get('week', 'Tuần 1')
            active_year = db_session.query(SchoolYear).filter(SchoolYear.is_active == True).first()
            if not active_year: return redirect(url_for('weekly'))
            
            current_scores = db_session.query(WeeklyScore).join(Branch).filter(WeeklyScore.week == week_name, Branch.school_year_id == active_year.id).all()
            try: week_num = int(week_name.replace("Tuần ", "").strip()); prev_week_name = f"Tuần {week_num - 1}"
            except: prev_week_name = None
                
            prev_rank_map = {}
            if prev_week_name:
                prev_scores = db_session.query(WeeklyScore).join(Branch).filter(WeeklyScore.week == prev_week_name, Branch.school_year_id == active_year.id).all()
                prev_data = {}
                for sc in prev_scores:
                    grp = sc.branch.group or "Nhóm 1"
                    if grp not in prev_data: prev_data[grp] = []
                    prev_data[grp].append(sc)
                for grp, lst in prev_data.items():
                    lst.sort(key=lambda x: float(x.total_score or 0), reverse=True)
                    rk = 1
                    for i, s in enumerate(lst):
                        if i > 0 and float(s.total_score or 0) < float(lst[i-1].total_score or 0): rk = i + 1
                        prev_rank_map[s.branch_id] = rk

            report_data = {}; start_date_str = ""; end_date_str = ""
            for sc in current_scores:
                if sc.start_date: start_date_str = sc.start_date
                if sc.end_date: end_date_str = sc.end_date
                b = sc.branch; grp = b.group or "Nhóm 1"
                if grp not in report_data: report_data[grp] = []
                
                so_luong_diem_tot = int(sc.count_9 or 0) + int(sc.count_10 or 0)
                if "2" in str(grp): 
                    so_luong_diem_tot += int(sc.count_8 or 0)
                    
                report_data[grp].append({
                    'branch_name': b.name, 'total_score': float(sc.total_score or 0),
                    'diem_tot': so_luong_diem_tot,
                    'note': sc.note or "", 'prev_rank': prev_rank_map.get(b.id, "N/A")
                })
                
            if start_date_str and "-" in start_date_str:
                try: 
                    p = start_date_str.split('-')
                    start_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass
            if end_date_str and "-" in end_date_str:
                try:
                    p = end_date_str.split('-')
                    end_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass

            # [BẢN VÁ]: Gắn mặc định dấu chấm nếu CSDL không có dữ liệu ngày
            if not start_date_str: start_date_str = "......"
            if not end_date_str: end_date_str = "......"
                
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.worksheet.page import PageMargins
            import io
            from flask import send_file

            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Thi Đua Tuần"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT 
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.5, footer=0.5)
            
            font_title = Font(name="Times New Roman", size=11, bold=True)
            font_main_header = Font(name="Times New Roman", size=14, bold=True)
            font_data = Font(name="Times New Roman", size=11)
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font_header_color = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
            thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
            
            ws.merge_cells("A1:D1"); ws.cell(row=1, column=1, value="ĐOÀN TRƯỜNG THPT THANH HÒA").font = font_title; ws.cell(row=1, column=1).alignment = align_center
            ws.merge_cells("E1:G1"); ws.cell(row=1, column=5, value="ĐOÀN TNCS HỒ CHÍ MINH").font = font_title; ws.cell(row=1, column=5).alignment = align_center
            ws.merge_cells("A2:G2"); ws["A2"] = f"BẢNG TỔNG HỢP KẾT QUẢ THI ĐUA - {week_name.upper()}"; ws["A2"].font = font_main_header; ws["A2"].alignment = align_center
            
            # [BẢN VÁ]: Luôn ghi ra dòng Thời gian, không dùng lệnh IF ẩn đi nữa
            ws.merge_cells("A3:G3")
            ws["A3"] = f"(Từ ngày {start_date_str} đến ngày {end_date_str})"
            ws["A3"].font = Font(name="Times New Roman", size=12, italic=True)
            ws["A3"].alignment = align_center
            
            current_row = 5
            for group in sorted(report_data.keys()):
                group_items = report_data[group]
                group_items.sort(key=lambda x: x['total_score'], reverse=True)
                
                curr_rank = 1
                for i, item in enumerate(group_items):
                    if i > 0 and item['total_score'] < group_items[i-1]['total_score']: curr_rank = i + 1
                    item['current_rank'] = curr_rank

                ws.cell(row=current_row, column=1, value=str(group)).font = Font(name="Times New Roman", size=11, bold=True)
                current_row += 1
                
                headers = ["STT", "Lớp", "X.H Tuần Trước", "Hạng Hiện Tại", "Tổng điểm", "SL Điểm Tốt", "Ghi chú"]
                for col_idx, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                    cell.font = font_header_color; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border
                current_row += 1
                
                for idx, row_data in enumerate(group_items, 1):
                    val_tot = int(row_data["total_score"]) if row_data["total_score"].is_integer() else row_data["total_score"]
                    row_values = [idx, row_data["branch_name"], row_data["prev_rank"], row_data["current_rank"], val_tot, row_data["diem_tot"], row_data["note"]]
                    for col_idx, val in enumerate(row_values, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = font_data; cell.border = thin_border
                        if col_idx <= 6: cell.alignment = align_center
                        else: cell.alignment = align_left
                    current_row += 1
                current_row += 1 
            
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
            ws.cell(row=current_row, column=5, value="TM/BCH ĐOÀN TRƯỜNG").font = font_title; ws.cell(row=current_row, column=5).alignment = align_center
            
            for col_letter, width in {'A': 5.5, 'B': 8.5, 'C': 11.5, 'D': 11.5, 'E': 11.5, 'F': 11.0, 'G': 30.0}.items():
                ws.column_dimensions[col_letter].width = width
            
            log_system_action("XUẤT EXCEL", f"Xuất Báo cáo Điểm {week_name}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name=f"Bao_Cao_Thi_Dua_{week_name.replace(' ', '_')}.xlsx", as_attachment=True)
    except Exception as e:
        flash(f"Lỗi xuất file Excel: {str(e)}", "error")
        return redirect(url_for('weekly'))
# ==========================================
# MODULE: XUẤT EXCEL THI ĐUA THÁNG
# ==========================================
@app.route('/export_monthly_excel', methods=['POST'])
def export_monthly_excel():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học nào được kích hoạt!", "error")
                return redirect(url_for('monthly'))

            selected_month = request.form.get('month', 'Tháng...')
            selected_weeks = request.form.getlist('weeks')

            if not selected_weeks:
                flash("Vui lòng chọn ít nhất 1 tuần để xuất báo cáo!", "error")
                return redirect(url_for('monthly'))

            groups_data = {}
            branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
            
            for b in branches:
                grp = b.group or "Nhóm 1"
                if grp not in groups_data:
                    groups_data[grp] = []
                    
                scores = db_session.query(WeeklyScore).filter(
                    WeeklyScore.branch_id == b.id,
                    WeeklyScore.week.in_(selected_weeks)
                ).all()
                
                week_scores = {s.week: (s.total_score or 0.0) for s in scores}
                total_score = sum(week_scores.values())
                
                groups_data[grp].append({
                    'branch_name': b.name,
                    'week_scores': week_scores,
                    'total_score': total_score,
                    'gvcn': b.gvcn
                })
            
            for grp, lst in groups_data.items():
                lst.sort(key=lambda x: x['total_score'], reverse=True)
                rk = 1
                for i, d in enumerate(lst):
                    if i > 0 and d['total_score'] < lst[i-1]['total_score']:
                        rk = i + 1
                    d['rank'] = rk

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Báo Cáo Thi Đua"

            font_title = Font(name="Times New Roman", size=12, bold=True)
            font_main_header = Font(name="Times New Roman", size=16, bold=True)
            font_header_table = Font(name="Times New Roman", size=12, bold=True)
            font_data = Font(name="Times New Roman", size=12)
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            num_cols = 2 + len(selected_weeks) + 3 
            last_col_letter = get_column_letter(num_cols)

            ws.merge_cells("A1:C1")
            ws["A1"] = "ĐOÀN TRƯỜNG THPT THANH HÒA"
            ws["A1"].font = font_title
            ws["A1"].alignment = align_center

            ws.merge_cells(f"E1:{last_col_letter}1" if num_cols >= 5 else f"D1:{last_col_letter}1")
            ws["E1" if num_cols >= 5 else "D1"] = "ĐOÀN TNCS HỒ CHÍ MINH"
            ws["E1" if num_cols >= 5 else "D1"].font = font_title
            ws["E1" if num_cols >= 5 else "D1"].alignment = align_center

            ws.merge_cells(f"A3:{last_col_letter}3")
            ws["A3"] = f"ĐIỂM THI ĐUA {selected_month.upper()}"
            ws["A3"].font = font_main_header
            ws["A3"].alignment = align_center
            
            week_str = ", ".join(selected_weeks)
            ws.merge_cells(f"A4:{last_col_letter}4")
            ws["A4"] = f"(Tổng hợp điểm từ các tuần: {week_str})"
            ws["A4"].font = Font(name="Times New Roman", size=12, italic=True)
            ws["A4"].alignment = align_center
            
            mid_col = num_cols // 2 if num_cols > 3 else 2
            ws.merge_cells(f"{get_column_letter(mid_col)}5:{get_column_letter(mid_col+2)}5")
            ws[f"{get_column_letter(mid_col)}5"] = f"NĂM HỌC: {active_year.name}"
            ws[f"{get_column_letter(mid_col)}5"].font = Font(name="Times New Roman", size=12, bold=True)
            ws[f"{get_column_letter(mid_col)}5"].alignment = align_center

            current_row = 6
            for grp in sorted(groups_data.keys()):
                ws.cell(row=current_row, column=1, value=str(grp)).font = Font(name="Times New Roman", size=12, bold=True, italic=True)
                current_row += 1
                
                headers = ["STT", "LỚP"] + [w.replace("Tuần ", "T.") for w in selected_weeks] + ["ĐIỂM", "HẠNG\nHIỆN TẠI", "GVCN"]
                for col_idx, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                    cell.font = font_header_table
                    cell.alignment = align_center
                    cell.border = border_thin
                current_row += 1
                
                for idx, item in enumerate(groups_data[grp], 1):
                    val_tot = int(item['total_score']) if float(item['total_score']).is_integer() else item['total_score']
                    row_vals = [idx, item['branch_name']]
                    for w in selected_weeks:
                        ws_score = item['week_scores'].get(w, 0.0)
                        row_vals.append(int(ws_score) if float(ws_score).is_integer() else ws_score)
                    row_vals.extend([val_tot, item['rank'], item['gvcn']])
                    
                    for col_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = font_data
                        cell.border = border_thin
                        if col_idx == 2 or col_idx == num_cols:
                            cell.alignment = align_center if col_idx == 2 else align_left
                        else:
                            cell.alignment = align_center
                    current_row += 1
                current_row += 1 

            current_row += 1
            start_col = num_cols - 2 if num_cols >= 3 else num_cols
            ws.merge_cells(start_row=current_row, start_column=start_col-1, end_row=current_row, end_column=num_cols)
            sign_cell = ws.cell(row=current_row, column=start_col-1, value="TM. BCH ĐOÀN TRƯỜNG")
            sign_cell.font = font_title
            sign_cell.alignment = align_center

            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 10
            for i in range(len(selected_weeks)):
                ws.column_dimensions[get_column_letter(3+i)].width = 7
            ws.column_dimensions[get_column_letter(num_cols - 2)].width = 10 
            ws.column_dimensions[get_column_letter(num_cols - 1)].width = 12 
            ws.column_dimensions[get_column_letter(num_cols)].width = 25 

            log_system_action("XUẤT EXCEL", f"Xuất Báo cáo Điểm {selected_month}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name=f"Bao_Cao_Diem_{selected_month.replace(' ', '_')}.xlsx", as_attachment=True)

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi xuất Excel: {str(e)}", "error")
        return redirect(url_for('monthly'))

# ==========================================
# MODULE: THI ĐUA HỌC KỲ (LẤY DỮ LIỆU TỪ THÁNG & LƯU HỆ THỐNG)
# ==========================================
@app.route('/semester', methods=['GET', 'POST'])
def semester():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            
            available_months = []
            used_by_other_semesters = set()
            current_semester_months = set()
            semester_data = {}
            
            selected_semester = request.form.get('semester') or request.args.get('semester') or 'Học kỳ 1'
            selected_months = request.form.getlist('months')
            action = request.form.get('action', 'view')

            if active_year:
                months_db = db_session.query(MonthlyRecord.month_name).filter(
                    MonthlyRecord.school_year_id == active_year.id,
                    MonthlyRecord.month_name.like('Tháng%')
                ).distinct().all()
                
                school_order = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
                raw_months = [m[0] for m in months_db if m[0]]
                available_months = sorted(raw_months, key=lambda x: school_order.index(x) if x in school_order else 99)

                all_sem_records = db_session.query(MonthlyRecord).filter(
                    MonthlyRecord.school_year_id == active_year.id,
                    MonthlyRecord.month_name.like("Học kỳ%")
                ).all()
                
                for rec in all_sem_records:
                    if rec.weeks_used:
                        m_list = [m.strip() for m in rec.weeks_used.split(",") if m.strip()]
                        if rec.month_name == selected_semester:
                            current_semester_months.update(m_list)
                        else:
                            used_by_other_semesters.update(m_list)

                if request.method == 'GET' and not selected_months:
                    selected_months = list(current_semester_months)
                    selected_months = sorted(selected_months, key=lambda x: school_order.index(x) if x in school_order else 99)

                if selected_months:
                    prev_sem_name = "Học kỳ 1" if selected_semester == "Học kỳ 2" else None
                    prev_ranks = {}
                    if prev_sem_name:
                        p_records = db_session.query(MonthlyRecord).filter(
                            MonthlyRecord.school_year_id == active_year.id,
                            MonthlyRecord.month_name == prev_sem_name
                        ).all()
                        for pr in p_records:
                            prev_ranks[pr.branch_id] = pr.rank

                    branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                    for b in branches:
                        grp = b.group or "Nhóm 1"
                        if grp not in semester_data: semester_data[grp] = []
                        
                        m_scores = db_session.query(MonthlyRecord).filter(
                            MonthlyRecord.branch_id == b.id,
                            MonthlyRecord.month_name.in_(selected_months)
                        ).all()
                        
                        month_scores_dict = {s.month_name: (s.total_score or 0.0) for s in m_scores}
                        total_score = sum(month_scores_dict.values())
                        
                        semester_data[grp].append({
                            'branch_id': b.id,
                            'branch_name': b.name,
                            'gvcn': b.gvcn,
                            'month_scores': month_scores_dict,
                            'total_score': total_score,
                            'prev_rank': prev_ranks.get(b.id, "-")
                        })
                        
                    for grp, lst in semester_data.items():
                        lst.sort(key=lambda x: x['total_score'], reverse=True)
                        rk = 1
                        for i, d in enumerate(lst):
                            if i > 0 and d['total_score'] < lst[i-1]['total_score']: rk = i + 1
                            d['rank'] = rk
                            
                            p_rk = d['prev_rank']
                            diff_val = 0
                            change_str = "-"
                            if p_rk != "-" and isinstance(p_rk, int):
                                diff_val = p_rk - rk
                                if diff_val > 0: change_str = f"▲ Tăng {diff_val}"
                                elif diff_val < 0: change_str = f"▼ Giảm {abs(diff_val)}"
                                else: change_str = "▬ Giữ nguyên"
                            d['change_str'] = change_str
                            d['diff_val'] = diff_val

                    if action == 'save':
                        db_session.query(MonthlyRecord).filter(
                            MonthlyRecord.school_year_id == active_year.id,
                            MonthlyRecord.month_name == selected_semester
                        ).delete()
                        
                        months_str = ", ".join(selected_months)
                        for grp, lst in semester_data.items():
                            for d in lst:
                                db_session.add(MonthlyRecord(
                                    school_year_id=active_year.id,
                                    month_name=selected_semester,
                                    branch_id=d['branch_id'],
                                    total_score=d['total_score'],
                                    rank=d['rank'],
                                    weeks_used=months_str
                                ))
                        db_session.commit()
                        
                        log_system_action("LƯU ĐIỂM HỌC KỲ", f"Đã tính toán và chốt sổ điểm {selected_semester} (gộp từ: {months_str}).")
                        flash(f"✅ Đã lưu thành công dữ liệu {selected_semester} vào hệ thống!", "success")
                        return redirect(url_for('semester', semester=selected_semester))

            return render_template(
                'semester.html', 
                available_months=available_months, 
                used_by_other_semesters=list(used_by_other_semesters),
                semester_data=semester_data,
                selected_semester=selected_semester, 
                selected_months=selected_months
            )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi phân hệ thi đua học kỳ: {str(e)}", "error")
        return redirect(url_for('dashboard'))

# ==========================================
# MODULE: XUẤT EXCEL HỌC KỲ
# ==========================================
@app.route('/export_semester_excel', methods=['POST'])
def export_semester_excel():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return redirect(url_for('semester'))

            selected_semester = request.form.get('semester', 'Học kỳ 1')
            selected_months = request.form.getlist('months')
            is_hk1 = (selected_semester == "Học kỳ 1")

            if not selected_months:
                flash("Vui lòng chọn ít nhất 1 tháng để xuất Excel!", "error")
                return redirect(url_for('semester'))

            semester_data = {}
            prev_sem_name = "Học kỳ 1" if not is_hk1 else None
            prev_ranks = {}
            if prev_sem_name:
                p_records = db_session.query(MonthlyRecord).filter(MonthlyRecord.school_year_id == active_year.id, MonthlyRecord.month_name == prev_sem_name).all()
                for pr in p_records: prev_ranks[pr.branch_id] = pr.rank

            branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
            for b in branches:
                grp = b.group or "Nhóm 1"
                if grp not in semester_data: semester_data[grp] = []
                m_scores = db_session.query(MonthlyRecord).filter(MonthlyRecord.branch_id == b.id, MonthlyRecord.month_name.in_(selected_months)).all()
                
                month_scores_dict = {s.month_name: (s.total_score or 0.0) for s in m_scores}
                semester_data[grp].append({
                    'branch_name': b.name, 'gvcn': b.gvcn, 'month_scores': month_scores_dict,
                    'total_score': sum(month_scores_dict.values()), 'prev_rank': prev_ranks.get(b.id, "-")
                })
                
            for grp, lst in semester_data.items():
                lst.sort(key=lambda x: x['total_score'], reverse=True)
                rk = 1
                for i, d in enumerate(lst):
                    if i > 0 and d['total_score'] < lst[i-1]['total_score']: rk = i + 1
                    d['rank'] = rk
                    p_rk = d['prev_rank']
                    if p_rk != "-" and isinstance(p_rk, int):
                        diff_val = p_rk - rk
                        if diff_val > 0: d['change_str'] = f"▲ Tăng {diff_val}"
                        elif diff_val < 0: d['change_str'] = f"▼ Giảm {abs(diff_val)}"
                        else: d['change_str'] = "▬ Giữ nguyên"
                    else: d['change_str'] = "-"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Báo Cáo Học Kỳ"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_margins = PageMargins(left=0.75, right=0.5, top=0.75, bottom=0.75, header=0.5, footer=0.5)

            font_bold = Font(name='Times New Roman', size=11, bold=True)
            font_normal = Font(name='Times New Roman', size=11)
            font_title = Font(name='Times New Roman', size=14, bold=True)
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center')
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

            month_headers = [m.replace("Tháng ", "T.") for m in selected_months]
            if is_hk1: base_headers = ["STT", "LỚP"] + month_headers + ["TỔNG ĐIỂM", "HẠNG\n HIỆN TẠI", "GVCN"]
            else: base_headers = ["STT", "LỚP"] + month_headers + ["TỔNG ĐIỂM", "XH. KỲ TRƯỚC", "HẠNG\n HIỆN TẠI", "TĂNG/GIẢM", "GVCN"]
            
            total_cols = len(base_headers)
            last_col_letter = get_column_letter(total_cols)

            ws.merge_cells('A1:C1')
            ws['A1'] = "ĐOÀN TRƯỜNG THPT THANH HÒA"
            ws['A1'].font = font_bold
            ws['A1'].alignment = align_left

            ws.merge_cells(f'E1:{last_col_letter}1' if total_cols >= 5 else f'D1:{last_col_letter}1')
            ws['E1' if total_cols >= 5 else 'D1'] = "ĐOÀN TNCS HỒ CHÍ MINH"
            ws['E1' if total_cols >= 5 else 'D1'].font = Font(name='Times New Roman', size=14, bold=True)
            ws['E1' if total_cols >= 5 else 'D1'].alignment = align_center

            ws.merge_cells(f'A3:{last_col_letter}3')
            ws['A3'] = f"ĐIỂM THI ĐUA {selected_semester.upper()}"
            ws['A3'].font = font_title
            ws['A3'].alignment = align_center
            ws.row_dimensions[3].height = 25.0
            
            mid_col = total_cols // 2 if total_cols > 3 else 2
            ws.merge_cells(f'{get_column_letter(mid_col)}4:{get_column_letter(mid_col+2)}4')
            ws[f'{get_column_letter(mid_col)}4'] = f"NĂM HỌC: {active_year.name}"
            ws[f'{get_column_letter(mid_col)}4'].font = font_bold
            ws[f'{get_column_letter(mid_col)}4'].alignment = align_center

            current_row = 6
            for grp in sorted(semester_data.keys()):
                ws.cell(row=current_row, column=1, value=str(grp)).font = Font(name='Times New Roman', size=12, bold=True, italic=True)
                current_row += 1

                for col, h in enumerate(base_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=h)
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.border = border_thin
                    cell.fill = fill_header
                current_row += 1

                for idx, r in enumerate(semester_data[grp], 1):
                    row_data = [idx, r['branch_name']]
                    for m in selected_months:
                        sc = r['month_scores'].get(m, 0.0)
                        row_data.append(int(sc) if float(sc).is_integer() else sc)
                    
                    tot = r['total_score']
                    row_data.append(int(tot) if float(tot).is_integer() else tot)

                    if is_hk1:
                        row_data.append(r['rank'])
                        row_data.append(r['gvcn'])
                    else:
                        row_data.append(r['prev_rank'])
                        row_data.append(r['rank'])
                        row_data.append(r['change_str'])
                        row_data.append(r['gvcn'])

                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        cell.font = font_normal
                        cell.border = border_thin
                        cell.alignment = align_center if col != 2 and col != total_cols else align_left
                        
                        if not is_hk1 and base_headers[col-1] == "TĂNG/GIẢM":
                            if "Tăng" in str(val): cell.font = Font(name='Times New Roman', size=11, bold=True, color="008000")
                            elif "Giảm" in str(val): cell.font = Font(name='Times New Roman', size=11, bold=True, color="FF0000")

                    current_row += 1
                current_row += 1

            current_row += 1
            sign_col = total_cols - 1 if total_cols >= 3 else total_cols
            ws.merge_cells(start_row=current_row, start_column=sign_col-1, end_row=current_row, end_column=total_cols)
            sign_cell = ws.cell(row=current_row, column=sign_col-1, value="TM. BCH ĐOÀN TRƯỜNG")
            sign_cell.font = font_bold
            sign_cell.alignment = align_center

            ws.column_dimensions['A'].width = 5.5
            ws.column_dimensions['B'].width = 10.0
            for c in range(3, total_cols - 1):
                ws.column_dimensions[get_column_letter(c)].width = 7.0
            ws.column_dimensions[get_column_letter(total_cols - 2)].width = 10.0 
            ws.column_dimensions[get_column_letter(total_cols - 1)].width = 12.0 
            ws.column_dimensions[last_col_letter].width = 25.0 

            log_system_action("XUẤT EXCEL", f"Xuất Báo cáo Điểm {selected_semester}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name=f"Bao_Cao_{selected_semester.replace(' ', '_')}.xlsx", as_attachment=True)

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi xuất Excel: {str(e)}", "error")
        return redirect(url_for('semester'))

# ==========================================
# MODULE: TỔNG KẾT NĂM HỌC
# ==========================================
@app.route('/yearly', methods=['GET', 'POST'])
def yearly():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            
            available_semesters = []
            yearly_data = {}
            
            selected_year_name = request.form.get('year_name') or request.args.get('year_name') or 'Năm học'
            if active_year and selected_year_name == 'Năm học':
                selected_year_name = f"Năm học {active_year.name}"
                
            selected_sems = request.form.getlist('semesters')
            action = request.form.get('action', 'view')

            if active_year:
                sems_db = db_session.query(MonthlyRecord.month_name).filter(
                    MonthlyRecord.school_year_id == active_year.id,
                    MonthlyRecord.month_name.like('Học kỳ%')
                ).distinct().all()
                
                available_semesters = sorted([m[0] for m in sems_db if m[0]])

                if request.method == 'GET' and not selected_sems:
                    selected_sems = available_semesters

                if selected_sems:
                    branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
                    for b in branches:
                        grp = b.group or "Nhóm 1"
                        if grp not in yearly_data: yearly_data[grp] = []
                        
                        sem_scores_records = db_session.query(MonthlyRecord).filter(
                            MonthlyRecord.branch_id == b.id,
                            MonthlyRecord.month_name.in_(selected_sems)
                        ).all()
                        
                        hk1_score = "-"
                        hk2_score = "-"
                        total_score = 0.0
                        
                        for s in sem_scores_records:
                            val = s.total_score or 0.0
                            if "1" in s.month_name: hk1_score = val
                            if "2" in s.month_name: hk2_score = val
                            total_score += val
                            
                        yearly_data[grp].append({
                            'branch_id': b.id,
                            'branch_name': b.name,
                            'gvcn': b.gvcn,
                            'hk1_score': hk1_score,
                            'hk2_score': hk2_score,
                            'total_score': total_score
                        })
                        
                    for grp, lst in yearly_data.items():
                        lst.sort(key=lambda x: x['total_score'], reverse=True)
                        rk = 1
                        for i, d in enumerate(lst):
                            if i > 0 and d['total_score'] < lst[i-1]['total_score']: rk = i + 1
                            d['rank'] = rk

                    if action == 'save':
                        db_session.query(MonthlyRecord).filter(
                            MonthlyRecord.school_year_id == active_year.id,
                            MonthlyRecord.month_name == selected_year_name
                        ).delete()
                        
                        sems_str = ", ".join(selected_sems)
                        for grp, lst in yearly_data.items():
                            for d in lst:
                                db_session.add(MonthlyRecord(
                                    school_year_id=active_year.id,
                                    month_name=selected_year_name,
                                    branch_id=d['branch_id'],
                                    total_score=d['total_score'],
                                    rank=d['rank'],
                                    weeks_used=sems_str
                                ))
                        db_session.commit()
                        
                        log_system_action("LƯU ĐIỂM NĂM HỌC", f"Đã tính toán và chốt sổ điểm {selected_year_name} (gộp từ: {sems_str}).")
                        flash(f"✅ Đã lưu thành công dữ liệu Tổng kết {selected_year_name}!", "success")
                        return redirect(url_for('yearly', year_name=selected_year_name))

            return render_template(
                'yearly.html', 
                available_semesters=available_semesters, 
                yearly_data=yearly_data,
                selected_year_name=selected_year_name, 
                selected_sems=selected_sems,
                active_year=active_year
            )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi phân hệ Tổng kết Năm học: {str(e)}", "error")
        return redirect(url_for('dashboard'))

@app.route('/export_yearly_excel', methods=['POST'])
def export_yearly_excel():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return redirect(url_for('yearly'))

            selected_year_name = request.form.get('year_name', 'Năm học')
            selected_sems = request.form.getlist('semesters')

            if not selected_sems:
                flash("Vui lòng chọn ít nhất 1 học kỳ để xuất Excel!", "error")
                return redirect(url_for('yearly'))

            yearly_data = {}
            branches = db_session.query(Branch).filter(Branch.school_year_id == active_year.id).all()
            for b in branches:
                grp = b.group or "Nhóm 1"
                if grp not in yearly_data: yearly_data[grp] = []
                
                sem_scores = db_session.query(MonthlyRecord).filter(MonthlyRecord.branch_id == b.id, MonthlyRecord.month_name.in_(selected_sems)).all()
                hk1 = "-"; hk2 = "-"; tot = 0.0
                for s in sem_scores:
                    val = s.total_score or 0.0
                    if "1" in str(s.month_name): hk1 = val
                    if "2" in str(s.month_name): hk2 = val
                    tot += val
                    
                yearly_data[grp].append({'branch_name': b.name, 'gvcn': b.gvcn, 'hk1_score': hk1, 'hk2_score': hk2, 'total_score': tot})
                
            for grp, lst in yearly_data.items():
                lst.sort(key=lambda x: x['total_score'], reverse=True)
                rk = 1
                for i, d in enumerate(lst):
                    if i > 0 and d['total_score'] < lst[i-1]['total_score']: rk = i + 1
                    d['rank'] = rk

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng Kết Năm Học"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_margins = PageMargins(left=0.75, right=0.5, top=0.75, bottom=0.75, header=0.5, footer=0.5)

            font_bold = Font(name='Times New Roman', size=11, bold=True)
            font_normal = Font(name='Times New Roman', size=11)
            font_title = Font(name='Times New Roman', size=14, bold=True)
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center')
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

            ws.merge_cells('A1:C1'); ws['A1'] = "ĐOÀN TRƯỜNG THPT THANH HÒA"; ws['A1'].font = font_bold; ws['A1'].alignment = align_left
            ws.merge_cells('E1:G1'); ws['E1'] = "ĐOÀN TNCS HỒ CHÍ MINH"; ws['E1'].font = Font(name='Times New Roman', size=14, bold=True); ws['E1'].alignment = align_center
            ws.merge_cells('A3:G3'); ws['A3'] = f"TỔNG KẾT THI ĐUA NĂM HỌC - {selected_year_name.upper()}"; ws['A3'].font = font_title; ws['A3'].alignment = align_center

            current_row = 5
            for grp in sorted(yearly_data.keys()):
                ws.cell(row=current_row, column=1, value=str(grp)).font = font_bold; current_row += 1

                headers = ["STT", "LỚP", "ĐIỂM HK1", "ĐIỂM HK2", "TỔNG ĐIỂM NĂM", "HẠNG HIỆN TẠI", "GVCN"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=h)
                    cell.font = font_bold; cell.alignment = align_center; cell.border = border_thin; cell.fill = fill_header
                current_row += 1

                for idx, r in enumerate(yearly_data[grp], 1):
                    h1 = int(r['hk1_score']) if isinstance(r['hk1_score'], float) and r['hk1_score'].is_integer() else r['hk1_score']
                    h2 = int(r['hk2_score']) if isinstance(r['hk2_score'], float) and r['hk2_score'].is_integer() else r['hk2_score']
                    tt = int(r['total_score']) if isinstance(r['total_score'], float) and r['total_score'].is_integer() else r['total_score']
                    
                    row_data = [idx, r['branch_name'], h1, h2, tt, r['rank'], r['gvcn']]
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        cell.font = font_normal; cell.border = border_thin
                        cell.alignment = align_center if col in [1, 3, 4, 5, 6] else align_left
                    current_row += 1
                current_row += 1

            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
            ws.cell(row=current_row, column=5, value="TM. BCH ĐOÀN TRƯỜNG").font = font_bold; ws.cell(row=current_row, column=5).alignment = align_center

            widths = {'A': 5.5, 'B': 14.0, 'C': 12.0, 'D': 12.0, 'E': 15.0, 'F': 18.0, 'G': 25.0}
            for col, width in widths.items(): ws.column_dimensions[col].width = width

            log_system_action("XUẤT EXCEL", f"Xuất báo cáo Điểm {selected_year_name}")
            out = io.BytesIO(); wb.save(out); out.seek(0)
            return send_file(out, download_name=f"Tong_Ket_Nam_Hoc.xlsx", as_attachment=True)

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi xuất Excel: {str(e)}", "error")
        return redirect(url_for('yearly'))

# =========================================================================
# THUẬT TOÁN ĐỌC DỮ LIỆU TỪ CƠ SỞ DỮ LIỆU SQLITE CHO BIỂU MẪU BÁO CÁO
# =========================================================================
def calculate_trimmed_good_points_web(session, week_name, branch_name, branch_group, max_mon, max_tot):
    """Tính lại số điểm tốt dựa trên dữ liệu thô từ CSDL và barem động hiện hành"""
    try:
        records = session.query(RawScore).filter_by(week=week_name, branch_name=branch_name).all()
        raw_scores = [{"c10": r.c10, "c9": r.c9, "c8": r.c8} for r in records]
    except Exception:
        raw_scores = []
        
    if not raw_scores: return 0
    f10, f9, f8 = 0, 0, 0
    for r in raw_scores:
        try:
            c10, c9, c8 = int(r.get('c10', 0)), int(r.get('c9', 0)), int(r.get('c8', 0))
        except ValueError: continue
        
        if "1" in str(branch_group): c8 = 0 
        
        k10 = min(c10, max_mon)
        k9 = min(c9, max_mon - k10)
        k8 = min(c8, max_mon - k10 - k9)
        f10 += k10; f9 += k9; f8 += k8
        
    total = f10 + f9 + f8
    return min(total, max_tot)

@app.route('/templates', methods=['GET', 'POST'])
def templates_report():
    try:
        with session_scope() as db_session:
            years = db_session.query(SchoolYear).order_by(SchoolYear.id.desc()).all()
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            
            selected_year_name = request.args.get('year_name', active_year.name if active_year else '')
            report_type = request.args.get('report_type', 'Báo cáo Tuần')
            time_val = request.args.get('time_val', 'Tuần 1')
            group_val = request.args.get('group_val', 'Tất cả')

            selected_year = db_session.query(SchoolYear).filter_by(name=selected_year_name).first()
            
            time_options = []
            if "Tuần" in report_type:
                time_options = [f"Tuần {i}" for i in range(1, 39)]
            elif "Tháng" in report_type:
                time_options = ["Tháng 8", "Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
            elif "Học kỳ" in report_type:
                time_options = ["Học kỳ 1", "Học kỳ 2", "Cả năm học"]

            data_list = []
            start_date_str = "…"
            end_date_str = "…"

            if selected_year:
                max_tot, max_mon = 14, 4
                try:
                    settings = db_session.query(ScoreSettings).filter_by(school_year_id=selected_year.id).first()
                    if settings:
                        if hasattr(settings, 'max_diem_tot'): max_tot = int(settings.max_diem_tot)
                        if hasattr(settings, 'max_diem_mon'): max_mon = int(settings.max_diem_mon)
                except Exception: pass

                branch_query = db_session.query(Branch).filter_by(school_year_id=selected_year.id)
                if group_val != "Tất cả":
                    if group_val.startswith("Khối"):
                        khoi = group_val.replace("Khối", "").strip()
                        from sqlalchemy import or_
                        branch_query = branch_query.filter(or_(Branch.name.startswith(khoi), Branch.name.like(f"% {khoi}%")))
                    else:
                        branch_query = branch_query.filter_by(group=group_val)
                branches = branch_query.all()
                branch_ids = [b.id for b in branches]

                if branches:
                    if "Báo cáo Tuần" in report_type:
                        scores = db_session.query(WeeklyScore).filter(WeeklyScore.week == time_val, WeeklyScore.branch_id.in_(branch_ids)).all()
                        if scores:
                            start_date_str = scores[0].start_date or "…"
                            end_date_str = scores[0].end_date or "…"
                        
                        for branch in branches:
                            sc = next((s for s in scores if s.branch_id == branch.id), None)
                            gvcn_val = branch.gvcn if branch.gvcn else ""
                            so_diem_tot = calculate_trimmed_good_points_web(db_session, time_val, branch.name, branch.group, max_mon, max_tot) if sc else 0
                            data_list.append({
                                "Chi đoàn": branch.name,
                                "Nhóm": branch.group or "Nhóm 1",
                                "Sĩ số": branch.si_so,
                                "Xếp loại": getattr(sc, 'week_rating', '-') if sc else '-',
                                "Điểm Trừ VP": getattr(sc, 'score_kem', 0) if sc else 0,
                                "Số Điểm Tốt": so_diem_tot,
                                "Tổng Điểm": getattr(sc, 'total_score', 0) if sc else 0,
                                "Ghi chú VP": getattr(sc, 'note', '') if sc else 'Chưa nhập điểm',
                                "Giáo viên chủ nhiệm": gvcn_val
                            })

                    elif "Báo cáo Tháng" in report_type:
                        month_scores = db_session.query(MonthlyRecord).filter(MonthlyRecord.month_name == time_val, MonthlyRecord.school_year_id == selected_year.id, MonthlyRecord.branch_id.in_(branch_ids)).all()
                        weekly_scores_all = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(branch_ids)).all()

                        for branch in branches:
                            m_sc = next((m for m in month_scores if m.branch_id == branch.id), None)
                            gvcn_val = branch.gvcn if branch.gvcn else ""
                            if m_sc:
                                ghi_chu_gop = []
                                tong_diem_tru = 0
                                tong_diem_tot = 0
                                if getattr(m_sc, 'weeks_used', None):
                                    weeks = [w.strip() for w in m_sc.weeks_used.split(",")]
                                    branch_weeks = [w for w in weekly_scores_all if w.branch_id == branch.id and w.week in weeks]
                                    for bw in branch_weeks:
                                        tong_diem_tru += (bw.score_kem or 0)
                                        if bw.note and bw.note.strip():
                                            ghi_chu_gop.append(f"[{bw.week.replace('Tuần ', 'T')}] {bw.note.strip()}")
                                    for w in weeks:
                                        tong_diem_tot += calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot)
                                
                                xep_loai = getattr(m_sc, 'rating', 'Tốt')
                                data_list.append({
                                    "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                                    "Xếp loại": xep_loai, "Điểm Trừ VP": tong_diem_tru, "Số Điểm Tốt": tong_diem_tot,
                                    "Tổng Điểm": m_sc.total_score, "Ghi chú VP": " | ".join(ghi_chu_gop), "Giáo viên chủ nhiệm": gvcn_val
                                })
                            else:
                                data_list.append({
                                    "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                                    "Xếp loại": "-", "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                                    "Ghi chú VP": "Chưa tổng hợp tháng", "Giáo viên chủ nhiệm": gvcn_val
                                })

                    elif "Báo cáo Học kỳ" in report_type:
                        if time_val == "Học kỳ 1": target_weeks = [f"Tuần {i}" for i in range(1, 19)]
                        elif time_val == "Học kỳ 2": target_weeks = [f"Tuần {i}" for i in range(19, 38)]
                        elif time_val == "Cả năm học": target_weeks = [f"Tuần {i}" for i in range(1, 38)]
                        else: target_weeks = []

                        scores = db_session.query(WeeklyScore).filter(WeeklyScore.week.in_(target_weeks), WeeklyScore.branch_id.in_(branch_ids)).all() if target_weeks else []
                        for branch in branches:
                            branch_scores = [sc for sc in scores if sc.branch_id == branch.id]
                            gvcn_val = branch.gvcn if branch.gvcn else ""
                            if branch_scores:
                                tong_diem = sum((sc.total_score or 0) for sc in branch_scores)
                                tong_diem_tru_vp = sum((sc.score_kem or 0) for sc in branch_scores)
                                ghi_chu_gop = [f"[{sc.week.replace('Tuần ', 'T')}] {sc.note.strip()}" for sc in branch_scores if sc.note and sc.note.strip()]
                                tong_diem_tot = sum(calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot) for w in target_weeks)
                                
                                data_list.append({
                                    "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                                    "Xếp loại": "Tốt" if tong_diem >= 90 else "Khá", "Điểm Trừ VP": tong_diem_tru_vp,
                                    "Số Điểm Tốt": tong_diem_tot, "Tổng Điểm": round(tong_diem, 2),
                                    "Ghi chú VP": " | ".join(ghi_chu_gop), "Giáo viên chủ nhiệm": gvcn_val
                                })
                            else:
                                data_list.append({
                                    "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                                    "Xếp loại": "-", "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                                    "Ghi chú VP": "Chưa có dữ liệu", "Giáo viên chủ nhiệm": gvcn_val
                                })

            if data_list:
                df = pd.DataFrame(data_list)
                if "Nhóm" in df.columns:
                    df["Hạng"] = df.groupby("Nhóm")["Tổng Điểm"].rank(method="min", ascending=False).astype(int)
                    df = df.sort_values(by=["Nhóm", "Hạng"])
                data_list = df.to_dict('records')

            return render_template(
                'templates_report.html',
                years=years,
                selected_year_name=selected_year_name,
                report_type=report_type,
                time_val=time_val,
                group_val=group_val,
                time_options=time_options,
                data_list=data_list,
                start_date=start_date_str,
                end_date=end_date_str
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải biểu mẫu báo cáo: {e}", "error")
        return redirect(url_for('dashboard'))
    
# ==========================================
# MODULE: XUẤT EXCEL BIỂU MẪU (A4 DỌC, CHIA NHÓM)
# ==========================================
@app.route('/preview_templates_report', methods=['GET', 'POST'])
def preview_templates_report():
    try:
        # Hỗ trợ nhận dữ liệu linh hoạt từ cả GET và POST
        if request.method == 'POST':
            report_type = request.form.get('report_type', 'Báo cáo Tuần')
            time_val = request.form.get('time_val', 'Tuần 1')
            year_name = request.form.get('year_name', '')
            group_val = request.form.get('group_val', 'Tất cả')
        else:
            report_type = request.args.get('report_type', 'Báo cáo Tuần')
            time_val = request.args.get('time_val', 'Tuần 1')
            year_name = request.args.get('year_name', '')
            group_val = request.args.get('group_val', 'Tất cả')

        with session_scope() as db_session:
            selected_year = db_session.query(SchoolYear).filter_by(name=year_name).first()
            if not selected_year:
                flash("Không tìm thấy năm học!", "error")
                return redirect(url_for('templates_report'))

            max_tot, max_mon = 14, 4
            try:
                settings = db_session.query(ScoreSettings).filter_by(school_year_id=selected_year.id).first()
                if settings:
                    if hasattr(settings, 'max_diem_tot'): max_tot = int(settings.max_diem_tot)
                    if hasattr(settings, 'max_diem_mon'): max_mon = int(settings.max_diem_mon)
            except Exception: pass

            branch_query = db_session.query(Branch).filter_by(school_year_id=selected_year.id)
            if group_val != "Tất cả":
                if group_val.startswith("Khối"):
                    khoi = group_val.replace("Khối", "").strip()
                    from sqlalchemy import or_
                    branch_query = branch_query.filter(or_(Branch.name.startswith(khoi), Branch.name.like(f"% {khoi}%")))
                else:
                    branch_query = branch_query.filter_by(group=group_val)
            branches = branch_query.all()
            branch_ids = [b.id for b in branches]

            data_list = []
            start_str, end_str = "…", "…"

            if "Báo cáo Tuần" in report_type:
                scores = db_session.query(WeeklyScore).filter(WeeklyScore.week == time_val, WeeklyScore.branch_id.in_(branch_ids)).all()
                if scores:
                    start_str = scores[0].start_date or "…"
                    end_str = scores[0].end_date or "…"
                for branch in branches:
                    sc = next((s for s in scores if s.branch_id == branch.id), None)
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    so_diem_tot = calculate_trimmed_good_points_web(db_session, time_val, branch.name, branch.group, max_mon, max_tot) if sc else 0
                    data_list.append({
                        "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                        "Điểm Trừ VP": getattr(sc, 'score_kem', 0) if sc else 0, "Số Điểm Tốt": so_diem_tot,
                        "Tổng Điểm": getattr(sc, 'total_score', 0) if sc else 0,
                        "Giáo viên chủ nhiệm": gvcn_val
                    })
            elif "Báo cáo Tháng" in report_type:
                month_scores = db_session.query(MonthlyRecord).filter(MonthlyRecord.month_name == time_val, MonthlyRecord.school_year_id == selected_year.id, MonthlyRecord.branch_id.in_(branch_ids)).all()
                weekly_scores_all = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(branch_ids)).all()
                for branch in branches:
                    m_sc = next((m for m in month_scores if m.branch_id == branch.id), None)
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    if m_sc:
                        tong_diem_tru = 0; tong_diem_tot = 0
                        if getattr(m_sc, 'weeks_used', None):
                            weeks = [w.strip() for w in m_sc.weeks_used.split(",")]
                            branch_weeks = [w for w in weekly_scores_all if w.branch_id == branch.id and w.week in weeks]
                            for bw in branch_weeks:
                                tong_diem_tru += (bw.score_kem or 0)
                            for w in weeks:
                                tong_diem_tot += calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot)
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": tong_diem_tru, "Số Điểm Tốt": tong_diem_tot, "Tổng Điểm": m_sc.total_score,
                            "Giáo viên chủ nhiệm": gvcn_val
                        })
                    else:
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                            "Giáo viên chủ nhiệm": gvcn_val
                        })
            elif "Báo cáo Học kỳ" in report_type:
                if time_val == "Học kỳ 1": target_weeks = [f"Tuần {i}" for i in range(1, 19)]
                elif time_val == "Học kỳ 2": target_weeks = [f"Tuần {i}" for i in range(19, 38)]
                elif time_val == "Cả năm học": target_weeks = [f"Tuần {i}" for i in range(1, 38)]
                else: target_weeks = []

                scores = db_session.query(WeeklyScore).filter(WeeklyScore.week.in_(target_weeks), WeeklyScore.branch_id.in_(branch_ids)).all() if target_weeks else []
                for branch in branches:
                    branch_scores = [sc for sc in scores if sc.branch_id == branch.id]
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    if branch_scores:
                        tong_diem = sum((sc.total_score or 0) for sc in branch_scores)
                        tong_diem_tru_vp = sum((sc.score_kem or 0) for sc in branch_scores)
                        tong_diem_tot = sum(calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot) for w in target_weeks)
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": tong_diem_tru_vp, "Số Điểm Tốt": tong_diem_tot, "Tổng Điểm": round(tong_diem, 2),
                            "Giáo viên chủ nhiệm": gvcn_val
                        })
                    else:
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                            "Giáo viên chủ nhiệm": gvcn_val
                        })

            df = pd.DataFrame(data_list)
            if not df.empty and "Nhóm" in df.columns:
                df["Hạng"] = df.groupby("Nhóm")["Tổng Điểm"].rank(method="min", ascending=False).astype(int)
                df = df.sort_values(by=["Nhóm", "Hạng"])
                data_list = df.to_dict('records')

            return render_template(
                'preview_templates_excel.html',
                report_type=report_type,
                time_val=time_val,
                year_name=year_name,
                group_val=group_val,
                start_str=start_str,
                end_str=end_str,
                data_list=data_list
            )
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi xem trước: {e}", "error")
        return redirect(url_for('templates_report'))
# ==========================================
# MODULE: XUẤT EXCEL BIỂU MẪU CHÍNH THỨC (A4 DỌC, CHIA NHÓM)
# ==========================================
@app.route('/export_templates_excel', methods=['POST'])
def export_templates_excel():
    try:
        report_type = request.form.get('report_type', 'Báo cáo Tuần')
        time_val = request.form.get('time_val', 'Tuần 1')
        year_name = request.form.get('year_name', '')
        group_val = request.form.get('group_val', 'Tất cả')

        with session_scope() as db_session:
            selected_year = db_session.query(SchoolYear).filter_by(name=year_name).first()
            if not selected_year:
                flash("Không tìm thấy năm học!", "error")
                return redirect(url_for('templates_report'))

            max_tot, max_mon = 14, 4
            try:
                settings = db_session.query(ScoreSettings).filter_by(school_year_id=selected_year.id).first()
                if settings:
                    if hasattr(settings, 'max_diem_tot'): max_tot = int(settings.max_diem_tot)
                    if hasattr(settings, 'max_diem_mon'): max_mon = int(settings.max_diem_mon)
            except Exception: pass

            branch_query = db_session.query(Branch).filter_by(school_year_id=selected_year.id)
            if group_val != "Tất cả":
                if group_val.startswith("Khối"):
                    khoi = group_val.replace("Khối", "").strip()
                    from sqlalchemy import or_
                    branch_query = branch_query.filter(or_(Branch.name.startswith(khoi), Branch.name.like(f"% {khoi}%")))
                else:
                    branch_query = branch_query.filter_by(group=group_val)
            branches = branch_query.all()
            branch_ids = [b.id for b in branches]

            if not branches:
                flash("Không có dữ liệu chi đoàn để xuất!", "error")
                return redirect(url_for('templates_report'))

            data_list = []
            start_str, end_str = "…", "…"

            if "Báo cáo Tuần" in report_type:
                scores = db_session.query(WeeklyScore).filter(WeeklyScore.week == time_val, WeeklyScore.branch_id.in_(branch_ids)).all()
                if scores:
                    start_str = scores[0].start_date or "…"
                    end_str = scores[0].end_date or "…"
                for branch in branches:
                    sc = next((s for s in scores if s.branch_id == branch.id), None)
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    so_diem_tot = calculate_trimmed_good_points_web(db_session, time_val, branch.name, branch.group, max_mon, max_tot) if sc else 0
                    data_list.append({
                        "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                        "Điểm Trừ VP": getattr(sc, 'score_kem', 0) if sc else 0, "Số Điểm Tốt": so_diem_tot,
                        "Tổng Điểm": getattr(sc, 'total_score', 0) if sc else 0,
                        "Ghi chú VP": getattr(sc, 'note', '') if sc else 'Chưa nhập điểm',
                        "Giáo viên chủ nhiệm": gvcn_val
                    })
            elif "Báo cáo Tháng" in report_type:
                month_scores = db_session.query(MonthlyRecord).filter(MonthlyRecord.month_name == time_val, MonthlyRecord.school_year_id == selected_year.id, MonthlyRecord.branch_id.in_(branch_ids)).all()
                weekly_scores_all = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(branch_ids)).all()
                for branch in branches:
                    m_sc = next((m for m in month_scores if m.branch_id == branch.id), None)
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    if m_sc:
                        ghi_chu_gop = []; tong_diem_tru = 0; tong_diem_tot = 0
                        if getattr(m_sc, 'weeks_used', None):
                            weeks = [w.strip() for w in m_sc.weeks_used.split(",")]
                            branch_weeks = [w for w in weekly_scores_all if w.branch_id == branch.id and w.week in weeks]
                            for bw in branch_weeks:
                                tong_diem_tru += (bw.score_kem or 0)
                                if bw.note and bw.note.strip():
                                    ghi_chu_gop.append(f"[{bw.week.replace('Tuần ', 'T')}] {bw.note.strip()}")
                            for w in weeks:
                                tong_diem_tot += calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot)
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": tong_diem_tru, "Số Điểm Tốt": tong_diem_tot, "Tổng Điểm": m_sc.total_score,
                            "Ghi chú VP": " | ".join(ghi_chu_gop), "Giáo viên chủ nhiệm": gvcn_val
                        })
                    else:
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                            "Ghi chú VP": "Chưa tổng hợp tháng", "Giáo viên chủ nhiệm": gvcn_val
                        })
            elif "Báo cáo Học kỳ" in report_type:
                if time_val == "Học kỳ 1": target_weeks = [f"Tuần {i}" for i in range(1, 19)]
                elif time_val == "Học kỳ 2": target_weeks = [f"Tuần {i}" for i in range(19, 38)]
                elif time_val == "Cả năm học": target_weeks = [f"Tuần {i}" for i in range(1, 38)]
                else: target_weeks = []

                scores = db_session.query(WeeklyScore).filter(WeeklyScore.week.in_(target_weeks), WeeklyScore.branch_id.in_(branch_ids)).all() if target_weeks else []
                for branch in branches:
                    branch_scores = [sc for sc in scores if sc.branch_id == branch.id]
                    gvcn_val = branch.gvcn if branch.gvcn else ""
                    if branch_scores:
                        tong_diem = sum((sc.total_score or 0) for sc in branch_scores)
                        tong_diem_tru_vp = sum((sc.score_kem or 0) for sc in branch_scores)
                        ghi_chu_gop = [f"[{sc.week.replace('Tuần ', 'T')}] {sc.note.strip()}" for sc in branch_scores if sc.note and sc.note.strip()]
                        tong_diem_tot = sum(calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot) for w in target_weeks)
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": tong_diem_tru_vp, "Số Điểm Tốt": tong_diem_tot, "Tổng Điểm": round(tong_diem, 2),
                            "Ghi chú VP": " | ".join(ghi_chu_gop), "Giáo viên chủ nhiệm": gvcn_val
                        })
                    else:
                        data_list.append({
                            "Chi đoàn": branch.name, "Nhóm": branch.group or "Nhóm 1", "Sĩ số": branch.si_so,
                            "Điểm Trừ VP": 0, "Số Điểm Tốt": 0, "Tổng Điểm": 0,
                            "Ghi chú VP": "Chưa có dữ liệu", "Giáo viên chủ nhiệm": gvcn_val
                        })

            df = pd.DataFrame(data_list)
            if not df.empty and "Nhóm" in df.columns:
                df["Hạng"] = df.groupby("Nhóm")["Tổng Điểm"].rank(method="min", ascending=False).astype(int)
                df = df.sort_values(by=["Nhóm", "Hạng"])

            # Khởi tạo file Excel với định dạng A4 Dọc
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bao_Cao_Thi_Dua"

            # Cấu hình khổ giấy A4 Dọc và chế độ xem trước
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # <--- Định dạng A4 Dọc
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.views.sheetView[0].view = "pageBreakPreview"

            # Thiết lập Header văn bản hành chính
            ws['B1'] = "ĐOÀN TNCS HỒ CHÍ MINH"
            ws['B1'].font = Font(name='Times New Roman', bold=True, size=11)
            ws['B1'].alignment = Alignment(horizontal='center')

            ws['G1'] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
            ws['G1'].font = Font(name='Times New Roman', bold=True, size=11)
            ws['G1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('B2:C2')
            ws['B2'] = "BCH TRƯỜNG THPT THANH HÒA"
            ws['B2'].font = Font(name='Times New Roman', bold=True, size=11)
            ws['B2'].alignment = Alignment(horizontal='center')

            ws.merge_cells('F2:H2')
            ws['F2'] = "Độc lập - Tự do - Hạnh phúc"
            ws['F2'].font = Font(name='Times New Roman', bold=True, size=11)
            ws['F2'].alignment = Alignment(horizontal='center')

            ws.merge_cells('B4:H4')
            ws['B4'] = f"BẢNG {report_type.upper()} - {time_val.upper()}"
            ws['B4'].font = Font(name='Times New Roman', bold=True, size=15)
            ws['B4'].alignment = Alignment(horizontal='center', vertical='center')

            if not start_str or start_str == "…": start_str = "........"
            if not end_str or end_str == "…": end_str = "........"

            ws.merge_cells('B5:H5')
            ws['B5'] = f"(Thời gian: Từ ngày {start_str} đến ngày {end_str})"
            ws['B5'].font = Font(name='Times New Roman', italic=True, size=11)
            ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            group_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")

            current_row = 7
            # Đã lược bỏ 'Ghi chú VP'
            columns_to_print = ['STT', 'Chi đoàn', 'Sĩ số', 'Điểm Trừ VP', 'Số Điểm Tốt', 'Hạng', 'Tổng Điểm', 'Giáo viên chủ nhiệm']
            
            grouped = df.groupby('Nhóm')
            for name, group in grouped:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
                g_cell = ws.cell(row=current_row, column=1, value=f"PHÂN BẢNG THI ĐUA: {str(name).upper()}")
                g_cell.font = Font(name='Times New Roman', bold=True, size=12, color="0F5132")
                g_cell.fill = group_fill
                g_cell.alignment = Alignment(horizontal='left', vertical='center')
                for c in range(1, 9):
                    ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1

                for col_num, header_name in enumerate(columns_to_print, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header_name)
                    cell.font = Font(name='Times New Roman', bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = header_fill
                current_row += 1

                for stt, (_, row_data) in enumerate(group.iterrows(), 1):
                    cell_stt = ws.cell(row=current_row, column=1, value=stt)
                    cell_stt.border = thin_border
                    cell_stt.alignment = Alignment(horizontal='center')

                    # Đã lược bỏ 'Ghi chú VP' khỏi danh sách khóa dữ liệu
                    col_keys = ['Chi đoàn', 'Sĩ số', 'Điểm Trừ VP', 'Số Điểm Tốt', 'Hạng', 'Tổng Điểm', 'Giáo viên chủ nhiệm']
                    for col_num, col_name in enumerate(col_keys, 2):
                        cell = ws.cell(row=current_row, column=col_num)
                        val = row_data.get(col_name, "")
                        if isinstance(val, float) and val.is_integer():
                            val = int(val)
                        cell.value = val
                        cell.font = Font(name='Times New Roman', size=11)
                        cell.border = thin_border
                        
                        if col_name in ['Sĩ số', 'Hạng', 'Tổng Điểm', 'Số Điểm Tốt', 'Điểm Trừ VP']:
                            cell.alignment = Alignment(horizontal='center')
                        else:
                            cell.alignment = Alignment(horizontal='left')
                    current_row += 1
                current_row += 1

            # Điều chỉnh lại độ rộng các cột tối ưu chuẩn A4 Dọc khi không có cột ghi chú
            widths = {'A': 6, 'B': 15, 'C': 9, 'D': 14, 'E': 14, 'F': 10, 'G': 12, 'H': 28}
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            log_system_action("XUẤT EXCEL", f"Xuất biểu mẫu A4 dọc chia nhóm: {report_type} - {time_val}")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            file_name = f"Bao_Cao_{report_type}_{time_val}.xlsx".replace(" ", "_")
            return send_file(out, download_name=file_name, as_attachment=True)
            
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi xuất Excel: {e}", "error")
        return redirect(url_for('templates_report'))
    

# ==========================================
# MODULE: SAO LƯU, PHỤC HỒI & QUẢN LÝ DỮ LIỆU (CHỈ DÀNH CHO ADMIN)
# ==========================================
def perform_backup_internal(actor="Hệ thống"):
    import glob, shutil
    from datetime import datetime
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"Data_ThiDua_Backup_{timestamp}"
    backup_path = os.path.join(backup_dir, backup_filename)
    temp_dir = os.path.join(backup_dir, f"temp_{timestamp}")
    os.makedirs(temp_dir, exist_ok=True)
    
    # Gom dữ liệu
    for folder in ["data", "config", "database"]:
        if os.path.exists(folder):
            shutil.copytree(folder, os.path.join(temp_dir, folder))
            
    # Nén zip
    shutil.make_archive(backup_path, 'zip', temp_dir)
    shutil.rmtree(temp_dir) # Dọn dẹp thư mục tạm
    
    # THUẬT TOÁN 1: Dọn rác tự động - Quét và chỉ giữ lại đúng 20 bản sao lưu mới nhất
    files = glob.glob(os.path.join(backup_dir, "*.zip"))
    files.sort(key=os.path.getmtime, reverse=True)
    if len(files) > 20:
        for old_file in files[20:]:
            try: os.remove(old_file)
            except: pass
            
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {actor} đã tạo bản sao lưu: {backup_filename}.zip")
    return f"{backup_filename}.zip"

# THUẬT TOÁN 3: Lập lịch sao lưu ngầm (Chạy nền)
def background_auto_backup():
    import time
    from datetime import datetime
    while True:
        now = datetime.now()
        # Nếu là Chủ nhật (weekday == 6) và thời gian rơi vào 23h55' đêm
        if now.weekday() == 6 and now.hour == 23 and now.minute == 55:
            try:
                perform_backup_internal(actor="BOT Lịch ngầm Tự động")
            except Exception as e:
                print("Lỗi auto backup:", e)
            time.sleep(3600) # Ngủ đông 1 tiếng để tránh chạy đúp lệnh trong cùng 1 đêm
        time.sleep(45) # Quét đồng hồ mỗi 45 giây

@app.route('/backup')
def backup_manager():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Bạn không có quyền truy cập chức năng này!", "error")
        return redirect(url_for('dashboard'))
        
    import glob
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)
    
    files = glob.glob(os.path.join(backup_dir, "*.zip"))
    backups = []
    for f in files:
        stat = os.stat(f)
        backups.append({
            'filename': os.path.basename(f),
            'size': round(stat.st_size / 1024 / 1024, 2),
            'time': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y - %H:%M:%S'),
            'raw_time': stat.st_mtime
        })
        
    backups.sort(key=lambda x: x['raw_time'], reverse=True)
    return render_template('backup.html', backups=backups)

@app.route('/create_backup', methods=['POST'])
def create_backup():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        return redirect(url_for('dashboard'))
    try:
        filename = perform_backup_internal(actor=session.get('username'))
        log_system_action("SAO LƯU DỮ LIỆU", f"Đã tạo bản sao lưu hệ thống: {filename}")
        flash("✅ Đã tạo sao lưu, đóng gói và tự động dọn rác thành công!", "success")
    except Exception as e:
        flash(f"Lỗi tạo sao lưu: {e}", "error")
    return redirect(url_for('backup_manager'))

# THUẬT TOÁN 2: Phục hồi 1 chạm (1-Click Restore)
@app.route('/restore_backup/<filename>', methods=['POST'])
def restore_backup(filename):
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        return redirect(url_for('dashboard'))
    try:
        import shutil
        import zipfile
        import os
        import time
        
        file_path = os.path.join("backups", filename)
        
        if not os.path.exists(file_path):
            flash("Không tìm thấy file sao lưu trên máy chủ!", "error")
            return redirect(url_for('backup_manager'))
            
        temp_extract = os.path.join("backups", "temp_restore")
        os.makedirs(temp_extract, exist_ok=True)
        
        # 1. Xả nén file Zip
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_extract)
            
        # --- [BẢN VÁ LỖI WINERROR 32]: NGẮT KẾT NỐI CSDL TRƯỚC KHI XÓA ---
        try:
            from database.database import engine
            engine.dispose()  # Ép SQLAlchemy nhả file thi_dua.db ra khỏi RAM
            time.sleep(1.5)   # Đợi 1.5 giây để HĐH Windows thu hồi file handle
        except Exception as e:
            print("Cảnh báo ngắt kết nối DB:", e)
        # -----------------------------------------------------------------
            
        # 2. Xóa dữ liệu cũ và ghi đè dữ liệu mới
        for folder in ["data", "config", "database"]:
            src = os.path.join(temp_extract, folder)
            if os.path.exists(src):
                if os.path.exists(folder):
                    shutil.rmtree(folder, ignore_errors=True) # Xóa thư mục hiện tại
                    time.sleep(0.5) # Nghỉ nhịp để HĐH làm mới cây thư mục
                    
                    # Cố gắng copy thư mục, nếu kẹt file thì copy đè từng file
                    if not os.path.exists(folder):
                        shutil.copytree(src, folder)
                    else:
                        for root, dirs, files in os.walk(src):
                            for f in files:
                                src_file = os.path.join(root, f)
                                dst_file = os.path.join(folder, os.path.relpath(src_file, src))
                                os.makedirs(os.path.dirname(dst_file), exist_ok=True)
                                shutil.copy2(src_file, dst_file)
                else:
                    shutil.copytree(src, folder)
                    
        shutil.rmtree(temp_extract, ignore_errors=True) # Dọn rác xả nén
        
        log_system_action("PHỤC HỒI DỮ LIỆU", f"Đã phục hồi hệ thống từ bản sao lưu: {filename}")
        flash(f"🔄 ĐÃ PHỤC HỒI THÀNH CÔNG DỮ LIỆU TỪ FILE {filename}!", "success")
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Lỗi phục hồi dữ liệu: {e}", "error")
        
    return redirect(url_for('backup_manager'))

@app.route('/download_backup/<filename>')
def download_backup(filename):
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        return redirect(url_for('dashboard'))
        
    file_path = os.path.join("backups", filename)
    if os.path.exists(file_path):
        log_system_action("SAO LƯU DỮ LIỆU", f"Tải xuống file sao lưu: {filename}")
        return send_file(file_path, as_attachment=True)
    else:
        flash("Không tìm thấy file sao lưu trên máy chủ!", "error")
        return redirect(url_for('backup_manager'))

@app.route('/delete_backup/<filename>', methods=['POST'])
def delete_backup(filename):
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        return redirect(url_for('dashboard'))
        
    file_path = os.path.join("backups", filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        log_system_action("SAO LƯU DỮ LIỆU", f"Đã xóa bản sao lưu: {filename}")
        flash(f"Đã xóa file {filename} thành công!", "success")
    else:
        flash("Không tìm thấy file trên máy chủ!", "error")
    return redirect(url_for('backup_manager'))

# ==========================================
# API TỰ ĐỘNG CẤP TÀI KHOẢN HÀNG LOẠT CHO GVCN
# ==========================================
@app.route('/auto_generate_gvcn', methods=['POST'])
def auto_generate_gvcn():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Chỉ Admin mới có quyền thực hiện!", "error")
        return redirect(url_for('users'))
        
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return redirect(url_for('users'))
                
            branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()
            count = 0
            
            for b in branches:
                username = b.name.strip().upper()
                exist = db_session.query(User).filter_by(username=username).first()
                
                if not exist:
                    # --- NÂNG CẤP: LẤY TÊN THẬT CỦA GVCN TỪ CHI ĐOÀN ---
                    # Nếu lớp đã có tên GVCN thì lấy tên đó, nếu trống thì mới dùng tạm "GVCN Lớp ..."
                    ten_gvcn = b.gvcn.strip() if b.gvcn and b.gvcn.strip() else f"GVCN Lớp {username}"
                    
                    new_user = User(
                        username=username, password_hash=username, # Pass mặc định = Tên lớp
                        full_name=ten_gvcn, role=UserRole.GVCN, is_active=True
                    )
                    db_session.add(new_user)
                    count += 1
                    try: sync_account_to_json(username, new_user.full_name, username, "Giáo viên chủ nhiệm", True)
                    except: pass
            
            log_system_action("CẤP TÀI KHOẢN", f"Đã tự động tạo {count} tài khoản GVCN.")
            flash(f"✅ Đã cấp phát tự động {count} tài khoản cho GVCN! (Tên đăng nhập = Mật khẩu = Tên lớp)", "success")
            return redirect(url_for('users'))
    except Exception as e:
        flash(f"Lỗi hệ thống: {e}", "error")
        return redirect(url_for('users'))

# ==========================================
# API: GVCN GỬI PHÚC KHẢO ĐIỂM
# ==========================================
@app.route('/submit_appeal', methods=['POST'])
def submit_appeal():
    if session.get('role') != 'Giáo viên chủ nhiệm': 
        return redirect(url_for('login'))
        
    score_id = request.form.get('score_id', type=int)
    reason = request.form.get('reason', '').strip()
    
    if not score_id or not reason: 
        return redirect(url_for('class_dashboard'))
        
    try:
        with session_scope() as db_session:
            score = db_session.query(WeeklyScore).filter_by(id=score_id).first()
            
            if not score:
                flash("Không tìm thấy dữ liệu điểm!", "error")
                return redirect(url_for('class_dashboard'))
            
            # 1. Kiểm tra xem tuần đã bị chốt cứng (khóa hoàn toàn) chưa
            if score.is_locked:
                flash("⛔ Tuần này đã chốt cứng, không thể gửi yêu cầu phúc khảo!", "error")
                return redirect(url_for('class_dashboard'))
            
            # 2. [NÂNG CẤP LÕI]: Kiểm tra giới hạn "Thời gian vàng" 3 ngày
            if score.is_appeal_expired:
                flash("⛔ Đã quá thời hạn 3 ngày. Hệ thống đã tự động khóa quyền khiếu nại đối với tuần này!", "error")
                return redirect(url_for('class_dashboard'))

            # 3. Nếu vượt qua mọi chốt chặn, tiến hành ghi nhận phúc khảo
            score.is_appealed = True
            score.appeal_reason = reason
            
            # Ghi log hệ thống rõ ràng
            log_system_action("PHÚC KHẢO", f"GVCN Lớp {score.branch.name} gửi khiếu nại Tuần {score.week}: {reason[:30]}...")
            flash("✅ Đã gửi Báo cáo sai sót / Phúc khảo đến Đoàn trường thành công!", "success")
            
    except Exception as e: 
        flash(f"Lỗi xử lý phúc khảo: {e}", "error")
        import traceback; traceback.print_exc() # Hỗ trợ debug nếu có lỗi ngầm
        
    return redirect(url_for('class_dashboard'))
# ==========================================
# MODULE: WEB APP MOBILE DÀNH CHO SAO ĐỎ
# ==========================================

@app.route('/auto_generate_saodo', methods=['POST'])
def auto_generate_saodo():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư']: return redirect(url_for('users'))
    try:
        with session_scope() as db_session:
            stars = db_session.query(RedStar).filter_by(is_active=True).all()
            count = 0
            for s in stars:
                # Tên đăng nhập: SD + ID của Sao đỏ (VD: SD1, SD15)
                username = f"SD{s.id}"
                exist = db_session.query(User).filter_by(username=username).first()
                if not exist:
                    new_user = User(
                        username=username, password_hash="123456", # Mật khẩu mặc định: 123456
                        full_name=f"SĐ: {s.full_name}", role=UserRole.SAO_DO, is_active=True
                    )
                    db_session.add(new_user)
                    count += 1
                    try: sync_account_to_json(username, new_user.full_name, "123456", "Sao đỏ", True)
                    except: pass
            
            log_system_action("CẤP TÀI KHOẢN", f"Đã tự động tạo {count} tài khoản Sao đỏ.")
            flash(f"✅ Đã cấp phát tự động {count} tài khoản Sao đỏ! (Tài khoản: SD + ID, Mật khẩu: 123456)", "success")
            return redirect(url_for('users'))
    except Exception as e:
        flash(f"Lỗi hệ thống: {e}", "error"); return redirect(url_for('users'))
# ==========================================
# API XUẤT DANH SÁCH TÀI KHOẢN SAO ĐỎ RA EXCEL
# ==========================================
@app.route('/export_saodo_accounts')
def export_saodo_accounts():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Bạn không có quyền thực hiện chức năng này!", "error")
        return redirect(url_for('users'))
        
    try:
        with session_scope() as db_session:
            # Lấy toàn bộ tài khoản Sao đỏ đang hoạt động
            sao_do_users = db_session.query(User).filter(User.role == UserRole.SAO_DO, User.is_active == True).all()
            
            if not sao_do_users:
                flash("Chưa có tài khoản Sao đỏ nào trên hệ thống!", "warning")
                return redirect(url_for('users'))
            
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            import io
            from flask import send_file
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DS_Tai_Khoan_Sao_Do"
            
            ws.merge_cells('A1:E1')
            ws['A1'] = "DANH SÁCH BÀN GIAO TÀI KHOẢN APP SAO ĐỎ"
            ws['A1'].font = Font(name="Times New Roman", size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            headers = ["STT", "Họ và Tên", "Chi đoàn", "Tên Đăng Nhập", "Mật Khẩu"]
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=3, column=col, value=h)
                c.font = Font(name="Times New Roman", size=12, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                
            for idx, u in enumerate(sao_do_users, 1):
                # Tách ID từ tên đăng nhập (VD: SD15 -> 15) để dò ra Tên Lớp
                branch_name = "Không rõ"
                try:
                    rs_id = int(u.username.replace('SD', ''))
                    rs = db_session.query(RedStar).filter_by(id=rs_id).first()
                    if rs and rs.branch:
                        branch_name = rs.branch.name
                except: pass
                
                # Làm sạch tên hiển thị
                real_name = u.full_name.replace("SĐ: ", "") if u.full_name else ""
                
                row_data = [idx, real_name, branch_name, u.username, u.password_hash]
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=idx+3, column=col, value=val)
                    c.font = Font(name="Times New Roman", size=12)
                    c.border = thin_border
                    if col in [1, 3, 4, 5]: c.alignment = Alignment(horizontal="center")
                    
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            
            log_system_action("XUẤT EXCEL", "Xuất danh sách bàn giao tài khoản Sao đỏ")
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, download_name="Danh_Sach_Tai_Khoan_Sao_Do.xlsx", as_attachment=True)
            
    except Exception as e:
        flash(f"Lỗi xuất Excel: {e}", "error")
        return redirect(url_for('users'))
    
@app.route('/mobile-sao-do', methods=['GET'])
def mobile_sao_do():
    if session.get('role') != 'Sao đỏ': return redirect(url_for('login'))
    
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return "Hệ thống chưa mở năm học mới."

            # Dịch ngược từ Username (VD: SD15) ra ID của Sao đỏ (15)
            username = session.get('username', '')
            try: star_id = int(username.replace('SD', '').strip())
            except: return "Tài khoản không hợp lệ."

            assignment = db_session.query(Assignment).filter_by(red_star_id=star_id).order_by(Assignment.week_number.desc()).first()
            all_branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()
            
            if not assignment:
                return render_template('sao_do_dashboard.html', assignment=None, all_branches=all_branches)
                
            current_week = f"Tuần {assignment.week_number}"
            
            # --- [VÁ LỖI AN TOÀN]: LẤY DANH SÁCH ĐỒNG ĐỘI CÙNG CA TRỰC ---
            teammates = []
            if assignment.duty_area:
                # Lấy tất cả người trực cùng ca, cùng tuần
                all_shift_assigns = db_session.query(Assignment).filter(
                    Assignment.week_number == assignment.week_number,
                    Assignment.shift == assignment.shift
                ).all()
                
                # Lọc ra những ai có trùng Object DutyArea và loại bản thân ra
                teammates = [a for a in all_shift_assigns if a.duty_area and a.duty_area.id == assignment.duty_area.id and a.red_star_id != star_id]
            # -----------------------------------------------------------
            
            # =========================================================
            # THUẬT TOÁN MỚI: TÍNH TOÁN NGÀY BẮT ĐẦU VÀ KẾT THÚC TUẦN
            # =========================================================
            start_date_str = ""
            end_date_str = ""
            
            existing_score = db_session.query(WeeklyScore).join(Branch).filter(
                WeeklyScore.week == current_week, 
                Branch.school_year_id == active_year.id
            ).first()
            
            if existing_score and existing_score.start_date: 
                start_date_str = existing_score.start_date
                end_date_str = existing_score.end_date or ""
            else:
                try:
                    import datetime as dt
                    if assignment and hasattr(assignment, 'date') and assignment.date:
                        py_date = assignment.date
                        start_date_str = py_date.strftime("%Y-%m-%d")
                        day_of_week = py_date.weekday()
                        if day_of_week <= 5: days_to_add = 5 - day_of_week
                        else: days_to_add = 6
                        end_date = py_date + dt.timedelta(days=days_to_add)
                        end_date_str = end_date.strftime("%Y-%m-%d")
                except Exception as e: pass

            if start_date_str and "-" in start_date_str:
                try: p = start_date_str.split('-'); start_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass
            if end_date_str and "-" in end_date_str:
                try: p = end_date_str.split('-'); end_date_str = f"{p[2]}/{p[1]}/{p[0]}"
                except: pass
                
            if not start_date_str: start_date_str = "..."
            if not end_date_str: end_date_str = "..."
            # =========================================================

            target_classes = []
            if assignment.duty_area:
                import json
                import os
                config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config", "class_zones.json")
                if os.path.exists(config_path):
                    with open(config_path, "r", encoding="utf-8") as f:
                        zones_map = json.load(f)
                        raw_classes = zones_map.get(assignment.duty_area.name, [])
                        class_names = [str(c).strip().upper() for c in raw_classes]
                        
                        target_classes = db_session.query(Branch).filter(
                            Branch.name.in_(class_names), Branch.school_year_id == active_year.id
                        ).all()

            violations_bank = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).all()
            
            existing_scores = {}
            scores_data_for_js = {} # [NÂNG CẤP]: Biến lưu trữ dữ liệu Ghi chú để đẩy ra JS
            
            for b in target_classes:
                sc = db_session.query(WeeklyScore).filter_by(branch_id=b.id, week=current_week).first()
                if sc: 
                    existing_scores[b.id] = sc
                    scores_data_for_js[b.id] = {
                        'note': sc.note or "",
                        'total_score': float(sc.total_score) if sc.total_score is not None else 100.0,
                        'score_tru': float(sc.score_tru) if sc.score_tru is not None else 0.0
                    }
                else:
                    scores_data_for_js[b.id] = {
                        'note': "",
                        'total_score': 100.0,
                        'score_tru': 0.0
                    }

            import json
            scores_data_json = json.dumps(scores_data_for_js)

            return render_template('sao_do_dashboard.html', 
                                   assignment=assignment, 
                                   target_classes=target_classes,
                                   all_branches=all_branches,
                                   violations_bank=violations_bank,
                                   existing_scores=existing_scores,
                                   scores_data_json=scores_data_json, # Truyền gói JSON ra ngoài
                                   current_week=current_week,
                                   teammates=teammates,
                                   start_date=start_date_str,
                                   end_date=end_date_str)
    except Exception as e:
        return f"Lỗi hệ thống Mobile: {e}"


@app.route('/submit_mobile_sao_do', methods=['POST'])
def submit_mobile_sao_do():
    if session.get('role') != 'Sao đỏ': return redirect(url_for('login'))
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            current_week = request.form.get('week_name')
            
            raw_branch_id = request.form.get('branch_id')
            if not active_year or not raw_branch_id: return redirect(url_for('mobile_sao_do'))
            branch_id = int(raw_branch_id)
            
            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            if not branch: return redirect(url_for('mobile_sao_do'))
            
            b_group = str(branch.group) if branch.group else "1"
            score = db_session.query(WeeklyScore).filter_by(branch_id=branch.id, week=current_week).first()
            old_note = score.note if score and score.note else ""
            
            max_mon, max_tot = 4, 14
            DIEM_8, DIEM_9, DIEM_10, TUAN_KHA, TUAN_TOT = 1.0, 3.0, 5.0, 20.0, 30.0
            settings = db_session.query(ScoreSettings).filter_by(school_year_id=active_year.id).first()
            if settings:
                DIEM_8, DIEM_9, DIEM_10 = float(settings.diem_8), float(settings.diem_9), float(settings.diem_10)
                TUAN_KHA, TUAN_TOT = float(settings.diem_tuan_kha), float(settings.diem_tuan_tot)
                max_tot = int(getattr(settings, 'max_diem_tot', 14))
                max_mon = int(getattr(settings, 'max_diem_mon', 4))
            
            form_rating = request.form.get('rating')
            rating = form_rating if form_rating and form_rating != 'Bình thường' else (score.week_rating if score else 'Bình thường')
            
            # 1. XỬ LÝ BAREM ĐIỂM TỐT (MÔN HỌC) TỪ APP GỬI XUỐNG
            raw_scores_json = request.form.get('raw_scores_json', '').strip()
            f10, f9, f8 = 0, 0, 0
            has_new_raw_scores = False
            
            if raw_scores_json and raw_scores_json != '[]':
                try:
                    import json
                    raw_list = json.loads(raw_scores_json)
                    if raw_list:
                        has_new_raw_scores = True
                        db_session.query(RawScore).filter_by(week=current_week, branch_name=branch.name).delete()
                        for item in raw_list:
                            subj = str(item.get("subj", "")).strip()
                            if not subj: continue
                            i10 = int(item.get("c10", 0) or 0); i9 = int(item.get("c9", 0) or 0); i8 = int(item.get("c8", 0) or 0)
                            if "1" in b_group: i8 = 0
                            db_session.add(RawScore(week=current_week, branch_name=branch.name, subject=subj, c10=i10, c9=i9, c8=i8))

                        remaining = max_tot
                        for item in raw_list:
                            subj = str(item.get("subj", "")).strip()
                            if not subj: continue
                            i10 = int(item.get("c10", 0) or 0); i9 = int(item.get("c9", 0) or 0); i8 = int(item.get("c8", 0) or 0)
                            if "1" in b_group: i8 = 0
                            
                            k10 = min(i10, max_mon); k9 = min(i9, max_mon - k10); k8 = min(i8, max_mon - k10 - k9)
                            a10 = min(k10, remaining); remaining -= a10
                            a9 = min(k9, remaining); remaining -= a9
                            a8 = min(k8, remaining); remaining -= a8
                            f10 += a10; f9 += a9; f8 += a8
                        db_session.flush()
                except Exception as ex: 
                    print("Lỗi xử lý JSON Barem Mobile:", ex)
            
            if not has_new_raw_scores:
                f10 = score.count_10 if score else 0
                f9 = score.count_9 if score else 0
                f8 = score.count_8 if score else 0
                if "1" in b_group: f8 = 0

            manual_note = request.form.get('manual_note', '').strip()
            
            # 2. THU THẬP VÀ ĐỒNG BỘ TOÀN BỘ CÁC LỖI VI PHẠM (WEEKLY VIOLATIONS)
            all_categories = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).all()
            new_violations = []
            diem_tru_auto = 0.0
            
            # Mảng gom ghi chú để hiển thị ra bảng điểm Web
            note_fragments = []
            if manual_note:
                note_fragments.append(manual_note)

            for cat in all_categories:
                qty_raw = request.form.get(f'viol_{cat.id}', '0')
                try: qty = int(qty_raw)
                except: qty = 0
                
                if qty > 0:
                    stu_name = request.form.get(f'student_{cat.id}', '').strip()
                    if getattr(cat, 'point_type', 'Điểm trừ') != 'Điểm cộng':
                        diem_tru_auto += float(cat.penalty_points * qty)
                    
                    new_violations.append({
                        'violation_id': cat.id, 
                        'quantity': qty, 
                        'student_name': stu_name if stu_name else None
                    })
                    
                    # Tạo đoạn ghi chú chuẩn hóa dạng "Tên lỗi xSL [Tên học sinh]"
                    if stu_name:
                        note_fragments.append(f"{cat.name} x{qty} [{stu_name}]")
                    else:
                        note_fragments.append(f"{cat.name} x{qty}")

            # Kết hợp ghi chú cũ và ghi chú mới từ App
            combined_note_raw = " ; ".join([f for f in [old_note, " ; ".join(note_fragments)] if f])

            # 3. THUẬT TOÁN GỘP NHÓM THÔNG MINH (TRÁNH TRÙNG LẶP LỖI TRONG TUẦN)
            import re
            sorted_cats = sorted(all_categories, key=lambda x: len(x.name), reverse=True)
            parsed_errors = {}
            
            parts = re.split(r'[,;+\n](?![^\[]*\])(?![^\(]*\))', combined_note_raw)
            for part in parts:
                part_clean = part.strip()
                if not part_clean: continue
                
                stu_name_raw = ""
                match_stu = re.search(r'\[(.*?)\]', part_clean)
                if match_stu: stu_name_raw = match_stu.group(1).strip()
                
                stu_name_normalized = " ".join(stu_name_raw.split()).title() if stu_name_raw else ""
                stu_name_key = stu_name_normalized.lower()
                
                matched = False
                for cat in sorted_cats:
                    if cat.name.lower() in part_clean.lower():
                        match_qty = re.search(r'(?:x|:|-)\s*(\d+)', part_clean.lower())
                        qty = int(match_qty.group(1)) if match_qty else 1
                        
                        key = (cat.name, stu_name_key, stu_name_normalized)
                        if key not in parsed_errors:
                            parsed_errors[key] = 0
                        parsed_errors[key] += qty
                        matched = True
                        break
                
                if not matched:
                    parsed_errors[("MANUAL", part_clean.lower(), part_clean)] = 1
                    
            final_parts = []
            diem_tru_final = 0.0
            
            for (cat_name, stu_key, stu_display), qty in parsed_errors.items():
                if cat_name == "MANUAL":
                    final_parts.append(stu_display)
                else:
                    if stu_display: 
                        final_parts.append(f"{cat_name} x{qty} [{stu_display}]")
                    else: 
                        final_parts.append(f"{cat_name} x{qty}")
                    
                    for cat in sorted_cats:
                        if cat.name == cat_name and getattr(cat, 'point_type', 'Điểm trừ') != 'Điểm cộng':
                            diem_tru_final += float(cat.penalty_points * qty)
                            break
                            
            final_note = " ; ".join(final_parts)

            # 4. TÍNH TOÁN ĐIỂM SỐ HOÀN CHỈNH
            if "1" in b_group: diem_quy_uoc = (f9 * DIEM_9) + (f10 * DIEM_10)
            elif "2" in b_group: diem_quy_uoc = (f8 * DIEM_8) + (f9 * DIEM_9) + (f10 * DIEM_10)
            else: diem_quy_uoc = (f9 * DIEM_9) + (f10 * DIEM_10)

            diem_xep_loai = 0.0
            if rating == "Tuần Tốt": diem_xep_loai = TUAN_TOT
            elif rating == "Tuần Khá": diem_xep_loai = TUAN_KHA

            truc = float(score.score_truc) if score and score.score_truc else 100.0
            cong = float(score.score_cong) if score and score.score_cong else 0.0
            
            # Công thức tổng điểm chuẩn
            total_val = truc + diem_xep_loai + diem_quy_uoc + cong - diem_tru_final

            # 5. LƯU VÀO CƠ SỞ DỮ LIỆU
            if score:
                score.week_rating = rating
                score.count_8 = f8; score.count_9 = f9; score.count_10 = f10
                score.note = final_note
                score.score_tru = diem_tru_final
                score.total_score = total_val
            else:
                score = WeeklyScore(
                    branch_id=branch.id, week=current_week, week_rating=rating,
                    count_8=f8, count_9=f9, count_10=f10,
                    score_truc=truc, score_cong=cong, score_tru=diem_tru_final,
                    note=final_note, total_score=total_val
                )
                db_session.add(score)
                db_session.flush()

            # Xóa các lỗi cũ thuộc tuần của lớp này để ghi đè danh sách lỗi chi tiết mới nhất
            db_session.query(WeeklyViolation).filter_by(weekly_score_id=score.id).delete()
            for v in new_violations:
                # Tìm lại ID của category từ tên hoặc lưu trực tiếp
                db_session.add(WeeklyViolation(
                    weekly_score_id=score.id, 
                    violation_id=v['violation_id'], 
                    quantity=v['quantity'], 
                    student_name=v['student_name']
                ))

            log_system_action("MOBILE SAO ĐỎ", f"SD{session.get('username')} đã chấm điểm và cập nhật lớp {branch.name}.")
            flash(f"Đã nộp điểm và đồng bộ vào hệ thống cho lớp {branch.name} thành công!", "success")
            return redirect(url_for('mobile_sao_do'))
            
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi: {e}", "error")
        return redirect(url_for('mobile_sao_do'))
    
# ==========================================
# MODULE: TRỢ LÝ AI PHÂN TÍCH VÀ VIẾT BÁO CÁO TUẦN
# ==========================================
@app.route('/api/ai_weekly_report/<week_name>')
def api_ai_weekly_report(week_name):
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year: return {"error": "Chưa có năm học kích hoạt!"}
            
            # 1. Lấy toàn bộ điểm số của tuần
            scores = db_session.query(WeeklyScore).join(Branch).filter(
                WeeklyScore.week == week_name,
                Branch.school_year_id == active_year.id
            ).all()
            
            if not scores:
                return {"error": f"Chưa có dữ liệu điểm của {week_name} để phân tích!"}
                
            # 2. Tổng hợp dữ liệu thô để đưa cho AI
            total_classes = len(scores)
            top_classes = sorted(scores, key=lambda x: float(x.total_score or 0), reverse=True)[:3]
            bottom_classes = sorted(scores, key=lambda x: float(x.total_score or 0))[:3]
            
            summary_lines = []
            for sc in scores:
                b_name = sc.branch.name
                tot = sc.total_score
                rating = sc.week_rating
                note = sc.note or "Không có"
                summary_lines.append(f"- Lớp {b_name}: Tổng điểm {tot}, Xếp loại {rating}, Ghi chú: {note}")
                
            data_context = "\n".join(summary_lines)
            
            # 3. Đọc cấu hình Groq API Key từ file config
            api_key = ""
            config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config", "groq_settings.json")
            if os.path.exists(config_path):
                with open(config_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    api_key = cfg.get("api_key", "")
                    
            if not api_key:
                return {"error": "Chưa cấu hình Groq API Key trong hệ thống!"}
                
            # 4. Gọi API Groq AI (Sử dụng thư viện requests chuẩn Python)
            import requests
            url = "https://api.groq.com/openai/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            
            prompt = f"""
            Bạn là một Trợ lý hành chính Đoàn trường THPT Thanh Hòa. Dựa trên dữ liệu tổng kết nề nếp và thi đua của {week_name} dưới đây, hãy viết một bản báo cáo sơ kết tuần ngắn gọn, trang trọng, mang văn phong nhà trường để gửi Ban Giám hiệu hoặc đăng nhóm Zalo giáo viên:
            
            DỮ LIỆU THI ĐUA:
            {data_context}
            
            YÊU CẦU BÁO CÁO:
            1. Tiêu đề trang trọng.
            2. Đánh giá chung về tình hình nề nếp toàn trường trong tuần.
            3. Tuyên dương cụ thể các lớp dẫn đầu có thành tích xuất sắc.
            4. Nhắc nhở, rút kinh nghiệm đối với các lớp còn nhiều lỗi vi phạm (đi học muộn, vắng, v.v.).
            5. Định hướng nề nếp cho tuần tiếp theo.
            Trình bày bằng các gạch đầu dòng rõ ràng, lịch sự, chuyên nghiệp.
            """
            
            payload = {
                "model": "openai/gpt-oss-120b",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7
            }
            
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            if response.status_code == 200:
                res_json = response.json()
                ai_text = res_json['choices'][0]['message']['content']
                return {"success": True, "report": ai_text}
            else:
                return {"error": f"Lỗi kết nối AI API: {response.text}"}
                
    except Exception as e:
        return {"error": str(e)}
# ==========================================
# MODULE: TRA CỨU SỔ ĐEN CÁ NHÂN TOÀN TRƯỜNG
# ==========================================
@app.route('/blacklist', methods=['GET'])
def blacklist():
    # Chỉ Admin/Bí thư mới được xem Sổ đen toàn trường
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Bí thư', 'Bí thư Đoàn trường']:
        flash("Bạn không có quyền truy cập Sổ đen!", "error")
        return redirect(url_for('dashboard'))
        
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('dashboard'))
                
            branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()
            
            # Lấy từ khóa tìm kiếm từ giao diện
            search_name = request.args.get('search_name', '').strip()
            search_branch = request.args.get('search_branch', '')
            
            # Truy vấn: Tìm tất cả các Lỗi vi phạm CÓ GHI TÊN học sinh
            query = db_session.query(
                WeeklyViolation, WeeklyScore, Branch, ViolationCategory
            ).join(
                WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id
            ).join(
                Branch, WeeklyScore.branch_id == Branch.id
            ).join(
                ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id
            ).filter(
                Branch.school_year_id == active_year.id,
                WeeklyViolation.student_name != None,
                WeeklyViolation.student_name != ''
            )
            
            # Áp dụng bộ lọc nếu có
            if search_name:
                query = query.filter(WeeklyViolation.student_name.ilike(f"%{search_name}%"))
            if search_branch:
                query = query.filter(Branch.id == search_branch)
                
            # Sắp xếp theo thứ tự Tuần mới nhất giảm dần
            results = query.order_by(WeeklyScore.id.desc(), Branch.name).all()
            
            violation_data = []
            for v, sc, b, c in results:
                violation_data.append({
                    'week': sc.week,
                    'branch_name': b.name,
                    'student_name': v.student_name,
                    'violation_name': c.name,
                    'quantity': v.quantity,
                    'penalty': float(c.penalty_points * v.quantity) if getattr(c, 'point_type', 'Điểm trừ') != 'Điểm cộng' else 0
                })
                
            return render_template('blacklist.html', 
                                   branches=branches, 
                                   violations=violation_data,
                                   search_name=search_name,
                                   search_branch=search_branch)
    except Exception as e:
        flash(f"Lỗi tải sổ đen: {e}", "error")
        return redirect(url_for('dashboard'))
    
from flask import send_file

@app.route('/manifest.json')
def serve_manifest():
    return send_file('static/manifest.json', mimetype='application/manifest+json')

@app.route('/sw.js')
def serve_sw():
    return send_file('static/sw.js', mimetype='application/javascript')
from flask import send_from_directory

# ==========================================
# CẤU HÌNH APP HÓA (PWA) CHO ĐIỆN THOẠI
# ==========================================
@app.route('/sw.js')
def service_worker():
    # Phục vụ tệp sw.js từ thư mục static ra thẳng thư mục gốc mà không bị redirect
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')

@app.route('/manifest.json')
def manifest():
    # Tương tự với file cấu hình PWA
    return send_from_directory('static', 'manifest.json', mimetype='application/json')

@app.route('/sao_do_quick_submit_form', methods=['POST'])
def sao_do_quick_submit_form():
    if session.get('role') != 'Sao đỏ': return redirect(url_for('login'))
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            week_name = request.form.get('week_name')
            
            raw_branch_id = request.form.get('branch_id')
            raw_viol_id = request.form.get('violation_id')
            student_name = request.form.get('student_name', '').strip()
            
            if not all([week_name, raw_branch_id, raw_viol_id]):
                flash("Lỗi: Vui lòng chọn đầy đủ Lớp và Lỗi vi phạm!", "error")
                return redirect(url_for('mobile_sao_do'))
                
            branch_id = int(raw_branch_id)
            violation_id = int(raw_viol_id)
                
            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            violation = db_session.query(ViolationCategory).filter_by(id=violation_id).first()
            if not branch or not violation:
                flash("Lỗi: Dữ liệu Lớp hoặc Lỗi không hợp lệ!", "error")
                return redirect(url_for('mobile_sao_do'))
                
            score = db_session.query(WeeklyScore).filter_by(branch_id=branch.id, week=week_name).first()
            if not score:
                score = WeeklyScore(
                    branch_id=branch.id, week=week_name, week_rating='Bình thường', 
                    count_8=0, count_9=0, count_10=0, score_truc=100.0, 
                    score_cong=0.0, score_tru=0.0, note='', total_score=100.0
                )
                db_session.add(score)
                db_session.flush() 
                
            old_note = score.note if score and score.note else ""
            
            # Thêm lỗi mới vào chuỗi
            if student_name:
                # Chuẩn hóa tên hiển thị ngay khi nhập
                std_clean = " ".join(student_name.split()).title()
                note_add = f"{violation.name} x1 [{std_clean}]"
            else:
                note_add = f"{violation.name} x1"
                
            raw_combined_note = f"{old_note} ; {note_add}" if old_note else note_add
            
            # Gộp nhóm thông minh ngay lập tức
            import re
            all_categories = db_session.query(ViolationCategory).filter_by(school_year_id=active_year.id).all()
            sorted_cats = sorted(all_categories, key=lambda x: len(x.name), reverse=True)
            parsed_errors = {}
            
            parts = re.split(r'[,;+\n](?![^\[]*\])(?![^\(]*\))', raw_combined_note)
            for part in parts:
                part_clean = part.strip()
                if not part_clean: continue
                
                stu_name_raw = ""
                match_stu = re.search(r'\[(.*?)\]', part_clean)
                if match_stu: stu_name_raw = match_stu.group(1).strip()
                
                stu_name_normalized = " ".join(stu_name_raw.split()).title() if stu_name_raw else ""
                stu_name_key = stu_name_normalized.lower()
                
                matched = False
                for cat in sorted_cats:
                    if cat.name.lower() in part_clean.lower():
                        match_qty = re.search(r'(?:x|:|-)\s*(\d+)', part_clean.lower())
                        qty = int(match_qty.group(1)) if match_qty else 1
                        
                        key = (cat.name, stu_name_key, stu_name_normalized)
                        if key not in parsed_errors:
                            parsed_errors[key] = 0
                        parsed_errors[key] += qty
                        matched = True
                        break
                
                if not matched:
                    parsed_errors[("MANUAL", part_clean.lower(), part_clean)] = 1
                    
            final_parts = []
            diem_tru_auto = 0.0
            
            for (cat_name, stu_key, stu_display), qty in parsed_errors.items():
                if cat_name == "MANUAL":
                    final_parts.append(stu_display)
                else:
                    if stu_display: 
                        final_parts.append(f"{cat_name} x{qty} [{stu_display}]")
                    else: 
                        final_parts.append(f"{cat_name} x{qty}")
                        
                    for cat in sorted_cats:
                        if cat.name == cat_name and getattr(cat, 'point_type', 'Điểm trừ') != 'Điểm cộng':
                            diem_tru_auto += float(cat.penalty_points * qty)
                            break
                            
            score.note = " ; ".join(final_parts)
            score.score_tru = diem_tru_auto
            
            truc = float(score.score_truc) if score.score_truc else 100.0
            cong = float(score.score_cong) if score.score_cong else 0.0
            b_group = str(branch.group) if branch.group else "1"
            DIEM_8, DIEM_9, DIEM_10, TUAN_KHA, TUAN_TOT = 1.0, 3.0, 5.0, 20.0, 30.0
            settings = db_session.query(ScoreSettings).filter_by(school_year_id=active_year.id).first()
            if settings:
                DIEM_8, DIEM_9, DIEM_10 = float(settings.diem_8), float(settings.diem_9), float(settings.diem_10)
                TUAN_KHA, TUAN_TOT = float(settings.diem_tuan_kha), float(settings.diem_tuan_tot)
            
            if "1" in b_group: diem_quy_uoc = (int(score.count_9 or 0) * DIEM_9) + (int(score.count_10 or 0) * DIEM_10)
            elif "2" in b_group: diem_quy_uoc = (int(score.count_8 or 0) * DIEM_8) + (int(score.count_9 or 0) * DIEM_9) + (int(score.count_10 or 0) * DIEM_10)
            else: diem_quy_uoc = (int(score.count_9 or 0) * DIEM_9) + (int(score.count_10 or 0) * DIEM_10)

            diem_xep_loai = 0.0
            if score.week_rating == "Tuần Tốt": diem_xep_loai = TUAN_TOT
            elif score.week_rating == "Tuần Khá": diem_xep_loai = TUAN_KHA
            
            score.total_score = truc + diem_xep_loai + diem_quy_uoc + cong - diem_tru_auto

            log_system_action("MOBILE TRỰC CỔNG", f"SD{session.get('username')} ghi nhận nhanh {branch.name}: {violation.name} - {student_name}")
            flash(f"⚡ Đã ghi nhận lỗi của {branch.name} vào Sổ đen thành công!", "success")
            return redirect(url_for('mobile_sao_do'))
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi hệ thống: {e}", "error")
        return redirect(url_for('mobile_sao_do'))
# =====================================================================
# MODULE: PHIẾU PHÂN TÍCH CHUYÊN SÂU LỚP HỌC (DÀNH CHO HỌP GVCN)
# =====================================================================
@app.route('/class_monthly_analysis', methods=['GET', 'POST'])
def class_monthly_analysis():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('dashboard'))

            # Lấy danh sách tháng đã chốt, sắp xếp theo thứ tự năm học
            months_db = db_session.query(MonthlyRecord.month_name).filter(
                MonthlyRecord.school_year_id == active_year.id,
                MonthlyRecord.month_name.like('Tháng%')
            ).distinct().all()
            
            school_order = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
            raw_months = [m[0] for m in months_db if m[0]]
            available_months = sorted(raw_months, key=lambda x: school_order.index(x) if x in school_order else 99)
            
            branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()

            selected_month = request.args.get('month', available_months[-1] if available_months else "")
            selected_branch_id = request.args.get('branch_id', type=int)

            analysis_data = None

            if selected_month and selected_branch_id:
                branch = db_session.query(Branch).filter_by(id=selected_branch_id).first()
                month_rec = db_session.query(MonthlyRecord).filter_by(
                    school_year_id=active_year.id, 
                    month_name=selected_month, 
                    branch_id=selected_branch_id
                ).first()

                if month_rec and month_rec.weeks_used:
                    valid_weeks = [w.strip() for w in month_rec.weeks_used.split(',') if w.strip()]
                    
                    # 1. TRÍCH XUẤT DỮ LIỆU NỀ NẾP (KỶ LUẬT)
                    violations = db_session.query(WeeklyViolation, ViolationCategory).join(
                        ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id
                    ).join(
                        WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id
                    ).filter(
                        WeeklyScore.week.in_(valid_weeks),
                        WeeklyScore.branch_id == selected_branch_id
                    ).all()

                    viol_summary = {}
                    student_issues = {}
                    student_bonus = {} 

                    for v, cat in violations:
                        qty = v.quantity or 1
                        # Gom nhóm lỗi
                        if getattr(cat, 'point_type', 'Điểm trừ') == 'Điểm trừ':
                            viol_summary[cat.name] = viol_summary.get(cat.name, 0) + qty
                            
                            # Gom nhóm cá nhân vi phạm
                            if v.student_name:
                                names = [n.strip().title() for n in str(v.student_name).replace(';', ',').split(',') if n.strip()]
                                for name in names:
                                    if name not in student_issues: student_issues[name] = []
                                    student_issues[name].append(f"{cat.name} (x{qty})")
                        else:
                            # Điểm cộng (Học sinh xuất sắc)
                            if v.student_name:
                                names = [n.strip().title() for n in str(v.student_name).replace(';', ',').split(',') if n.strip()]
                                for name in names:
                                    student_bonus[name] = student_bonus.get(name, 0) + qty

                    # [ĐÃ THÁO KHÓA]: Hiển thị TOÀN BỘ lỗi, sắp xếp từ nhiều đến ít (Không giới hạn Top 3 nữa)
                    all_violations_sorted = sorted(viol_summary.items(), key=lambda x: x[1], reverse=True)
                    
                    # [ĐÃ THÁO KHÓA]: Hiển thị TOÀN BỘ học sinh có tên trong sổ đen (Kể cả 1 lần)
                    all_violators_sorted = {k: v for k, v in sorted(student_issues.items(), key=lambda item: len(item[1]), reverse=True)}

                    # 2. TRÍCH XUẤT DỮ LIỆU HỌC TẬP (RAW SCORES)
                    raw_scores = db_session.query(RawScore).filter(
                        RawScore.week.in_(valid_weeks),
                        RawScore.branch_name == branch.name
                    ).all()

                    subject_scores = {}
                    total_good_points = 0
                    for rs in raw_scores:
                        subj = rs.subject.strip()
                        c10, c9, c8 = int(rs.c10 or 0), int(rs.c9 or 0), int(rs.c8 or 0)
                        if "1" in str(branch.group): c8 = 0
                        
                        points = c10 + c9 + c8
                        total_good_points += points
                        subject_scores[subj] = subject_scores.get(subj, 0) + points

                    # Chỉ lấy Top 3 môn tốt nhất, bỏ phần quét môn yếu kém
                    top_subjects = sorted(subject_scores.items(), key=lambda x: x[1], reverse=True)[:3]

                    # Thống kê Tuần Tốt / Khá
                    scores = db_session.query(WeeklyScore).filter(
                        WeeklyScore.week.in_(valid_weeks),
                        WeeklyScore.branch_id == selected_branch_id
                    ).all()
                    
                    tuan_tot = sum(1 for s in scores if s.week_rating == 'Tuần Tốt')
                    tuan_kha = sum(1 for s in scores if s.week_rating == 'Tuần Khá')

                    analysis_data = {
                        "branch": branch,
                        "month_name": selected_month,
                        "weeks_used": month_rec.weeks_used,
                        "total_score": month_rec.total_score,
                        "rank": month_rec.rank,
                        "top_violations": all_violations_sorted, 
                        "frequent_violators": all_violators_sorted, 
                        "student_bonus": sorted(student_bonus.items(), key=lambda x: x[1], reverse=True),
                        "total_good_points": total_good_points,
                        "top_subjects": top_subjects,
                        "tuan_tot": tuan_tot,
                        "tuan_kha": tuan_kha
                    }
            return render_template('class_monthly_analysis.html', 
                                   available_months=available_months, 
                                   branches=branches, 
                                   selected_month=selected_month, 
                                   selected_branch_id=selected_branch_id,
                                   data=analysis_data)
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải trang phân tích: {e}", "error")
        return redirect(url_for('dashboard'))
# =====================================================================
# MODULE: XUẤT PDF PHIẾU PHÂN TÍCH TOÀN BỘ CÁC LỚP
# =====================================================================
@app.route('/export_all_monthly_analysis', methods=['GET'])
def export_all_monthly_analysis():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                return "Chưa có năm học kích hoạt!"

            selected_month = request.args.get('month')
            if not selected_month:
                return "Lỗi: Vui lòng chọn tháng trước khi xuất!"

            branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).order_by(Branch.name).all()
            all_data = []

            for branch in branches:
                month_rec = db_session.query(MonthlyRecord).filter_by(
                    school_year_id=active_year.id, 
                    month_name=selected_month, 
                    branch_id=branch.id
                ).first()

                if month_rec and month_rec.weeks_used:
                    valid_weeks = [w.strip() for w in month_rec.weeks_used.split(',') if w.strip()]
                    
                    # NỀ NẾP
                    violations = db_session.query(WeeklyViolation, ViolationCategory).join(
                        ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id
                    ).join(
                        WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id
                    ).filter(WeeklyScore.week.in_(valid_weeks), WeeklyScore.branch_id == branch.id).all()

                    viol_summary = {}; student_issues = {}; student_bonus = {} 
                    for v, cat in violations:
                        qty = v.quantity or 1
                        if getattr(cat, 'point_type', 'Điểm trừ') == 'Điểm trừ':
                            viol_summary[cat.name] = viol_summary.get(cat.name, 0) + qty
                            if v.student_name:
                                names = [n.strip().title() for n in str(v.student_name).replace(';', ',').split(',') if n.strip()]
                                for name in names:
                                    if name not in student_issues: student_issues[name] = []
                                    student_issues[name].append(f"{cat.name} (x{qty})")
                        else:
                            if v.student_name:
                                names = [n.strip().title() for n in str(v.student_name).replace(';', ',').split(',') if n.strip()]
                                for name in names: student_bonus[name] = student_bonus.get(name, 0) + qty

                    all_violations_sorted = sorted(viol_summary.items(), key=lambda x: x[1], reverse=True)
                    all_violators_sorted = {k: v for k, v in sorted(student_issues.items(), key=lambda item: len(item[1]), reverse=True)}

                    # HỌC TẬP
                    raw_scores = db_session.query(RawScore).filter(RawScore.week.in_(valid_weeks), RawScore.branch_name == branch.name).all()
                    subject_scores = {}; total_good_points = 0
                    for rs in raw_scores:
                        subj = rs.subject.strip()
                        c10, c9, c8 = int(rs.c10 or 0), int(rs.c9 or 0), int(rs.c8 or 0)
                        if "1" in str(branch.group): c8 = 0
                        points = c10 + c9 + c8
                        total_good_points += points
                        subject_scores[subj] = subject_scores.get(subj, 0) + points

                    top_subjects = sorted(subject_scores.items(), key=lambda x: x[1], reverse=True)[:3]
                    scores = db_session.query(WeeklyScore).filter(WeeklyScore.week.in_(valid_weeks), WeeklyScore.branch_id == branch.id).all()
                    tuan_tot = sum(1 for s in scores if s.week_rating == 'Tuần Tốt')
                    tuan_kha = sum(1 for s in scores if s.week_rating == 'Tuần Khá')

                    all_data.append({
                        "branch": branch, "month_name": selected_month, "weeks_used": month_rec.weeks_used,
                        "total_score": month_rec.total_score, "rank": month_rec.rank,
                        "top_violations": all_violations_sorted, "frequent_violators": all_violators_sorted, 
                        "student_bonus": sorted(student_bonus.items(), key=lambda x: x[1], reverse=True),
                        "total_good_points": total_good_points, "top_subjects": top_subjects,
                        "tuan_tot": tuan_tot, "tuan_kha": tuan_kha
                    })

            return render_template('class_monthly_analysis_all.html', all_data=all_data, selected_month=selected_month)
    except Exception as e:
        import traceback; traceback.print_exc()
        return f"<h1>Lỗi xuất PDF toàn trường: {e}</h1>"
# =====================================================================
# MODULE: PHIẾU PHÂN TÍCH ĐÁNH GIÁ CUỐI HỌC KỲ (HỌP GVCN)
# =====================================================================
@app.route('/class_semester_analysis', methods=['GET', 'POST'])
def class_semester_analysis():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('dashboard'))

            # [ĐÃ SỬA]: Quét danh sách Học kỳ từ bảng MonthlyRecord giống hệt logic của đồng chí
            semesters_db = db_session.query(MonthlyRecord.month_name).filter(
                MonthlyRecord.school_year_id == active_year.id,
                MonthlyRecord.month_name.like('Học kỳ%')
            ).distinct().all()
            
            available_semesters = [s[0] for s in semesters_db if s[0]]
            if not available_semesters: available_semesters = ["Học kỳ 1", "Học kỳ 2"]

            branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()

            selected_semester = request.args.get('semester', available_semesters[0])
            selected_branch_id = request.args.get('branch_id', type=int)

            analysis_data = None

            if selected_semester and selected_branch_id:
                branch = db_session.query(Branch).filter_by(id=selected_branch_id).first()
                
                # [ĐÃ SỬA]: Truy vấn dữ liệu Học kỳ từ bảng MonthlyRecord
                semester_rec = db_session.query(MonthlyRecord).filter_by(
                    school_year_id=active_year.id, 
                    month_name=selected_semester, 
                    branch_id=selected_branch_id
                ).first()

                if semester_rec and semester_rec.weeks_used:
                    # Trong bảng này, weeks_used đang chứa chuỗi các tháng (VD: "Tháng 9, Tháng 10")
                    valid_months = [m.strip() for m in semester_rec.weeks_used.split(',') if m.strip()]
                    
                    # --- TIÊU CHÍ 1: KẾT QUẢ THI ĐUA CHUNG ---
                    monthly_records = db_session.query(MonthlyRecord).filter(
                        MonthlyRecord.school_year_id == active_year.id,
                        MonthlyRecord.branch_id == selected_branch_id,
                        MonthlyRecord.month_name.in_(valid_months)
                    ).all()
                    
                    school_order = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
                    monthly_records = sorted(monthly_records, key=lambda x: school_order.index(x.month_name) if x.month_name in school_order else 99)
                    
                    monthly_stats = [{"month": m.month_name, "score": float(m.total_score), "rank": m.rank} for m in monthly_records]
                    avg_score = round(sum(m['score'] for m in monthly_stats) / len(monthly_stats), 2) if monthly_stats else 0
                    
                    progress_note = "Chưa đủ dữ liệu để so sánh"
                    if len(monthly_stats) >= 2:
                        first_m = monthly_stats[0]
                        last_m = monthly_stats[-1]
                        if last_m['rank'] < first_m['rank']:
                            progress_note = f"Tiến bộ rõ rệt (Từ Hạng {first_m['rank']} vươn lên Hạng {last_m['rank']})"
                        elif last_m['rank'] > first_m['rank']:
                            progress_note = f"Có dấu hiệu giảm sút (Từ Hạng {first_m['rank']} rớt xuống Hạng {last_m['rank']})"
                        else:
                            progress_note = f"Duy trì kết quả ổn định ở Hạng {last_m['rank']}"

                    valid_weeks = []
                    for m_rec in monthly_records:
                        if m_rec.weeks_used: valid_weeks.extend([w.strip() for w in m_rec.weeks_used.split(',') if w.strip()])
                    valid_weeks = list(set(valid_weeks))

                    # --- TIÊU CHÍ 2: NỀN NẾP & KỶ LUẬT ---
                    violations = db_session.query(WeeklyViolation, ViolationCategory).join(
                        ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id
                    ).join(
                        WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id
                    ).filter(WeeklyScore.week.in_(valid_weeks), WeeklyScore.branch_id == selected_branch_id).all()

                    viol_summary = {}
                    student_issues = {}
                    total_violations = 0
                    
                    for v, cat in violations:
                        qty = v.quantity or 1
                        if getattr(cat, 'point_type', 'Điểm trừ') == 'Điểm trừ':
                            total_violations += qty
                            viol_summary[cat.name] = viol_summary.get(cat.name, 0) + qty
                            
                            if v.student_name:
                                names = [n.strip().title() for n in str(v.student_name).replace(';', ',').split(',') if n.strip()]
                                for name in names:
                                    if name not in student_issues: student_issues[name] = []
                                    student_issues[name].append(f"{cat.name} (x{qty})")

                    top_violations = sorted(viol_summary.items(), key=lambda x: x[1], reverse=True)
                    severe_violators = {k: v for k, v in sorted(student_issues.items(), key=lambda item: len(item[1]), reverse=True) if len(v) >= 3} 

                    analysis_data = {
                        "branch": branch,
                        "semester_name": selected_semester,
                        "months_used": semester_rec.weeks_used, 
                        "total_score": semester_rec.total_score,
                        "rank": semester_rec.rank,
                        "monthly_stats": monthly_stats,
                        "avg_score": avg_score,
                        "progress_note": progress_note,
                        "total_violations": total_violations,
                        "top_violations": top_violations,
                        "severe_violators": severe_violators
                    }

            return render_template('class_semester_analysis.html', 
                                   available_semesters=available_semesters, 
                                   branches=branches, 
                                   selected_semester=selected_semester, 
                                   selected_branch_id=selected_branch_id,
                                   data=analysis_data)
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải trang phân tích học kỳ: {e}", "error")
        return redirect(url_for('dashboard'))
    
# =====================================================================
# MODULE: BÁO CÁO TỔNG QUAN TOÀN TRƯỜNG (DÀNH CHO BÍ THƯ / HIỆU TRƯỞNG)
# =====================================================================
@app.route('/school_monthly_analysis', methods=['GET'])
def school_monthly_analysis():
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                flash("Chưa có năm học kích hoạt!", "error")
                return redirect(url_for('dashboard'))

            months_db = db_session.query(MonthlyRecord.month_name).filter(
                MonthlyRecord.school_year_id == active_year.id,
                MonthlyRecord.month_name.like('Tháng%')
            ).distinct().all()
            
            school_order = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]
            raw_months = [m[0] for m in months_db if m[0]]
            available_months = sorted(raw_months, key=lambda x: school_order.index(x) if x in school_order else 99)

            selected_month = request.args.get('month', available_months[-1] if available_months else "")
            analysis_data = None

            if selected_month:
                # 1. Lấy toàn bộ danh sách Chi đoàn của năm học hiện tại làm bản đồ tra cứu chuẩn tuyệt đối
                all_branches = db_session.query(Branch).filter_by(school_year_id=active_year.id).all()
                branch_map = {b.id: b for b in all_branches}

                # 2. Lấy tất cả các bản ghi điểm tháng đã chốt
                records = db_session.query(MonthlyRecord).filter(
                    MonthlyRecord.school_year_id == active_year.id, 
                    MonthlyRecord.month_name == selected_month
                ).order_by(MonthlyRecord.rank).all()
                
                processed_records = []
                valid_weeks = set()
                
                for m_rec in records:
                    # Tra cứu thông tin lớp an toàn tuyệt đối qua branch_id
                    b_rec = branch_map.get(m_rec.branch_id)
                    
                    b_name = str(b_rec.name).strip() if b_rec and b_rec.name else f"Chi đoàn {m_rec.branch_id}"
                    b_gvcn = str(b_rec.gvcn).strip() if b_rec and b_rec.gvcn else "Chưa cập nhật GVCN"

                    processed_records.append({
                        "rank": m_rec.rank,
                        "total_score": round(float(m_rec.total_score or 0), 1),
                        "branch_name": b_name,
                        "branch_gvcn": b_gvcn
                    })
                    if m_rec.weeks_used:
                        valid_weeks.update([w.strip() for w in m_rec.weeks_used.split(',') if w.strip()])
                        
                valid_weeks = list(valid_weeks)

                if processed_records:
                    top_classes = processed_records[:5] 
                    bottom_classes = processed_records[-5:] if len(processed_records) > 5 else [] 
                    bottom_classes.reverse() 

                    violations = db_session.query(WeeklyViolation, ViolationCategory).join(
                        ViolationCategory, WeeklyViolation.violation_id == ViolationCategory.id
                    ).join(
                        WeeklyScore, WeeklyViolation.weekly_score_id == WeeklyScore.id
                    ).filter(WeeklyScore.week.in_(valid_weeks)).all()

                    viol_summary = {}
                    total_violations = 0
                    for v, cat in violations:
                        qty = v.quantity or 1
                        if getattr(cat, 'point_type', 'Điểm trừ') == 'Điểm trừ':
                            viol_summary[cat.name] = viol_summary.get(cat.name, 0) + qty
                            total_violations += qty

                    top_violations_school = sorted(viol_summary.items(), key=lambda x: x[1], reverse=True)[:5] 

                    branch_group_map = {b.name: b.group for b in all_branches}

                    raw_scores = db_session.query(RawScore).filter(RawScore.week.in_(valid_weeks)).all()
                    subject_scores = {}
                    total_good_points = 0
                    
                    for rs in raw_scores:
                        subj = rs.subject.strip()
                        b_name = rs.branch_name.strip()
                        c10, c9, c8 = int(rs.c10 or 0), int(rs.c9 or 0), int(rs.c8 or 0)
                        
                        grp = branch_group_map.get(b_name, "")
                        if "1" in str(grp): c8 = 0 
                        
                        points = c10 + c9 + c8
                        total_good_points += points
                        subject_scores[subj] = subject_scores.get(subj, 0) + points

                    top_subjects_school = sorted(subject_scores.items(), key=lambda x: x[1], reverse=True)[:5]

                    analysis_data = {
                        "month_name": selected_month,
                        "weeks_used": ", ".join(valid_weeks),
                        "top_classes": top_classes,
                        "bottom_classes": bottom_classes,
                        "total_violations": total_violations,
                        "top_violations_school": top_violations_school,
                        "total_good_points": total_good_points,
                        "top_subjects_school": top_subjects_school,
                        "total_classes": len(processed_records)
                    }

            return render_template('school_monthly_analysis.html', 
                                   available_months=available_months, 
                                   selected_month=selected_month, 
                                   data=analysis_data)
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f"Lỗi tải trang báo cáo toàn trường: {e}", "error")
        return redirect(url_for('dashboard'))
# ==========================================
# API: LẤY DANH SÁCH THÁNG ĐỘNG CHO APP GVCN
# ==========================================
@app.route('/api/gvcn/get_months', methods=['GET'])
def api_gvcn_get_months():
    if session.get('role') not in ['Giáo viên chủ nhiệm', 'Quản trị viên', 'Admin', 'Bí thư Đoàn trường', 'Bí thư']:
        return {"success": False, "error": "Không có quyền truy cập."}, 401
    
    try:
        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                latest_year = db_session.query(SchoolYear).order_by(SchoolYear.id.desc()).first()
                school_year_id = latest_year.id if latest_year else None
            else:
                school_year_id = active_year.id

            if not school_year_id:
                return {"success": True, "months": []}, 200

            # Lấy danh sách các tháng đã được ghi nhận/tạo trong MonthlyRecord của năm học
            months_query = db_session.query(MonthlyRecord.month_name).filter_by(school_year_id=school_year_id).distinct().all()
            months = [m[0] for m in months_query if m[0]]

            # Nếu chưa có bản ghi nào, trả về danh sách tháng mặc định để phòng hờ
            if not months:
                months = ["Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5"]

            return {"success": True, "months": months}, 200
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

# =====================================================================
# API: XEM BẢNG XẾP HẠNG TOÀN TRƯỜNG (HỖ TRỢ TUẦN, THÁNG, HỌC KỲ KÈM CHI TIẾT)
# =====================================================================
@app.route('/api/gvcn/leaderboard/<path:week_name>', methods=['GET'])
def api_gvcn_leaderboard(week_name):
    if not session.get('role'):
        return {
            "success": False,
            "error": "Phiên đăng nhập không hợp lệ hoặc bạn không có quyền xem Bảng xếp hạng."
        }, 401

    try:
        week_name = (week_name or '').strip()
        if not week_name:
            return {
                "success": False,
                "error": "Thiếu thông tin mốc thời gian."
            }, 400

        with session_scope() as db_session:
            active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
            if not active_year:
                latest_year = db_session.query(SchoolYear).order_by(SchoolYear.id.desc()).first()
                school_year_id = latest_year.id if latest_year else None
            else:
                school_year_id = active_year.id

            if not school_year_id:
                return {
                    "success": False,
                    "error": "Chưa có năm học nào trong hệ thống."
                }, 200

            branches = db_session.query(Branch).filter_by(school_year_id=school_year_id).all()
            group_1_data = []
            group_2_data = []

            # Thiết lập cài đặt điểm tốt tối đa nếu có
            max_tot, max_mon = 14, 4
            try:
                settings = db_session.query(ScoreSettings).filter_by(school_year_id=school_year_id).first()
                if settings:
                    if hasattr(settings, 'max_diem_tot'): max_tot = int(settings.max_diem_tot)
                    if hasattr(settings, 'max_diem_mon'): max_mon = int(settings.max_diem_mon)
            except Exception: pass

            branch_ids = [b.id for b in branches]

            if "Tháng" in week_name:
                # --- XỬ LÝ DỮ LIỆU THÁNG ---
                month_scores = db_session.query(MonthlyRecord).filter(
                    MonthlyRecord.month_name == week_name, 
                    MonthlyRecord.school_year_id == school_year_id
                ).all()
                weekly_scores_all = db_session.query(WeeklyScore).filter(WeeklyScore.branch_id.in_(branch_ids)).all()

                for branch in branches:
                    m_sc = next((m for m in month_scores if m.branch_id == branch.id), None)
                    grp = str(branch.group or "").strip()
                    
                    tong_diem_tru = 0
                    tong_diem_cong = 0
                    tong_diem_tot = 0
                    ghi_chu_gop = []

                    if m_sc and getattr(m_sc, 'weeks_used', None):
                        weeks = [w.strip() for w in m_sc.weeks_used.split(",")]
                        branch_weeks = [w for w in weekly_scores_all if w.branch_id == branch.id and w.week in weeks]
                        for bw in branch_weeks:
                            tong_diem_tru += float(bw.score_tru or bw.score_kem or 0)
                            tong_diem_cong += float(bw.score_cong or 0)
                            if bw.note and bw.note.strip() and bw.note.strip() != 'Không có vi phạm':
                                ghi_chu_gop.append(f"[{bw.week}] {bw.note.strip()}")
                        for w in weeks:
                            tong_diem_tot += calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot)

                    item = {
                        "branch_name": str(branch.name or ""),
                        "group": grp,
                        "week_rating": f"Tổng hợp {week_name}",
                        "total_score": round(float(m_sc.total_score or 0), 1) if m_sc else 0.0,
                        "note": " | ".join(ghi_chu_gop) if ghi_chu_gop else "Không có vi phạm",
                        "cong": round(tong_diem_cong, 1),
                        "tru": round(tong_diem_tru, 1),
                        "diem_tot": int(tong_diem_tot)
                    }

                    if "1" in grp: group_1_data.append(item)
                    else: group_2_data.append(item)

            elif "Học kỳ" in week_name or "Cả năm" in week_name:
                # --- XỬ LÝ DỮ LIỆU HỌC KỲ / NĂM HỌC ---
                if "Học kỳ 1" in week_name: target_weeks = [f"Tuần {i}" for i in range(1, 19)]
                elif "Học kỳ 2" in week_name: target_weeks = [f"Tuần {i}" for i in range(19, 38)]
                else: target_weeks = [f"Tuần {i}" for i in range(1, 38)]

                scores = db_session.query(WeeklyScore).filter(
                    WeeklyScore.week.in_(target_weeks), 
                    WeeklyScore.branch_id.in_(branch_ids)
                ).all() if target_weeks else []

                for branch in branches:
                    branch_scores = [sc for sc in scores if sc.branch_id == branch.id]
                    tong_diem = sum((sc.total_score or 0) for sc in branch_scores)
                    tong_diem_tru = sum((sc.score_tru or sc.score_kem or 0) for sc in branch_scores)
                    tong_diem_cong = sum((sc.score_cong or 0) for sc in branch_scores)
                    tong_diem_tot = sum(calculate_trimmed_good_points_web(db_session, w, branch.name, branch.group, max_mon, max_tot) for w in target_weeks)
                    
                    ghi_chu_gop = [f"[{sc.week}] {sc.note.strip()}" for sc in branch_scores if sc.note and sc.note.strip() and sc.note.strip() != 'Không có vi phạm']
                    grp = str(branch.group or "").strip()

                    item = {
                        "branch_name": str(branch.name or ""),
                        "group": grp,
                        "week_rating": f"Tổng hợp {week_name}",
                        "total_score": round(float(tong_diem), 1),
                        "note": " | ".join(ghi_chu_gop) if ghi_chu_gop else "Không có vi phạm",
                        "cong": round(tong_diem_cong, 1),
                        "tru": round(tong_diem_tru, 1),
                        "diem_tot": int(tong_diem_tot)
                    }

                    if "1" in grp: group_1_data.append(item)
                    else: group_2_data.append(item)

            else:
                # --- XỬ LÝ DỮ LIỆU TUẦN (MẶC ĐỊNH) ---
                scores = db_session.query(WeeklyScore).filter(
                    WeeklyScore.week == week_name,
                    WeeklyScore.branch_id.in_(branch_ids)
                ).all()

                score_map = {sc.branch_id: sc for sc in scores}
                for branch in branches:
                    sc = score_map.get(branch.id)
                    grp = str(branch.group or "").strip()
                    diem_tot = getattr(sc, 'diem_tot', (sc.count_8 or 0) + (sc.count_9 or 0) + (sc.count_10 or 0)) if sc else 0
                    
                    item = {
                        "branch_name": str(branch.name or ""),
                        "group": grp,
                        "week_rating": str(sc.week_rating or "Bình thường") if sc else "Chưa có dữ liệu",
                        "total_score": round(float(sc.total_score or 0), 1) if sc else 0.0,
                        "note": str(sc.note or "Không có vi phạm") if sc else "Chưa nhập điểm",
                        "cong": float(sc.score_cong or 0) if sc else 0.0,
                        "tru": float(sc.score_tru or sc.score_kem or 0) if sc else 0.0,
                        "diem_tot": int(diem_tot)
                    }

                    if "1" in grp: group_1_data.append(item)
                    else: group_2_data.append(item)

            group_1_data.sort(key=lambda x: x["total_score"], reverse=True)
            group_2_data.sort(key=lambda x: x["total_score"], reverse=True)

            return {
                "success": True,
                "week_name": week_name,
                "group_1": group_1_data,
                "group_2": group_2_data
            }, 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": f"{type(e).__name__}: {str(e)}"
        }, 500
# =====================================================================
# API: CẬP NHẬT THÔNG TIN LIÊN HỆ LỚP (ĐÃ VÁ LỖI JSONIFY VÀ BỌC BẢO VỆ KÉP)
# =====================================================================
@app.route('/update_gvcn_info', methods=['POST'])
def update_branch_info():
    if not session.get('role'):
        return {"success": False, "error": "Vui lòng đăng nhập!"}, 401
        
    try:
        data = request.get_json()
        branch_id = data.get('branch_id')
        
        with session_scope() as db_session:
            branch = db_session.query(Branch).filter_by(id=branch_id).first()
            if not branch:
                return {"success": False, "error": "Không tìm thấy Chi đoàn!"}, 404
                
            # Kiểm tra bảo mật nội bộ
            if session.get('role') == 'Giáo viên chủ nhiệm':
                if session.get('username') != branch.name:
                    return {"success": False, "error": "Bạn chỉ có quyền sửa thông tin lớp chủ nhiệm của mình!"}, 403
                    
            # Cập nhật thông tin
            branch.gvcn = data.get('gvcn', '').strip()
            branch.phone_gvcn = data.get('phone_gvcn', '').strip()
            branch.class_monitor = data.get('class_monitor', '').strip()
            branch.phone_monitor = data.get('phone_monitor', '').strip()
            
            # Bọc bảo vệ hàm log để chắc chắn không gây sập nếu chưa được định nghĩa
            try:
                log_system_action("CẬP NHẬT HỒ SƠ", f"Cập nhật thông tin liên hệ lớp {branch.name}")
            except Exception:
                pass
                
            return {"success": True, "message": "Đã cập nhật thông tin thành công!"}, 200
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Lỗi hệ thống: {str(e)}"}, 500
@app.route('/bgh/dashboard', methods=['GET'])
def bgh_dashboard():
    if session.get('role') not in ['Quản trị viên', 'Admin', 'Ban Giám hiệu', 'Bí thư Đoàn trường', 'Bí thư']:
        flash("Bạn không có quyền truy cập khu vực điều hành của Ban Giám hiệu!", "error")
        return redirect(url_for('dashboard'))

    selected_week = request.args.get('week', 'Tuần 1')

    with session_scope() as db_session:
        active_year = db_session.query(SchoolYear).filter_by(is_active=True).first()
        school_year_id = active_year.id if active_year else None

        if not school_year_id:
            return render_template('bgh_dashboard.html', groups_data={}, ai_summary="Chưa kích hoạt năm học.", weeks=[f"Tuần {i}" for i in range(1, 38)])

        branches = db_session.query(Branch).filter_by(school_year_id=school_year_id).all()
        branch_ids = [b.id for b in branches]

        scores = db_session.query(WeeklyScore).filter(
            WeeklyScore.week == selected_week,
            WeeklyScore.branch_id.in_(branch_ids)
        ).all()
        score_map = {sc.branch_id: sc for sc in scores}

        # Gom nhóm dữ liệu theo tên Nhóm (Ví dụ: Nhóm 1, Nhóm 2 hoặc theo Khối)
        groups_dict = {}

        for branch in branches:
            group_name = branch.group if branch.group else "Nhóm Chung"
            if group_name not in groups_dict:
                groups_dict[group_name] = []

            sc = score_map.get(branch.id)
            score = float(sc.total_score or 0) if sc else 0.0
            tru = float(sc.score_tru or sc.score_kem or 0) if sc else 0.0

            groups_dict[group_name].append({
                "branch_name": branch.name,
                "gvcn": branch.gvcn or "Chưa cập nhật",
                "score_tru": tru,
                "total_score": score
            })

        # Sắp xếp và xếp hạng độc lập cho từng nhóm
        total_school_score = 0
        total_classes_count = 0
        total_violations = 0
        top_classes = []
        warning_classes = []

        for group_name in groups_dict:
            # Sắp xếp điểm giảm dần trong từng nhóm riêng biệt
            groups_dict[group_name].sort(key=lambda x: x["total_score"], reverse=True)
            
            # Gán hạng (Rank) riêng cho từng lớp trong nhóm
            for idx, c in enumerate(groups_dict[group_name], start=1):
                c["rank"] = idx
                total_school_score += c["total_score"]
                total_violations += c["score_tru"]
                total_classes_count += 1

            # Lấy top đầu và cảnh báo của nhóm
            if groups_dict[group_name]:
                top_classes.append({"group": group_name, "class": groups_dict[group_name][0]})
                valid_low = [c for c in groups_dict[group_name] if c['total_score'] > 0]
                if valid_low:
                    warning_classes.append({"group": group_name, "class": valid_low[-1]})

        avg_school_score = round(total_school_score / total_classes_count, 1) if total_classes_count > 0 else 0

        # --- [NÂNG CẤP MỚI]: TRỢ LÝ AI PHÂN TÍCH CHUYÊN SÂU & THỰC CHIẾN ---
        all_flattened_classes = []
        for g_name, cls_list in groups_dict.items():
            for c in cls_list:
                c['group_name'] = g_name
                all_flattened_classes.append(c)
        
        # Sắp xếp theo số điểm trừ (score_tru) giảm dần để tìm điểm nóng vi phạm
        all_flattened_classes.sort(key=lambda x: x["score_tru"], reverse=True)
        top_violated_classes = [c for c in all_flattened_classes if c["score_tru"] > 0]

        if total_classes_count == 0:
            ai_summary = f"Trong {selected_week}, chưa có dữ liệu điểm số được ghi nhận trên hệ thống để phân tích."
        else:
            insights_sentences = []
            insights_sentences.append(f"📊 **Bức tranh tổng thể {selected_week}:** Điểm trung bình toàn trường đạt **{avg_school_score} điểm** với tổng mức điểm trừ vi phạm ghi nhận là **-{total_violations}đ**.")
            
            if top_violated_classes:
                worst = top_violated_classes[0]
                insights_sentences.append(f"⚠️ **Điểm nóng cần lưu ý:** Lớp **{worst['branch_name']}** ({worst['group_name']}) đang dẫn đầu danh sách vi phạm với mức trừ **-{worst['score_tru']}đ** (GVCN: {worst['gvcn']}).")
            else:
                insights_sentences.append(f"✨ **Tín hiệu tích cực:** Nề nếp toàn trường trong tuần này rất ổn định, không ghi nhận các trường hợp vi phạm nặng.")

            if avg_school_score >= 90:
                insights_sentences.append(f"💡 **Đánh giá của Trợ lý AI:** Nề nếp học sinh duy trì tốt, đề nghị Đoàn trường tiếp tục phát huy và tuyên dương các lớp giữ vững phong độ.")
            else:
                insights_sentences.append(f"💡 **Đề xuất hành động cho BGH:** Đề nghị Đoàn trường phối hợp với Ban giám thị và GVCN các lớp tốp dưới tăng cường nhắc nhở tác phong, đồng phục và giờ giấc trong đầu tuần tới.")

            ai_summary = "<br>".join(insights_sentences)

        return render_template(
            'bgh_dashboard.html',
            selected_week=selected_week,
            avg_school_score=avg_school_score,
            total_violations=total_violations,
            groups_data=groups_dict,
            top_classes=top_classes,
            warning_classes=warning_classes,
            ai_summary=ai_summary,
            weeks=[f"Tuần {i}" for i in range(1, 38)]
        )
if __name__ == "__main__":
    auto_init_accounts()
    init_db()
    create_mock_admin()
    
    import threading
    threading.Thread(target=background_auto_backup, daemon=True).start()
    
    print("=========================================================")
    print("🌟 BẢN CẬP NHẬT HOÀN HẢO 🌟")
    print("=========================================================")
    print("🚀 Máy chủ Web đang chạy. Hãy mở trình duyệt và truy cập http://127.0.0.1:8080")
    
    # [TINH CHỈNH NHỎ]: Tự động nhận diện Port từ Render hoặc mặc định là 8080
    import os
    port = int(os.environ.get("PORT", 8080))
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=port)